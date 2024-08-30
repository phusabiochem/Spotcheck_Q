#!/usr/bin/python3

from tkinter import *
import tkinter.font as font
from tkinter import messagebox
from tkinter import filedialog
from functools import partial
from tkinter import ttk
from time import sleep, gmtime, strftime
from datetime import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image as Img
from picamera import PiCamera
import cv2
import numpy as np
from matplotlib import pyplot
from PIL import ImageTk, Image
import serial
from fractions import Fraction
import os
import shutil
import awesometkinter as atk
import math
from enum import Enum
from ftplib import FTP
# ~ import Pmw
import subprocess

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders
import re
import dns.resolver
import socket


APP_BGD_COLOR = "white smoke"
# MAIN MENU DEFINE
MAIN_MENU_BUTTON_BGD_COLOR = "grey80"
MAIN_MENU_BUTTON_TXT_COLOR = "black"
MAIN_MENU_BUTTON_ACTIVE_BGD_COLOR = "lawn green"
MAIN_MENU_BUTTON_FRAME_BGD_COLOR = "white smoke"
MAIN_MENU_BUTTON_WIDTH = 15
MAIN_MENU_BUTTON_HEIGHT = 4
MAIN_FUNCTION_FRAME_BGD_COLOR = "white"
MAIN_TITLE_FRAME_BGD_COLOR = "dodger blue"
MAIN_BUTTON_FRAME_BGD_COLOR = "dodger blue"

LABEL_FRAME_TXT_COLOR = "black"
LABEL_FRAME_BGD_COLOR = "white"
LABEL_TXT_COLOR = "black"
LABEL_BGD_COLOR = "white"

MAIN_FUNCTION_BUTTON_WIDTH = 12
MAIN_FUNCTION_BUTTON_HEIGHT = 3
MAIN_FUNCTION_BUTTON_BGD_COLOR = "grey80"
MAIN_FUNCTION_BUTTON_TXT_COLOR = "black"

TITILE_FRAME_BGD_COLOR = "dodger blue"
TITILE_FRAME_TXT_COLOR = "white"

SWITCH_PAGE_BUTTON_WIDTH = 10
SWITCH_PAGE_BUTTON_HEIGHT = 2
SWITCH_PAGE_BUTTON_BGD_COLOR = "grey80"
SWITCH_PAGE_BUTTON_TXT_COLOR = "black"

CONFIRM_BUTTON_WIDTH = 10
CONFIRM_BUTTON_HEIGHT = 2
CONFIRM_BUTTON_BGD_COLOR = "grey80"
CONFIRM_BUTTON_TXT_COLOR = "black"

SAMPLE_BUTTON_WIDTH = 2
SAMPLE_BUTTON_HEIGHT = 2
SAMPLE_BUTTON_FRAME_BDG_COLOR = "grey96"
SAMPLE_BUTTON_BGD_COLOR = "lavender"
SAMPLE_BUTTON_TXT_COLOR = "black"
SAMPLE_BUTTON_ACTIVE_BGD_COLOR = "white"
SAMPLE_BUTTON_CHOOSE_BGD_COLOR = "turquoise2"
SAMPLE_BUTTON_DONE_BGD_COLOR = "lawn green"
SAMPLE_BUTTON_TMP_BGD_COLOR = "grey99"

RESULT_TABLE_FRAME_BGD_COLOR = "grey96"
RESULT_LABEL_BGD_COLOR = "lawn green"
RESULT_LABEL_ERROR_BGD_COLOR = "firebrick2"
RESULT_LABEL_TXT_COLOR = "black"

PROGRAM_BUTTON_BGD_COLOR = "cyan2"
PROGRAM_BUTTON_TXT_COLOR = "black"
PROGRAM_BUTTON_ACTIVE_BGD_COLOR = "lawn green"

LOGIN_BUTTON_BGD_COLOR = "grey85"
LOGIN_BUTTON_TXT_COLOR = "black"

NA_COLOR = "grey82"
ERROR_COLOR = "DodgerBLue2"
NEGATIVE_COLOR = "green3"
POSITIVE_COLOR = "red"
LOW_COPY_COLOR = "pink"

MAIN_MENU_BUTTON_FONT = ('Helvetica', 10, 'bold')
LABELFRAME_TXT_FONT = ('Helvetica', 12)
MAIN_FUCNTION_BUTTON_FONT = ('Helvetica', 10)
TITLE_TXT_FONT = ('Helvetica', 10, 'bold')
SWITCH_PAGE_BUTTON_FONT = ('Helvetica', 10)
LABEL_FRAME_TXT_FONT = ('Helvetica', 10)
LABEL_TXT_FONT = ('Helvetica', 10)
ENTRY_TXT_FONT = ('Helvetica', 11)
CONFIRM_BUTTON_TXT_FONT = ('Helvetica', 10)
SAMPLE_BUTTON_TXT_FONT = ('Helvetica', 10)
SAMPLE_LABEL_TXT_FONT = ('Helvetica', 13)
RESULT_LABEL_TXT_FONT = ('Helvetica', 10)
PROGRAM_BUTTON_TXT_FONT = ('Helvetica', 10)
LOGIN_LABEL_TXT_FONT = ('Helvetica', 15)
LOGIN_BUTTON_TXT_FONT = ('Helvetica', 10)

CAM_FRAMERATE_NUMERATOR = 1
CAM_FRAMERATE_DENOMINATOR = 6
CAM_SENSOR_MODE = 3
CAM_ROTATION = 180
CAM_ISO = 200
CAM_SLEEP_TIME = 2
CAM_SHUTTER_SPEED = 6000000
CAM_EXPOSURE_MODE = "off"

##### GLOBAL VARIABLE - START #####
# ~ NUM_1 = 1.08
# ~ NUM_2 = 1.15

serial_signal = 0
analysis_mode = 0

# ERROR LIST
class ERROR_LIST(Enum):
    CAMERA_ERROR = 1
    SERIAL_TIMEOUT_ERROR = 2
    SYSTEM_ERROR_1 = 3
    SYSTEM_ERROR_2 = 4

# SERIAL
ser = serial.Serial(
    port = '/dev/serial0',
    baudrate = 115200,
    parity = serial.PARITY_NONE,
    stopbits = serial.STOPBITS_ONE,
    bytesize = serial.EIGHTBITS,
    timeout = 1
)

##### GLOBAL VARIABLE - END #####

##### GLOBAL METHOD - START #####
def camera_capture(output):
    camera = PiCamera()
    camera.framerate = Fraction(CAM_FRAMERATE_NUMERATOR, CAM_FRAMERATE_DENOMINATOR)
    camera.sensor_mode = CAM_SENSOR_MODE
    camera.rotation = CAM_ROTATION
    camera.iso = CAM_ISO
    sleep(CAM_SLEEP_TIME)
    camera.shutter_speed = CAM_SHUTTER_SPEED
    camera.exposure_mode = CAM_EXPOSURE_MODE
    camera.capture(output)
    camera.close()
##### GLOBAL METHOD - END #####

##### CHECK FOLDER EXISTS - START #####
# ~ BUILD = 1
# ~ if(BUILD==1):
    # ~ dist_dir = os.path.abspath(os.getcwd())
    # ~ spotcheck_dist_dir = os.path.dirname(dist_dir)
    # ~ working_dir = os.path.dirname(spotcheck_dist_dir)
    # ~ parent_dir = os.path.dirname(working_dir)
# ~ else:
    # ~ working_dir = os.path.abspath(os.getcwd())
    # ~ parent_dir = os.path.dirname(working_dir)

working_dir = '/home/pi/Spotcheck'
parent_dir = '/home/pi'

print("working_dir: ", working_dir)
print("parent_dir: ", parent_dir)
##### In Spotcheck folder - Start #####
# Programs path
if not os.path.exists(working_dir + "/Programs"):
    f = os.path.join(working_dir + '/', "Programs")
    os.mkdir(f)
programs_path = working_dir + "/Programs/"

# Programs qualitaitve
if not os.path.exists(programs_path + "Screening"):
    f = os.path.join(programs_path, "Screening")
    os.mkdir(f)
programs_qualitative_path = programs_path + "Screening/"

# Programs quantitaitve
if not os.path.exists(programs_path + "/Quantitative"):
    f = os.path.join(programs_path + '/', "Quantitative")
    os.mkdir(f)
programs_quantitative_path = programs_path + "Quantitative/"
##### In Spotcheck folder - End #####

##### At Desktop - Start #####
# Spotcheck result path
if not os.path.exists(parent_dir + "/Desktop/Spotcheck_Results"):
    f = os.path.join(parent_dir + "/Desktop/", "Spotcheck_Results")
    os.mkdir(f)
results_path = parent_dir + "/Desktop/Spotcheck_Results/"

# Spotcheck results qualitative path
if not os.path.exists(results_path + "Screening"):
    f = os.path.join(results_path, "Screening")
    os.mkdir(f)
results_qualitative_path = results_path + "Screening/"

# Spotcheck results quantitative path
if not os.path.exists(results_path + "Quantitative"):
    f = os.path.join(results_path, "Quantitative")
    os.mkdir(f)
results_quantitative_path = results_path + "Quantitative/"

# Programs results path
if not os.path.exists(parent_dir + "/Desktop/Spotcheck_Programs"):
    f = os.path.join(parent_dir + "/Desktop/", "Spotcheck_Programs")
    os.mkdir(f)
results_programs_path = parent_dir + "/Desktop/Spotcheck_Programs/"

# Programe results qualitative path
if not os.path.exists(results_programs_path + "Screening"):
    f = os.path.join(results_programs_path, "Screening")
    os.mkdir(f)
results_programs_qualitative_path = results_programs_path + "Screening/"

# Programe results quantitative path
if not os.path.exists(results_programs_path + "Quantitative"):
    f = os.path.join(results_programs_path, "Quantitative")
    os.mkdir(f)
results_programs_quantitative_path = results_programs_path + "Quantitative/"


# Spotcheck ID path
if not os.path.exists(parent_dir + "/Desktop/Spotcheck_ID"):
    f = os.path.join(parent_dir + "/Desktop/", "Spotcheck_ID")
    os.mkdir(f)
id_path = parent_dir + "/Desktop/Spotcheck_ID/"

# ~ # Spotchek ID qualitative path
# ~ if not os.path.exists(id_path + "Qualitative"):
    # ~ f = os.path.join(id_path, "Qualitative")
    # ~ os.mkdir(f)
# ~ id_qualitative_path = id_path + "Qualitative/"

# ~ # Spotchek ID quantitative path
# ~ if not os.path.exists(id_path + "Quantitative"):
    # ~ f = os.path.join(id_path, "Quantitative")
    # ~ os.mkdir(f)
# ~ id_quantitative_path = id_path + "Quantitative/"


# Spotcheck ID old path
if not os.path.exists(parent_dir + "/Desktop/Spotcheck_ID/Spotcheck_ID_Old"):
    f = os.path.join(parent_dir + "/Desktop/Spotcheck_ID/", "Spotcheck_ID_Old")
    os.mkdir(f)
id_old_path = parent_dir + "/Desktop/Spotcheck_ID/Spotcheck_ID_Old/"
##### At Desktop - End #####
##### CHECK FOLDER EXISTS - END #####

##### COEFFICIENT HANDLE - START #####
if not os.path.exists(working_dir + "/coefficient.xlsx"):
    wb = Workbook()
    sheet = wb.active

    for i in range(0,48):
        if(i<6):
            pos1 = str(chr(65+i+1)) + "2"
            pos2 = str(chr(65+i+1)) + "11"
        if(i>=6 and i<12):
            pos1 = str(chr(65+i-5)) + "3"
            pos2 = str(chr(65+i-5)) + "12"
        if(i>=12 and i<18):
            pos1 = str(chr(65+i-11)) + "4"
            pos2 = str(chr(65+i-11)) + "13"
        if(i>=18 and i<24):
            pos1 = str(chr(65+i-17)) + "5"
            pos2 = str(chr(65+i-17)) + "14"
        if(i>=24 and i<30):
            pos1 = str(chr(65+i-23)) + "6"
            pos2 = str(chr(65+i-23)) + "15"
        if(i>=30 and i<36):
            pos1 = str(chr(65+i-29)) + "7"
            pos2 = str(chr(65+i-29)) + "16"
        if(i>=36 and i<42):
            pos1 = str(chr(65+i-35)) + "8"
            pos2 = str(chr(65+i-35)) + "17"
        if(i>=42):
            pos1 = str(chr(65+i-41)) + "9"
            pos2 = str(chr(65+i-41)) + "18"

        sheet[pos1] = 1
        sheet[pos2] = 40
    wb.save(working_dir + "/coefficient.xlsx")
    wb.close()

coefficient = list(range(48))
base_intensity = list(range(48))
wb = load_workbook(working_dir + "/coefficient.xlsx")
sheet = wb.active
for i in range(0,48):
    if(i<6):
        pos = str(chr(65+i+1)) + "2"
        pos1 = str(chr(65+i+1)) + "11"
    if(i>=6 and i<12):
        pos = str(chr(65+i-5)) + "3"
        pos1 = str(chr(65+i-5)) + "12"
    if(i>=12 and i<18):
        pos = str(chr(65+i-11)) + "4"
        pos1 = str(chr(65+i-11)) + "13"
    if(i>=18 and i<24):
        pos = str(chr(65+i-17)) + "5"
        pos1 = str(chr(65+i-17)) + "14"
    if(i>=24 and i<30):
        pos = str(chr(65+i-23)) + "6"
        pos1 = str(chr(65+i-23)) + "15"
    if(i>=30 and i<36):
        pos = str(chr(65+i-29)) + "7"
        pos1 = str(chr(65+i-29)) + "16"
    if(i>=36 and i<42):
        pos = str(chr(65+i-35)) + "8"
        pos1 = str(chr(65+i-35)) + "17"
    if(i>=42):
        pos = str(chr(65+i-41)) + "9"
        pos1 = str(chr(65+i-41)) + "18"

    coefficient[i] = float(sheet[pos].value)
    base_intensity[i] = float(sheet[pos1].value)

    wb.close()
##### COEFFICIENT HANDLE - END #####

##### COORDINATES HANDLE - START #####
if not os.path.exists(working_dir + "/coordinates.txt"):
    fw = open(working_dir + "/coordinates.txt",'w')
    fw.writelines(["0\n","0\n", "1\n","1\n"])
    fw.close()
fr = open(working_dir + "/coordinates.txt","r")
x1 = int(fr.readline())
y1 = int(fr.readline())
x2 = int(fr.readline())
y2 = int(fr.readline())

if not os.path.exists(working_dir + "/.coordinates.txt"):
    fw = open(working_dir + "/.coordinates.txt",'w')
    fw.close()
##### COORDINATES HANDLE - END #####

##### CONFIG PARAMETER HANDLE - START #####
if not os.path.exists(working_dir + "/config.txt"):
    fw = open(working_dir + "/config.txt",'w')
    fw.writelines(["1\n","0\n"])
    fw.close()
fr = open(working_dir + "/config.txt","r")
a = float(fr.readline())
b = float(fr.readline())
##### CONFIG PARAMETER HANDLE - END #####

##### MULTIPLIER HANDLE 1 - START #####
if not os.path.exists(working_dir + "/multiplier.txt"):
    fw = open(working_dir + "/multiplier.txt",'w')
    fw.writelines(["1\n","1\n"])
    fw.close()
fr = open(working_dir + "/multiplier.txt","r")
NUM_1 = float(fr.readline())
NUM_2 = float(fr.readline())

print('----- multiplier -----')
print('NUM_1: ', NUM_1)
print('NUM_2: ', NUM_2)
##### MULTIPLIER HANDLE 1 - END #####

##### MULTIPLIER HANDLE 2 - START #####
if not os.path.exists(working_dir + "/multiplier2.txt"):
    fw = open(working_dir + "/multiplier2.txt",'w')
    fw.writelines(["1\n","1\n","1\n"])
    fw.close()
fr = open(working_dir + "/multiplier2.txt","r")
NUM2_1 = float(fr.readline())
NUM2_2 = float(fr.readline())
NUM2_3 = float(fr.readline())

print('----- multiplier2 -----')
print('NUM2_1: ', NUM2_1)
print('NUM2_2: ', NUM2_2)
print('NUM2_3: ', NUM2_3)
##### MULTIPLIER HANDLE 2 - END #####


##### EMAIL HANDLE - START #####
if not os.path.exists(working_dir + "/.email.txt"):
    fw = open(working_dir + "/.email.txt",'w')
    fw.writelines(["0\n"])
    fw.close()
##### EMAIL HANDLE - END #####

##### SERVER HANDLE - START #####
if not os.path.exists(working_dir + "/.server.txt"):
    fw = open(working_dir + "/.server.txt",'w')
    fw.writelines(["0\n"])
    fw.close()    
else:
    fr = open(working_dir + "/.server.txt","r")
    server_active_0 = int(fr.readline().strip())
    fr.close()
    
##### SERVER HANDLE - END #####


class ScrollableFrame(Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        canvas = Canvas(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scrollable_frame = Frame(canvas, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=self.scrollable_frame)

        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", expand=TRUE)
        scrollbar.pack(side="right", fill="y")

class ScrollableFrame1(Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        canvas = Canvas(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR, height=370)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scrollable_frame = Frame(canvas, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=self.scrollable_frame)

        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", expand=TRUE)
        scrollbar.pack(side="right", fill="y")


class ScrollableFrame2(Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        canvas = Canvas(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR, height=370, width=519)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scrollable_frame = Frame(canvas, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=self.scrollable_frame)

        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", expand=TRUE)
        scrollbar.pack(side="right", fill="y")



# EMAIL - START
class AutoMail():
    def __init__(self, sender, password, recipient, subject, content, zip_file, zip_file_name):
        self.SMTP_SERVER = 'smtp.gmail.com'
        self.SMTP_PORT = 587
        self.sender = sender
        self.password = password
        self.recipient = recipient
        self.subject = subject
        self.content = content
        self.zip_file = zip_file
        self.zip_file_name = zip_file_name

    def send(self):
        emailData = MIMEMultipart()
        emailData['Subject'] = self.subject
        emailData['To'] = self.recipient
        emailData['From'] = self.sender

        emailData.attach(MIMEText(self.content))

    #     imageData = MIMEImage(open(image, 'rb').read(), 'jpg')
    #     imageData.add_header('Content-Disposition', 'attachment; filename="image.jpg"')
    #     emailData.attach(imageData)

        with open(self.zip_file,'rb') as file:
            emailData.attach(MIMEApplication(file.read(), Name= self.zip_file_name + '.zip'))

        session = smtplib.SMTP(self.SMTP_SERVER, self.SMTP_PORT)
        session.ehlo()
        session.starttls()
        session.ehlo()

        #session.login(mail_address, password)
        session.login(self.sender, self.password)

        session.sendmail(self.sender, self.recipient.split(','), emailData.as_string())
        session.quit
# EMAIL - END

class Process_Image():
    def __init__(self, image_path, start_point=(x1,y1), end_point=(x2,y2)):
        self.image = cv2.imread(image_path)
        self.start_point = start_point
        self.end_point = end_point

    def process(self, coefficient, mode=0, well_list=[]):
        print("coefficient: ", coefficient)
        blur_img = cv2.GaussianBlur(self.image.copy(), (35,35), 0)
        gray_img = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY)

        thresh, binary_img = cv2.threshold(gray_img.copy(), 30, maxval=255, type=cv2.THRESH_BINARY)
        contours, hierarchy = cv2.findContours(binary_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        print("Number of contours: " + str(len(contours)))

        contours = sorted(contours, key=lambda data:self.sorting_xy(data))

        if(mode==0):
            try:
                bounding_rect_0 = cv2.boundingRect(contours[0])
                bounding_rect_47 = cv2.boundingRect(contours[len(contours)-1])
                new_start_point = (bounding_rect_0[0]-12, bounding_rect_0[1]-12)
                new_end_point = (bounding_rect_47[0]+bounding_rect_47[2]+12, bounding_rect_47[1]+bounding_rect_47[3]+12)
                print('New start point:', new_start_point)
                print('New end point:', new_end_point)
                fw= open(working_dir + '/.coordinates.txt','w')
                fw.writelines("Start Point: " + str(new_start_point) + "\n")
                fw.writelines("End Point: " + str(new_end_point))
            except:
                print("Can't find new coordinates")
                pass

        contour_img = np.zeros_like(gray_img)
        contour_img = cv2.rectangle(contour_img, self.start_point, self.end_point, (255,255,255), -1)
        rect_w = self.end_point[0] - self.start_point[0]
        rect_h = self.end_point[1] - self.start_point[1]
        cell_w = round(rect_w/6)
        cell_h = round(rect_h/8)
        for i in range(1,6):
            contour_img = cv2.line(contour_img, (self.start_point[0] + i*cell_w, self.start_point[1]), (self.start_point[0] + i*cell_w, self.end_point[1]),(0,0,0), 4)
        for i in range(1,8):
            contour_img = cv2.line(contour_img, (self.start_point[0], self.start_point[1] + i*cell_h), (self.end_point[0], self.start_point[1] + i*cell_h),(0,0,0), 4)

        thresh1 , binary1_img = cv2.threshold(contour_img, 250, maxval=255, type=cv2.THRESH_BINARY)
        contours1, hierarchy1 = cv2.findContours(binary1_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)

        contours1 = sorted(contours1, key=lambda data:self.sorting_y(data))
        contours1_h1 = contours1[0:6]
        contours1_h2 = contours1[6:12]
        contours1_h3 = contours1[12:18]
        contours1_h4 = contours1[18:24]
        contours1_h5 = contours1[24:30]
        contours1_h6 = contours1[30:36]
        contours1_h7 = contours1[36:42]
        contours1_h8 = contours1[42:48]
        contours1_h1 = sorted(contours1_h1, key=lambda data:self.sorting_x(data))
        contours1_h2 = sorted(contours1_h2, key=lambda data:self.sorting_x(data))
        contours1_h3 = sorted(contours1_h3, key=lambda data:self.sorting_x(data))
        contours1_h4 = sorted(contours1_h4, key=lambda data:self.sorting_x(data))
        contours1_h5 = sorted(contours1_h5, key=lambda data:self.sorting_x(data))
        contours1_h6 = sorted(contours1_h6, key=lambda data:self.sorting_x(data))
        contours1_h7 = sorted(contours1_h7, key=lambda data:self.sorting_x(data))
        contours1_h8 = sorted(contours1_h8, key=lambda data:self.sorting_x(data))

        sorted_contours1 = contours1_h1 + contours1_h2 + contours1_h3 + contours1_h4 + contours1_h5 + contours1_h6 + contours1_h7 + contours1_h8

        list_intensities = []
        sum_intensities = []
        result_list = list(range(48))
        area = list(range(48))

        blur1_img = cv2.GaussianBlur(self.image.copy(), (25,25), 0)
        tmp_list = list(range(48))
        list_bgrvalue = []
        list_index = list(range(48))
        for i in range(len(sorted_contours1)):
            list_index[i] = []
            cimg = np.zeros_like(gray_img)
            cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
            pts = np.where(cimg == 255)
            list_bgrvalue.append(blur1_img[pts[0], pts[1]])
            for j in range(len(list_bgrvalue[i])):
                 list_index[i].append(round((list_bgrvalue[i][j][1]*3 + list_bgrvalue[i][j][2])))
            list_index[i].sort()
            list_intensities.append(sum(list_index[i][len(list_index[i])-250:]))
            area[i]= cv2.contourArea(sorted_contours1[i])
            tmp_list[i] = list_intensities[i]/1000
            result_list[i] = round(tmp_list[i],1)

        for i in range(1,7):
            result_list[6*i+1]=round(result_list[6*i+1]*(1-(0.02*round(result_list[6*i]/70) + 0.02*round(result_list[6*i+2]/70) + 0.02*round(result_list[6*i-5]/70) + 0.02*round(result_list[6*i+7]/70) + 0.003*round(result_list[6*i-6]/70) + 0.003*round(result_list[6*i-4]/70) + 0.003*round(result_list[6*i+6]/70) + 0.003*round(result_list[6*i+8]/70)+ 0.006*round(result_list[6*i+3]/70))),1)
            result_list[6*i+2]=round(result_list[6*i+2]*(1-(0.02*round(result_list[6*i+1]/70) + 0.02*round(result_list[6*i+3]/70) + 0.02*round(result_list[6*i-4]/70) + 0.02*round(result_list[6*i+8]/70) + 0.003*round(result_list[6*i-5]/70) + 0.003*round(result_list[6*i-3]/70) + 0.003*round(result_list[6*i+7]/70) + 0.003*round(result_list[6*i+9]/70)+ 0.006*round(result_list[6*i+4]/70)+ 0.006*round(result_list[6*i]/70))),1)
            result_list[6*i+3]=round(result_list[6*i+3]*(1-(0.02*round(result_list[6*i+2]/70) + 0.02*round(result_list[6*i+4]/70) + 0.02*round(result_list[6*i-3]/70) + 0.02*round(result_list[6*i+9]/70) + 0.003*round(result_list[6*i-4]/70) + 0.003*round(result_list[6*i-2]/70) + 0.003*round(result_list[6*i+8]/70) + 0.003*round(result_list[6*i+10]/70)+ 0.006*round(result_list[6*i+5]/70)+ 0.006*round(result_list[6*i+1]/70))),1)
            result_list[6*i+4]=round(result_list[6*i+4]*(1-(0.02*round(result_list[6*i+3]/70) + 0.02*round(result_list[6*i+5]/70) + 0.02*round(result_list[6*i-2]/70) + 0.02*round(result_list[6*i+10]/70) + 0.003*round(result_list[6*i-3]/70) + 0.003*round(result_list[6*i-1]/70) + 0.003*round(result_list[6*i+9]/70) + 0.003*round(result_list[6*i+11]/70)+ 0.006*round(result_list[6*i+2]/70))),1)

            result_list[6*i]=round(result_list[6*i]*(1-(0.02*round(result_list[6*i+1]/70) + 0.015*round(result_list[6*i-6]/70) + 0.015*round(result_list[6*i+6]/70) + 0.003*round(result_list[6*i-5]/70) + 0.003*round(result_list[6*i+7]/70)+0.006*round(result_list[6*i+2]/70))),1)
            result_list[6*i+5]=round(result_list[6*i+5]*(1-(0.02*round(result_list[6*i+4]/70) + 0.015*round(result_list[6*i-1]/70) + 0.015*round(result_list[6*i+11]/70) + 0.003*round(result_list[6*i-2]/70) + 0.003*round(result_list[6*i+10]/70)+0.006*round(result_list[6*i+3]/70))),1)

        for i in range(1,5):
            result_list[i]=round(result_list[i]*(1-(0.02*round(result_list[i-1]/70) + 0.02*round(result_list[i+1]/70) + 0.015*round(result_list[i+6]/70)+ 0.003*round(result_list[i+5]/70) + 0.003*round(result_list[i+7]/70))),1)
            result_list[i+42]=round(result_list[i+42]*(1-(0.02*round(result_list[i+41]/70) + 0.02*round(result_list[i+43]/70) + 0.015*round(result_list[i+36]/70)+ 0.003*round(result_list[i+35]/70) + 0.003*round(result_list[i+37]/70))),1)

        result_list[0]=round(result_list[0]*(1-(0.015*round(result_list[1]/70) + 0.015*round(result_list[6]/70))),1)
        result_list[5]=round(result_list[5]*(1-(0.015*round(result_list[4]/70) + 0.015*round(result_list[11]/70))),1)
        result_list[42]=round(result_list[42]*(1-(0.015*round(result_list[43]/70) + 0.015*round(result_list[36]/70))),1)
        result_list[47]=round(result_list[47]*(1-(0.015*round(result_list[46]/70) + 0.015*round(result_list[41]/70))),1)


        for i in range(len(sorted_contours1)):
            result_list[i] = round(result_list[i]*coefficient[i],1)

        for i in range(len(sorted_contours1)):
            if(result_list[i]>99):
                result_list[i]=99

        for i in range(len(sorted_contours1)):
            if ((i!=0) and ((i+1)%6==0)):
                print('%.1f'%(result_list[i]))
            else:
                print('%.1f'%(result_list[i]), end = ' | ')

        blurori_img = cv2.GaussianBlur(self.image.copy(), (25,25), 0)
        if(mode):
            for i in range(len(sorted_contours1)):
                if(well_list[i]=='N/A'):
                    cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,0), thickness = -1)
                else:
                    cv2.drawContours(blurori_img, sorted_contours1, i, (255,255,0), thickness = 2)
                    # if(well_list[i] < float(thr_set)):
                    #     cv2.drawContours(blurori_img, sorted_contours1, i, (0,255,0), thickness = 2)
                    # else:
                    #     cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,255), thickness = 2)
        else:
            for i in range(len(sorted_contours1)):
                cv2.drawContours(blurori_img, sorted_contours1, i, (255,255,0), thickness = 2)

        return (result_list, blurori_img)


    def sorting_y(self, contour):
        rect_y = cv2.boundingRect(contour)
        return rect_y[1]
    def sorting_x(self, contour):
        rect_x = cv2.boundingRect(contour)
        return rect_x[0]
    def sorting_xy(self, contour):
        rect_xy = cv2.boundingRect(contour)
        return math.sqrt(math.pow(rect_xy[0],2) + math.pow(rect_xy[1],2))


class SystemCheckFrame(Frame):
    def __init__(self, container):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container

        self.err = 0
        self.mode_check = 0 # 0 --> create qualitative program
                            # 1 --> analysis qualitative program
                            # 2 --> create quantitative program
                            # 3 --> analysis quantitative program

        # 3 main frame
        self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=5, fill=X)
        self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X)

        # In title frame
        self.title_label = Label(self.title_frame,
                                text = "SYSTEM CHECK",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

        # In work frame


        # In button frame
        self.back_button = Button(self.button_frame,
                                text = "Back",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.back_clicked)
        self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

        self.next_button = Button(self.button_frame,
                                text = "Next",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.next_clicked)
        self.next_button.pack(ipadx=30, ipady=10, side=RIGHT)


    def back_clicked(self):
        try:
            self.check_result_frame.destroy()
        except:
            pass

        try:
            self.progressbar.pack_forget()
            self.process_label.pack_forget()
        except:
            pass

        self.base_window.forget_page()

        if(self.mode_check == 0):
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_qualitative_1)
        elif(self.mode_check == 1):
            #self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_1)
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
        elif(self.mode_check == 2):
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_quantitative_1)
        elif(self.mode_check == 3):
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_1)
        self.base_window.switch_page()

    def next_clicked(self):
        if(self.err == 2):
            msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SYSTEM_ERROR_2'].value),
                                                ERROR_LIST(ERROR_LIST['SYSTEM_ERROR_2'].value).name,
                                                icon = "error")
        else:
            print("mode_check:", self.mode_check)
            if(self.mode_check == 0):
                self.base_window.forget_page()
                self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_qualitative_2)
                self.base_window.switch_page()
            elif(self.mode_check == 1):
                self.base_window.forget_page()
                self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_2)
                self.base_window.switch_page()
            elif(self.mode_check == 2):
                self.base_window.forget_page()
                self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_quantitative_2)
                self.base_window.switch_page()
            elif(self.mode_check == 3):
                self.base_window.forget_page()
                self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_2)
                self.base_window.switch_page()

    def serial_handle(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("1.Horizontal.TProgressbar", troughcolor ='grey85', background='green3')
        self.progressbar = ttk.Progressbar(self.work_frame,
                                    style="1.Horizontal.TProgressbar",
                                    length = 200,
                                    mode = 'determinate')
        self.progressbar.pack(ipadx=2, ipady=2)

        self.process_label = Label(self.work_frame,
                        bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
                        fg = 'black',
                        text = 'Checking background...',
                        font = LABEL_TXT_FONT)
        self.process_label.pack(ipadx=2, ipady=2, anchor=N)

        ser.flushInput()
        ser.flushOutput()

        self.progressbar['value']=5
        self.base_window.update_idletasks()
        sleep(0.5)

        data_send = 'P'
        print("Data send:", data_send)
        ser.write(data_send.encode())

        receive_data = StringVar()
        count = 0
        bled_ready = 0
        while(receive_data != 'C'):
            if(ser.in_waiting>0):
                receive_data = ser.readline().decode('utf-8').rstrip()
                print("Data received:", receive_data)

                self.progressbar['value']=10
                self.base_window.update_idletasks()
                sleep(0.5)

                if(receive_data == 'C'):
                    self.progressbar['value']=20
                    self.base_window.update_idletasks()
                    sleep(0.5)

                    bled_ready = 1
                    break;
            else:
                sleep(1)
                count += 1
                if(count > 15):
                    msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value),
                                                ERROR_LIST(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value).name,
                                                icon = "error")
                    break;

        if(bled_ready):
            self.progressbar['value']=45
            self.base_window.update_idletasks()
            print("self.mode_check: ", self.mode_check)
            try:
                if(self.mode_check==0):
                    camera_capture(results_programs_qualitative_path + self.base_window.new_qualitative_1.experiment_name + '/system_check_raw.jpg')
                elif(self.mode_check==1):
                    #camera_capture(self.base_window.qualitative_analysis_1.system_check_folder + '/system_check_raw.jpg')
                    camera_capture(self.base_window.qualitative_analysis_0.system_check_folder + '/system_check_raw.jpg')
                elif(self.mode_check==2):
                    camera_capture(results_programs_quantitative_path + self.base_window.new_quantitative_1.experiment_name + '/system_check_raw.jpg')
                elif(self.mode_check==3):
                    camera_capture(self.base_window.quantitative_analysis_1.system_check_folder + '/system_check_raw.jpg')


                self.progressbar['value']=65
                self.base_window.update_idletasks()
            except Exception as e :
                msg = messagebox.showerror("ERR "+ str(ERROR_LIST['CAMERA_ERROR'].value),
                                        ERROR_LIST(ERROR_LIST['CAMERA_ERROR'].value).name,
                                        icon = "error")
                if(msg=='ok'):
                    self.base_window.destroy()

            if(self.mode_check==0):
                self.result, self.image = Process_Image(results_programs_qualitative_path + self.base_window.new_qualitative_1.experiment_name + '/system_check_raw.jpg').process(coefficient)
                cv2.imwrite(results_programs_qualitative_path + self.base_window.new_qualitative_1.experiment_name + '/system_check_process.jpg', self.image)
            elif(self.mode_check==1):
                #self.result, self.image = Process_Image(self.base_window.qualitative_analysis_1.system_check_folder + '/system_check_raw.jpg').process(coefficient)
                self.result, self.image = Process_Image(self.base_window.qualitative_analysis_0.system_check_folder + '/system_check_raw.jpg').process(coefficient)
                #cv2.imwrite(self.base_window.qualitative_analysis_1.system_check_folder + '/system_check_process.jpg', self.image)
                cv2.imwrite(self.base_window.qualitative_analysis_0.system_check_folder + '/system_check_process.jpg', self.image)
            elif(self.mode_check==2):
                self.result, self.image = Process_Image(results_programs_quantitative_path + self.base_window.new_quantitative_1.experiment_name + '/system_check_raw.jpg').process(coefficient)
                cv2.imwrite(results_programs_quantitative_path + self.base_window.new_quantitative_1.experiment_name + '/system_check_process.jpg', self.image)
            elif(self.mode_check==3):
                self.result, self.image = Process_Image(self.base_window.quantitative_analysis_1.system_check_folder + '/system_check_raw.jpg').process(coefficient)
                cv2.imwrite(self.base_window.quantitative_analysis_1.system_check_folder + '/system_check_process.jpg', self.image)

            self.progressbar['value']=80
            self.base_window.update_idletasks()
            sleep(0.5)

            self.progressbar['value']=95
            self.base_window.update_idletasks()
            sleep(0.5)

            wb = Workbook()
            sheet = wb.active
            sheet["A2"] = "A"
            sheet["A3"] = "B"
            sheet["A4"] = "C"
            sheet["A5"] = "D"
            sheet["A6"] = "E"
            sheet["A7"] = "F"
            sheet["A8"] = "G"
            sheet["A9"] = "H"
            sheet["B1"] = "1"
            sheet["C1"] = "2"
            sheet["D1"] = "3"
            sheet["E1"] = "4"
            sheet["F1"] = "5"
            sheet["G1"] = "6"
            for i in range(0,48):
                if(i<6):
                    pos = str(chr(65+i+1)) + "2"
                if(i>=6 and i<12):
                    pos = str(chr(65+i-5)) + "3"
                if(i>=12 and i<18):
                    pos = str(chr(65+i-11)) + "4"
                if(i>=18 and i<24):
                    pos = str(chr(65+i-17)) + "5"
                if(i>=24 and i<30):
                    pos = str(chr(65+i-23)) + "6"
                if(i>=30 and i<36):
                    pos = str(chr(65+i-29)) + "7"
                if(i>=36 and i<42):
                    pos = str(chr(65+i-35)) + "8"
                if(i>=42):
                    pos = str(chr(65+i-41)) + "9"

                sheet[pos] = self.result[i]

            average_current_intensity = round(sum(self.result)/len(self.result),1)
            self.threshold = round(average_current_intensity*a+b,1)
            print("threshold: ", self.threshold)

            sheet['I2'] = "Average: " + str(average_current_intensity)
            sheet['I3'] = "Threshold: " + str(self.threshold)

            if(self.mode_check == 0):
                wb.save(results_programs_qualitative_path + self.base_window.new_qualitative_1.experiment_name + "/nonsample_value.xlsx")
            elif(self.mode_check == 1):
                #wb.save(self.base_window.qualitative_analysis_1.system_check_folder + "/nonsample_value.xlsx")
                wb.save(self.base_window.qualitative_analysis_0.system_check_folder + "/nonsample_value.xlsx")
            elif(self.mode_check == 2):
                wb.save(results_programs_quantitative_path + self.base_window.new_quantitative_1.experiment_name + "/nonsample_value.xlsx")
            elif(self.mode_check == 3):
                wb.save(self.base_window.quantitative_analysis_1.system_check_folder + "/nonsample_value.xlsx")

            wb.close()

            self.progressbar['value']=100
            self.base_window.update_idletasks()
            sleep(0.5)

            self.progressbar.destroy()
            self.process_label.destroy()

            average_base_intensity = round(sum(base_intensity)/len(base_intensity),1)
            tmp_value = round(average_base_intensity/average_current_intensity,2)
            print("average_base_intensity: ", average_base_intensity)
            print("average_current_intensity: ", average_current_intensity)

            self.check_result_frame = Frame(self.work_frame, bg=RESULT_TABLE_FRAME_BGD_COLOR)
            self.check_result_frame.pack()

            result_label = list(range(48))
            r=0
            c=-1
            for i in range(0,48):
                if(i<6):
                    t='A'+ str(i+1)
                if(i>=6 and i<12):
                    t='B'+ str(i-5)
                if(i>=12 and i<18):
                    t='C'+ str(i-11)
                if(i>=18 and i<24):
                    t='D'+ str(i-17)
                if(i>=24 and i<30):
                    t='E'+ str(i-23)
                if(i>=30 and i<36):
                    t='F'+ str(i-29)
                if(i>=36 and i<42):
                    t='G'+ str(i-35)
                if(i>=42):
                    t='H'+ str(i-41)

                c+=1
                if(c>5):
                    c=0
                    r+=1
                result_label[i] = Label(self.check_result_frame,
                                        text = t,
                                        width=6,
                                        height=3,
                                        bg = RESULT_LABEL_BGD_COLOR,
                                        font = RESULT_LABEL_TXT_FONT)
                result_label[i].grid(row=r,column=c, padx=2, pady=2)

                # Check system error 1
                if(self.result[i] > base_intensity[i]+base_intensity[i]*30/100 or self.result[i] < base_intensity[i]-base_intensity[i]*30/100):
                    result_label[i]['bg'] = RESULT_LABEL_ERROR_BGD_COLOR
                    self.err = 1

            # Check system error 2
            if(average_current_intensity > average_base_intensity + average_base_intensity*30/100 or
                average_current_intensity < average_base_intensity - average_base_intensity*30/100):
                    for i in range(0,48):
                        result_label[i]['bg'] = RESULT_LABEL_ERROR_BGD_COLOR
                    self.err = 0

            # Error handle
            if(self.err == 1):
                msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SYSTEM_ERROR_1'].value),
                                        ERROR_LIST(ERROR_LIST['SYSTEM_ERROR_1'].value).name,
                                        icon = "error")
            elif(self.err == 2):
                msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SYSTEM_ERROR_2'].value),
                                        ERROR_LIST(ERROR_LIST['SYSTEM_ERROR_2'].value).name,
                                        icon = "error")
            else:
                msg = messagebox.showinfo("", "Finished checking !")

        else:
            self.back_clicked()



class NewQualitativeFrame3(Frame):
    def __init__(self, container):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container

        # 3 main frame
        self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=5, fill=X)
        self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X)

        # In title frame
        self.title_label = Label(self.title_frame,
                                text = "CALIBRATION",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

    def serial_handle(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("1.Horizontal.TProgressbar", troughcolor ='grey85', background='green3')
        self.progressbar = ttk.Progressbar(self.work_frame,
                                    style="1.Horizontal.TProgressbar",
                                    length = 200,
                                    mode = 'determinate')
        self.progressbar.pack(ipadx=2, ipady=2)

        self.process_label = Label(self.work_frame,
                        bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
                        fg = 'black',
                        text = 'Processing...',
                        font = LABEL_TXT_FONT)
        self.process_label.pack(ipadx=2, ipady=2, anchor=N)

        ser.flushInput()
        ser.flushOutput()

        self.progressbar['value']=5
        self.base_window.update_idletasks()
        sleep(0.5)

        data_send = 'P'
        print("Data send:", data_send)
        ser.write(data_send.encode())

        receive_data = StringVar()
        count = 0
        bled_ready = 0
        while(receive_data != 'C'):
            if(ser.in_waiting>0):
                receive_data = ser.readline().decode('utf-8').rstrip()
                print("Data received:", receive_data)

                self.progressbar['value']=10
                self.base_window.update_idletasks()
                sleep(0.5)

                if(receive_data == 'C'):
                    self.progressbar['value']=20
                    self.base_window.update_idletasks()
                    sleep(0.5)

                    bled_ready = 1
                    break;
            else:
                sleep(1)
                count += 1
                if(count > 15):
                    msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value),
                                                ERROR_LIST(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value).name,
                                                icon = "error")
                    break;

        if(bled_ready):
            self.progressbar['value']=45
            self.base_window.update_idletasks()
            try:
                camera_capture(results_programs_qualitative_path + self.base_window.new_qualitative_1.experiment_name + '/raw.jpg')

                self.progressbar['value']=65
                self.base_window.update_idletasks()
            except Exception as e :
                msg = messagebox.showerror("ERR "+ str(ERROR_LIST['CAMERA_ERROR'].value),
                                        ERROR_LIST(ERROR_LIST['CAMERA_ERROR'].value).name,
                                        icon = "error")
                if(msg=='ok'):
                    self.base_window.destroy()

            self.result, self.image = Process_Image(results_programs_qualitative_path + self.base_window.new_qualitative_1.experiment_name + '/raw.jpg').process(coefficient, mode=1, well_list=self.base_window.new_qualitative_2.well_button)

            self.progressbar['value']=80
            self.base_window.update_idletasks()
            sleep(0.5)

            cv2.imwrite(results_programs_qualitative_path + self.base_window.new_qualitative_1.experiment_name + '/process.jpg', self.image)

            self.progressbar['value']=95
            self.base_window.update_idletasks()
            sleep(0.5)

            sum_value = 0
            active_well_number = 0

            # Save analysis file
            wb = Workbook()
            sheet = wb.active
            sheet["A2"] = "A"
            sheet["A3"] = "B"
            sheet["A4"] = "C"
            sheet["A5"] = "D"
            sheet["A6"] = "E"
            sheet["A7"] = "F"
            sheet["A8"] = "G"
            sheet["A9"] = "H"
            sheet["B1"] = "1"
            sheet["C1"] = "2"
            sheet["D1"] = "3"
            sheet["E1"] = "4"
            sheet["F1"] = "5"
            sheet["G1"] = "6"
            for i in range(0,48):
                if(i<6):
                    pos = str(chr(65+i+1)) + "2"
                if(i>=6 and i<12):
                    pos = str(chr(65+i-5)) + "3"
                if(i>=12 and i<18):
                    pos = str(chr(65+i-11)) + "4"
                if(i>=18 and i<24):
                    pos = str(chr(65+i-17)) + "5"
                if(i>=24 and i<30):
                    pos = str(chr(65+i-23)) + "6"
                if(i>=30 and i<36):
                    pos = str(chr(65+i-29)) + "7"
                if(i>=36 and i<42):
                    pos = str(chr(65+i-35)) + "8"
                if(i>=42):
                    pos = str(chr(65+i-41)) + "9"
                if(self.base_window.new_qualitative_2.well_button[i]['text'][0] != '#'):
                    active_well_number += 1
                    sum_value += round(self.result[i]/self.base_window.system_check.threshold,2)
                    sheet[pos] = round(self.result[i]/self.base_window.system_check.threshold,2)
                    print(self.result[i])
                else:
                    sheet[pos] = "N/A"
            wb.save(results_programs_qualitative_path + self.base_window.new_qualitative_1.experiment_name + "/analysis_value.xlsx")
            wb.close()

            self.final_value = round(sum_value/active_well_number,2)
            print("active_well_number:", active_well_number)
            print("sum_value:", sum_value)
            print("final_value:", self.final_value)

            # Save program file
            wb = Workbook()
            sheet = wb.active
            sheet["B2"] = "User name:"
            sheet["B3"] = "Comment:"
            sheet["D5"] = "A"
            sheet["D6"] = "B"
            sheet["D7"] = "C"
            sheet["D8"] = "D"
            sheet["D9"] = "E"
            sheet["D10"] = "F"
            sheet["D11"] = "G"
            sheet["D12"] = "H"
            sheet["E4"] = "1"
            sheet["F4"] = "2"
            sheet["G4"] = "3"
            sheet["H4"] = "4"
            sheet["I4"] = "5"
            sheet["J4"] = "6"

            sheet["C2"] = self.base_window.new_qualitative_1.user_name
            sheet["C3"] = self.base_window.new_qualitative_1.comments

            sheet["D2"] = "Base Value:"
            sheet["E2"] = self.final_value

            count = -1
            for i in range(0,48):
                if(i<6):
                    pos = str(chr(68+i+1)) + "5"
                if(i>=6 and i<12):
                    pos = str(chr(68+i-5)) + "6"
                if(i>=12 and i<18):
                    pos = str(chr(68+i-11)) + "7"
                if(i>=18 and i<24):
                    pos = str(chr(68+i-17)) + "8"
                if(i>=24 and i<30):
                    pos = str(chr(68+i-23)) + "9"
                if(i>=30 and i<36):
                    pos = str(chr(68+i-29)) + "10"
                if(i>=36 and i<42):
                    pos = str(chr(68+i-35)) + "11"
                if(i>=42):
                    pos = str(chr(68+i-41)) + "12"

                if(self.base_window.new_qualitative_2.well_button[i]['text'][0] != '#'):
                    count += 1
                    sheet[pos] = self.base_window.new_qualitative_2.well_button[i]['text']
                    sheet["B" + str(4 + count)] = self.base_window.new_qualitative_2.well_button[i]['text']
                    sheet["C" + str(4 + count)] = round(self.result[i]/self.base_window.system_check.threshold,2)
                else:
                    sheet[pos] = "N/A"
            wb.save(programs_qualitative_path + self.base_window.new_qualitative_1.experiment_name + ".xlsx")
            wb.close()

            self.progressbar['value']=100
            self.base_window.update_idletasks()
            sleep(0.5)

            self.progressbar.destroy()
            self.process_label.destroy()

            self.check_result_frame = Frame(self.work_frame, bg=RESULT_TABLE_FRAME_BGD_COLOR)
            self.check_result_frame.pack(side=LEFT, anchor=E)

            result_label = list(range(48))
            r=0
            c=-1
            for i in range(0,48):
                c+=1
                if(c>5):
                    c=0
                    r+=1
                result_label[i] = Label(self.check_result_frame,
                                        width=6,
                                        height=3,
                                        # ~ bg = RESULT_LABEL_BGD_COLOR,
                                        font = RESULT_LABEL_TXT_FONT)
                if(self.base_window.new_qualitative_2.well_button[i]['text'][0] != '#'):
                    result_label[i]['text'] = round(self.result[i]/self.base_window.system_check.threshold,2)
                    result_label[i]['bg'] = "cyan"
                else:
                    result_label[i]['text'] = "N/A"
                    result_label[i]['bg'] = "grey80"
                result_label[i].grid(row=r,column=c, padx=2, pady=2)

            self.pfi_label = Label(self.work_frame,
                                    bg = LABEL_BGD_COLOR,
                                    text = 'PFi value: ' + str(self.final_value),
                                    fg = 'blue',
                                    font = ("Helvetica", 13, 'bold'))
            self.pfi_label.pack(side=LEFT, ipadx=5, ipady=5, padx=100, pady=50, anchor=NE, expand=TRUE)

            self.title_label['text'] = "CALIBRATION RESULT"

            # In button frame
            self.finish_button = Button(self.button_frame,
                                    text = "FINISH",
                                    font = SWITCH_PAGE_BUTTON_FONT,
                                    # width = SWITCH_PAGE_BUTTON_WIDTH,
                                    # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                    bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                    fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.finish_clicked)
            self.finish_button.pack(ipady=10)

            msg = messagebox.showinfo("","COMPLETED")

        else:
            self.back_clicked()

    def finish_clicked(self):
        msg = messagebox.askquestion("","Do you want to go back to the main screen ?")
        if(msg=="yes"):
            self.base_window.forget_page()
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_option)
            # ~ self.base_window.page_num = PAGE_LIST.index('MAIN_MENU')
            self.base_window.switch_page()
            self.base_window.main_menu.reset()

class NewQuantitativeFrame3(NewQualitativeFrame3):
    def __init__(self, container):
        super().__init__(container)
        self.title_label['text'] = "CALIBRATION"

    def serial_handle(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("1.Horizontal.TProgressbar", troughcolor ='grey85', background='green3')
        self.progressbar = ttk.Progressbar(self.work_frame,
                                    style="1.Horizontal.TProgressbar",
                                    length = 200,
                                    mode = 'determinate')
        self.progressbar.pack(ipadx=2, ipady=2)

        self.process_label = Label(self.work_frame,
                        bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
                        fg = 'black',
                        text = 'Processing...',
                        font = LABEL_TXT_FONT)
        self.process_label.pack(ipadx=2, ipady=2, anchor=N)

        ser.flushInput()
        ser.flushOutput()

        self.progressbar['value']=5
        self.base_window.update_idletasks()
        sleep(0.5)

        data_send = 'P'
        print("Data send:", data_send)
        ser.write(data_send.encode())

        receive_data = StringVar()
        count = 0
        bled_ready = 0
        while(receive_data != 'C'):
            if(ser.in_waiting>0):
                receive_data = ser.readline().decode('utf-8').rstrip()
                print("Data received:", receive_data)

                self.progressbar['value']=10
                self.base_window.update_idletasks()
                sleep(0.5)

                if(receive_data == 'C'):
                    self.progressbar['value']=20
                    self.base_window.update_idletasks()
                    sleep(0.5)

                    bled_ready = 1
                    break;
            else:
                sleep(1)
                count += 1
                if(count > 15):
                    msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value),
                                                ERROR_LIST(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value).name,
                                                icon = "error")
                    break;

        if(bled_ready):
            self.progressbar['value']=45
            self.base_window.update_idletasks()
            try:
                camera_capture(results_programs_quantitative_path + self.base_window.new_quantitative_1.experiment_name + '/raw.jpg')

                self.progressbar['value']=65
                self.base_window.update_idletasks()
            except Exception as e :
                msg = messagebox.showerror("ERR "+ str(ERROR_LIST['CAMERA_ERROR'].value),
                                        ERROR_LIST(ERROR_LIST['CAMERA_ERROR'].value).name,
                                        icon = "error")
                if(msg=='ok'):
                    self.base_window.destroy()

            self.result, self.image = Process_Image(results_programs_quantitative_path + self.base_window.new_quantitative_1.experiment_name + '/raw.jpg').process(coefficient, mode=1, well_list=self.base_window.new_quantitative_2.well_button)

            self.progressbar['value']=80
            self.base_window.update_idletasks()
            sleep(0.5)

            cv2.imwrite(results_programs_quantitative_path + self.base_window.new_quantitative_1.experiment_name + '/process.jpg', self.image)

            self.progressbar['value']=95
            self.base_window.update_idletasks()
            sleep(0.5)

            sum_value = 0
            active_well_number = 0

            sum_concen_0 = 0
            sum_concen_1 = 0
            sum_concen_2 = 0
            sum_concen_3 = 0
            sum_concen_4 = 0
            sum_concen_5 = 0
            sum_concen_6 = 0
            sum_concen_7 = 0
            sum_concen_8 = 0
            sum_concen_9 = 0

            concen_0_number = 0
            concen_1_number = 0
            concen_2_number = 0
            concen_3_number = 0
            concen_4_number = 0
            concen_5_number = 0
            concen_6_number = 0
            concen_7_number = 0
            concen_8_number = 0
            concen_9_number = 0


            # Save analysis file
            wb = Workbook()
            sheet = wb.active
            sheet["A2"] = "A"
            sheet["A3"] = "B"
            sheet["A4"] = "C"
            sheet["A5"] = "D"
            sheet["A6"] = "E"
            sheet["A7"] = "F"
            sheet["A8"] = "G"
            sheet["A9"] = "H"
            sheet["B1"] = "1"
            sheet["C1"] = "2"
            sheet["D1"] = "3"
            sheet["E1"] = "4"
            sheet["F1"] = "5"
            sheet["G1"] = "6"
            for i in range(0,48):
                if(i<6):
                    pos = str(chr(65+i+1)) + "2"
                if(i>=6 and i<12):
                    pos = str(chr(65+i-5)) + "3"
                if(i>=12 and i<18):
                    pos = str(chr(65+i-11)) + "4"
                if(i>=18 and i<24):
                    pos = str(chr(65+i-17)) + "5"
                if(i>=24 and i<30):
                    pos = str(chr(65+i-23)) + "6"
                if(i>=30 and i<36):
                    pos = str(chr(65+i-29)) + "7"
                if(i>=36 and i<42):
                    pos = str(chr(65+i-35)) + "8"
                if(i>=42):
                    pos = str(chr(65+i-41)) + "9"
                if(self.base_window.new_quantitative_2.well_button[i]['text'][0] != '#'):
                    if(self.base_window.new_quantitative_2.concentration[i] == 0):
                        sum_concen_0 += self.result[i]/self.base_window.system_check.threshold
                        concen_0_number += 1
                    elif(self.base_window.new_quantitative_2.concentration[i] == 1):
                        sum_concen_1 += self.result[i]/self.base_window.system_check.threshold
                        concen_1_number += 1
                    elif(self.base_window.new_quantitative_2.concentration[i] == 2):
                        sum_concen_2 += self.result[i]/self.base_window.system_check.threshold
                        concen_2_number += 1
                    elif(self.base_window.new_quantitative_2.concentration[i] == 3):
                        sum_concen_3 += self.result[i]/self.base_window.system_check.threshold
                        concen_3_number += 1
                    elif(self.base_window.new_quantitative_2.concentration[i] == 4):
                        sum_concen_4 += self.result[i]/self.base_window.system_check.threshold
                        concen_4_number += 1
                    elif(self.base_window.new_quantitative_2.concentration[i] == 5):
                        sum_concen_5 += self.result[i]/self.base_window.system_check.threshold
                        concen_5_number += 1
                    elif(self.base_window.new_quantitative_2.concentration[i] == 6):
                        sum_concen_6 += self.result[i]/self.base_window.system_check.threshold
                        concen_6_number += 1
                    elif(self.base_window.new_quantitative_2.concentration[i] == 7):
                        sum_concen_7 += self.result[i]/self.base_window.system_check.threshold
                        concen_7_number += 1
                    elif(self.base_window.new_quantitative_2.concentration[i] == 8):
                        sum_concen_8 += self.result[i]/self.base_window.system_check.threshold
                        concen_8_number += 1
                    else:
                        sum_concen_9 += self.result[i]/self.base_window.system_check.threshold
                        concen_9_number += 1

                    sheet[pos] = round(self.result[i]/self.base_window.system_check.threshold,2)
                    print(self.result[i])
                else:
                    sheet[pos] = "N/A"

            print("sum_concen_0: ", sum_concen_0)
            print("concen_0_number: ",concen_0_number)
            print("sum_concen_1: ", sum_concen_1)
            print("concen_1_number: ",concen_1_number)
            print("sum_concen_2: ", sum_concen_2)
            print("concen_2_number: ",concen_2_number)
            print("sum_concen_3: ", sum_concen_3)
            print("concen_3_number: ",concen_3_number)
            print("sum_concen_4: ", sum_concen_4)
            print("concen_4_number: ",concen_4_number)
            print("sum_concen_5: ", sum_concen_5)
            print("concen_5_number: ",concen_5_number)
            print("sum_concen_6: ", sum_concen_6)
            print("concen_6_number: ",concen_6_number)
            print("sum_concen_7: ", sum_concen_7)
            print("concen_7_number: ",concen_7_number)
            print("sum_concen_8: ", sum_concen_8)
            print("concen_8_number: ",concen_8_number)
            print("sum_concen_9: ", sum_concen_9)
            print("concen_9_number: ",concen_9_number)
            wb.save(results_programs_quantitative_path + self.base_window.new_quantitative_1.experiment_name + "/analysis_value.xlsx")
            wb.close()

            number_concentration = 0
            pts_list = []
            if(concen_0_number != 0):
                number_concentration += 1
                avg_concen_0  = round(sum_concen_0/concen_0_number,2)
                print("avg_concen_0: ",avg_concen_0)
                # ~ concen_0_pt = [avg_concen_0,0]
                # ~ pts_list.append(concen_0_pt)
            if(concen_1_number != 0):
                number_concentration += 1
                avg_concen_1  = round(sum_concen_1/concen_1_number,2)
                print("avg_concen_1: ",avg_concen_1)
                concen_1_pt = [1, avg_concen_1]
                # ~ concen_1_pt = [1, 0.79]
                pts_list.append(concen_1_pt)
            if(concen_2_number != 0):
                number_concentration += 1
                avg_concen_2  = round(sum_concen_2/concen_2_number,2)
                print("avg_concen_2: ",avg_concen_2)
                concen_2_pt = [2, avg_concen_2]
                # ~ concen_2_pt = [2, 0.96]
                pts_list.append(concen_2_pt)
            if(concen_3_number != 0):
                number_concentration += 1
                avg_concen_3  = round(sum_concen_3/concen_3_number,2)
                print("avg_concen_3: ",avg_concen_3)
                concen_3_pt = [3, avg_concen_3]
                # ~ concen_3_pt = [3, 1.19]
                pts_list.append(concen_3_pt)
            if(concen_4_number != 0):
                number_concentration += 1
                avg_concen_4  = round(sum_concen_4/concen_4_number,2)
                print("avg_concen_4: ",avg_concen_4)
                concen_4_pt = [4, avg_concen_4]
                # ~ concen_4_pt = [4, 1.42]
                pts_list.append(concen_4_pt)
            if(concen_5_number != 0):
                number_concentration += 1
                avg_concen_5  = round(sum_concen_5/concen_5_number,2)
                print("avg_concen_5: ",avg_concen_5)
                concen_5_pt = [5, avg_concen_5]
                # ~ concen_5_pt = [5, 1.62]
                pts_list.append(concen_5_pt)
            if(concen_6_number != 0):
                number_concentration += 1
                avg_concen_6  = round(sum_concen_6/concen_6_number,2)
                print("avg_concen_6: ",avg_concen_6)
                concen_6_pt = [6, avg_concen_6]
                pts_list.append(concen_6_pt)
            if(concen_7_number != 0):
                number_concentration += 1
                avg_concen_7  = round(sum_concen_7/concen_7_number,2)
                print("avg_concen_7: ",avg_concen_7)
                concen_7_pt = [7, avg_concen_7]
                pts_list.append(concen_7_pt)
            if(concen_8_number != 0):
                number_concentration += 1
                avg_concen_8  = round(sum_concen_8/concen_8_number,2)
                print("avg_concen_8: ",avg_concen_8)
                concen_8_pt = [8, avg_concen_8]
                pts_list.append(concen_8_pt)
            if(concen_9_number != 0):
                number_concentration += 1
                avg_concen_9  = round(sum_concen_9/concen_9_number,2)
                print("avg_concen_9: ",avg_concen_9)
                concen_9_pt = [9, avg_concen_9]
                pts_list.append(concen_9_pt)

            pts_list = np.array(pts_list)
            print("pts_list: ", pts_list)

            x = pts_list[:,0]
            y = pts_list[:,1]

            a_value, b_value = np.polyfit(x,y,1)
            print("a_value: ", a_value)
            print("b_value: ", b_value)

            # Save program file
            wb = Workbook()
            sheet = wb.active
            sheet["B2"] = "User name:"
            sheet["B3"] = "Comment:"
            sheet["E5"] = "A"
            sheet["E6"] = "B"
            sheet["E7"] = "C"
            sheet["E8"] = "D"
            sheet["E9"] = "E"
            sheet["E10"] = "F"
            sheet["E11"] = "G"
            sheet["E12"] = "H"
            sheet["F4"] = "1"
            sheet["G4"] = "2"
            sheet["H4"] = "3"
            sheet["I4"] = "4"
            sheet["J4"] = "5"
            sheet["K4"] = "6"

            sheet["C2"] = self.base_window.new_quantitative_1.user_name
            sheet["C3"] = self.base_window.new_quantitative_1.comments

            sheet["D2"] = "a:"
            sheet["E2"] = a_value
            sheet["F2"] = "b:"
            sheet["G2"] = b_value
            sheet["H2"] = "N base value:"
            sheet["I2"] = avg_concen_0

            count = -1
            for i in range(0,48):
                if(i<6):
                    pos = str(chr(69+i+1)) + "5"
                if(i>=6 and i<12):
                    pos = str(chr(69+i-5)) + "6"
                if(i>=12 and i<18):
                    pos = str(chr(69+i-11)) + "7"
                if(i>=18 and i<24):
                    pos = str(chr(69+i-17)) + "8"
                if(i>=24 and i<30):
                    pos = str(chr(69+i-23)) + "9"
                if(i>=30 and i<36):
                    pos = str(chr(69+i-29)) + "10"
                if(i>=36 and i<42):
                    pos = str(chr(69+i-35)) + "11"
                if(i>=42):
                    pos = str(chr(68+i-41)) + "12"

                if(self.base_window.new_quantitative_2.well_button[i]['text'][0] != '#'):
                    count += 1
                    sheet[pos] = self.base_window.new_quantitative_2.well_button[i]['text']
                    sheet["B" + str(4 + count)] = self.base_window.new_quantitative_2.well_button[i]['text']
                    if(self.base_window.new_quantitative_2.concentration[i] == 0):
                        sheet["C" + str(4 + count)] = "0 copies"
                    if(self.base_window.new_quantitative_2.concentration[i] == 1):
                        sheet["C" + str(4 + count)] = "10e1 copies"
                    if(self.base_window.new_quantitative_2.concentration[i] == 2):
                        sheet["C" + str(4 + count)] = "10e2 copies"
                    if(self.base_window.new_quantitative_2.concentration[i] == 3):
                        sheet["C" + str(4 + count)] = "10e3 copies"
                    if(self.base_window.new_quantitative_2.concentration[i] == 4):
                        sheet["C" + str(4 + count)] = "10e4 copies"
                    if(self.base_window.new_quantitative_2.concentration[i] == 5):
                        sheet["C" + str(4 + count)] = "10e5 copies"
                    if(self.base_window.new_quantitative_2.concentration[i] == 6):
                        sheet["C" + str(4 + count)] = "10e6 copies"
                    if(self.base_window.new_quantitative_2.concentration[i] == 7):
                        sheet["C" + str(4 + count)] = "10e7 copies"
                    if(self.base_window.new_quantitative_2.concentration[i] == 8):
                        sheet["C" + str(4 + count)] = "10e8 copies"
                    if(self.base_window.new_quantitative_2.concentration[i] == 9):
                        sheet["C" + str(4 + count)] = "10e9 copies"
                    sheet["D" + str(4 + count)] = round(self.result[i]/self.base_window.system_check.threshold,2)
                else:
                    sheet[pos] = "N/A"
            wb.save(programs_quantitative_path + self.base_window.new_quantitative_1.experiment_name + ".xlsx")
            wb.close()

            self.progressbar['value']=100
            self.base_window.update_idletasks()
            sleep(0.5)

            self.progressbar.destroy()
            self.process_label.destroy()

            self.check_result_frame = Frame(self.work_frame, bg=RESULT_TABLE_FRAME_BGD_COLOR)
            self.check_result_frame.pack(side=LEFT, anchor=E)

            result_label = list(range(48))
            r=0
            c=-1
            for i in range(0,48):
                c+=1
                if(c>5):
                    c=0
                    r+=1
                result_label[i] = Label(self.check_result_frame,
                                        width=6,
                                        height=3,
                                        # ~ bg = RESULT_LABEL_BGD_COLOR,
                                        font = RESULT_LABEL_TXT_FONT)
                if(self.base_window.new_quantitative_2.well_button[i]['text'][0] != '#'):
                    result_label[i]['text'] = round(self.result[i]/self.base_window.system_check.threshold,2)
                    result_label[i]['bg'] = "cyan"
                else:
                    result_label[i]['text'] = "N/A"
                    result_label[i]['bg'] = "grey80"
                result_label[i].grid(row=r,column=c, padx=2, pady=2)


            self.pfi_label = Label(self.work_frame,
                                    bg = LABEL_BGD_COLOR,
                                    text = 'PFi value: ' + str(avg_concen_0),
                                    fg = 'blue',
                                    font = ("Helvetica", 13, 'bold'))
            self.pfi_label.pack(side=LEFT, ipadx=5, ipady=5, padx=100, pady=50, anchor=NE, expand=TRUE)

            self.title_label['text'] = "CALIBRATION RESULT"

            # In button frame
            self.finish_button = Button(self.button_frame,
                                    text = "FINISH",
                                    font = SWITCH_PAGE_BUTTON_FONT,
                                    # width = SWITCH_PAGE_BUTTON_WIDTH,
                                    # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                    bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                    fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.finish_clicked)
            self.finish_button.pack(ipady=10, side=RIGHT, ipadx=20, padx=40, anchor=W)

            def show_chart():
                print("x: ", x)
                print("y: ", y)
                pyplot.scatter(x,y)
                pyplot.plot(x, a_value*x + b_value,'r')
                pyplot.title("SC48 fit line")
                pyplot.xlabel('Concentration')
                pyplot.ylabel('Fi')
                pyplot.show()

            self.show_chart_button = Button(self.button_frame,
                                    text = "Chart",
                                    font = SWITCH_PAGE_BUTTON_FONT,
                                    # width = SWITCH_PAGE_BUTTON_WIDTH,
                                    # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                    bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                    fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = show_chart)
            self.show_chart_button.pack(ipady=10, ipadx=20, padx=40, side=LEFT, anchor=E)

            msg = messagebox.showinfo("","COMPLETED")

        else:
            self.back_clicked()

    def finish_clicked(self):
        msg = messagebox.askquestion("","Do you want to go back to the main screen ?")
        if(msg=="yes"):
            self.base_window.forget_page()
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
            # ~ self.base_window.page_num = PAGE_LIST.index('MAIN_MENU')
            self.base_window.switch_page()
            self.base_window.main_menu.reset()

class NewQualitativeFrame2(Frame):
    def __init__(self, container):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container

        # 3 main frame
        self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=5, fill=X)
        self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X)

        # In title frame
        self.title_label = Label(self.title_frame,
                                text = "CALIBRATION",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

        # In work frame
        # Sample button frame
        self.well_button_table_frame = Frame(self.work_frame, bg=SAMPLE_BUTTON_FRAME_BDG_COLOR)
        self.well_button_table_frame.pack(side=LEFT)
        self.well_button = list(range(48))
        r=0
        c=-1
        for i in range(0,48):
            c+=1
            if(c>5):
                c=0
                r+=1
            self.well_button[i] = Button(self.well_button_table_frame,
                                        bg = SAMPLE_BUTTON_BGD_COLOR,
                                        fg = SAMPLE_BUTTON_TXT_COLOR,
                                        activebackground = SAMPLE_BUTTON_ACTIVE_BGD_COLOR,
                                        justify = 'left',
                                        borderwidth = 0,
                                        text = '#',
                                        width = SAMPLE_BUTTON_WIDTH,
                                        height = SAMPLE_BUTTON_HEIGHT)
            self.well_button[i]['command'] = partial(self.well_button_clicked, i)
            self.well_button[i].grid(row=r, column=c, padx=2, pady=2)

        # Properties frame
        self.property_frame = Frame(self.work_frame, bg=SAMPLE_BUTTON_FRAME_BDG_COLOR, width=495)
        self.property_frame.pack(fill=BOTH, expand=TRUE, side=RIGHT)

        self.property_labelframe = LabelFrame(self.property_frame,
                                        text = "Sample Properties",
                                        font  = LABEL_FRAME_TXT_FONT,
                                        bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
                                        fg = LABEL_FRAME_TXT_COLOR)
        self.property_labelframe.pack(fill=BOTH, expand=TRUE, padx=10, pady=10)

        self.property_labelframe.rowconfigure(0, weight=1)
        self.property_labelframe.rowconfigure(1, weight=1)
        self.property_labelframe.rowconfigure(2, weight=1)
        self.property_labelframe.rowconfigure(3, weight=4)

        self.well_name_label = Label(self.property_labelframe,
                        bg = SAMPLE_BUTTON_CHOOSE_BGD_COLOR,
                        fg = LABEL_TXT_COLOR,
                        font = SAMPLE_LABEL_TXT_FONT)
        self.well_name_label.grid(row=0, column=0, columnspan=2, sticky=EW)

        # In button frame
        self.back_button = Button(self.button_frame,
                                text = "Back",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.back_clicked)
        self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

        self.note_label = Label(self.button_frame,
                            text='Please set and input primer mix kit',
                            fg = 'red',
                            bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR,
                            font = LABEL_TXT_FONT)
        self.note_label.pack(side=LEFT, ipadx=175)

        self.next_button = Button(self.button_frame,
                                text = "Next",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.next_clicked)
        self.next_button.pack(ipadx=30, ipady=10, side=RIGHT)

    def well_button_clicked(self,n):
        if(self.well_button[n]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR):
            for k in range (0,48):
                if(self.well_button[k]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR and self.well_button[k]['bg'] != SAMPLE_BUTTON_TMP_BGD_COLOR):
                    self.well_button[k]['bg'] = SAMPLE_BUTTON_BGD_COLOR
                else:
                    self.well_button[k]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
            self.well_button[n]['bg'] = SAMPLE_BUTTON_CHOOSE_BGD_COLOR
        else:
            for k in range (0,48):
                if(self.well_button[k]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR and self.well_button[k]['bg'] != SAMPLE_BUTTON_TMP_BGD_COLOR):
                    self.well_button[k]['bg'] = SAMPLE_BUTTON_BGD_COLOR
                if(self.well_button[k]['bg'] == SAMPLE_BUTTON_TMP_BGD_COLOR):
                    self.well_button[k]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
            self.well_button[n]['bg'] = SAMPLE_BUTTON_TMP_BGD_COLOR


        def ok_clicked(event=None):
            if(self.sample_name_entry.get()==''):
                self.well_button[n]['bg'] = SAMPLE_BUTTON_CHOOSE_BGD_COLOR
                self.well_button[n]['text'] = '#'
                msgbox = messagebox.showwarning("","Please enter sample name !")
            else:
                self.well_button[n]['text'] = self.sample_name_entry.get()
                self.well_button[n]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
                try:
                    if(n==42):
                        self.well_button_clicked(1)
                    elif(n==43):
                        self.well_button_clicked(2)
                    elif(n==44):
                        self.well_button_clicked(3)
                    elif(n==45):
                        self.well_button_clicked(4)
                    elif(n==46):
                        self.well_button_clicked(5)
                    elif(n==47):
                        self.well_button_clicked(0)
                    else:
                        self.well_button_clicked(n+6)
                except:
                    self.well_button_clicked(0)


        sample_name_label = Label(self.property_labelframe,
                                    text = "Sample Name",
                                    font = LABEL_TXT_FONT,
                                    bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
                                    fg = LABEL_TXT_COLOR)
        sample_name_label.grid(row=1, column=0, padx=78, pady=2, sticky=SE)

        self.sample_name_entry = Entry(self.property_labelframe, width=20, font=ENTRY_TXT_FONT)
        if(self.well_button[n]['bg'] == SAMPLE_BUTTON_TMP_BGD_COLOR):
            self.sample_name_entry.insert(0, self.well_button[n]['text'])
        #id_entry.bind("<Button-1>", enter_entry)
        self.sample_name_entry.bind("<Return>", ok_clicked)
        self.sample_name_entry.grid(row=2, column=0, padx=30, pady=0)
        self.sample_name_entry.focus_set()

        if(n<6):
            self.well_name_label['text'] = "A" + str(n+1)
        elif(n<12):
            self.well_name_label['text'] = "B" + str(n+1-6)
        elif(n<18):
            self.well_name_label['text'] = "C" + str(n+1-12)
        elif(n<24):
            self.well_name_label['text'] = "D" + str(n+1-18)
        elif(n<30):
            self.well_name_label['text'] = "E" + str(n+1-24)
        elif(n<36):
            self.well_name_label['text'] = "F" + str(n+1-30)
        elif(n<42):
            self.well_name_label['text'] = "G" + str(n+1-36)
        else:
            self.well_name_label['text'] = "H" + str(n+1-42)

        # ~ if(n<8):
            # ~ self.well_name_label['text'] = str(chr(65+n)) + '1'
        # ~ if(n>=8 and n<16):
            # ~ self.well_name_label['text'] = str(chr(65+n-8)) + '2'
        # ~ if(n>=16 and n<24):
            # ~ self.well_name_label['text'] = str(chr(65+n-16)) + '3'
        # ~ if(n>=24 and n<32):
            # ~ self.well_name_label['text'] = str(chr(65+n-24)) + '4'
        # ~ if(n>=32 and n<40):
            # ~ self.well_name_label['text'] = str(chr(65+n-32)) + '5'
        # ~ if(n>=40):
            # ~ self.well_name_label['text'] = str(chr(65+n-40)) + '6'

        self.ok_button = Button(self.property_labelframe,
                                text = "OK",
                                bg = CONFIRM_BUTTON_BGD_COLOR,
                                fg = CONFIRM_BUTTON_TXT_COLOR,
                                font = CONFIRM_BUTTON_TXT_FONT,
                                borderwidth = 0,
                                command = ok_clicked)
        self.ok_button.grid(row=3, column=0, columnspan=2, ipadx=30, ipady=10)

    def back_clicked(self):
        self.base_window.frame_list.remove(self.base_window.system_check)
        del self.base_window.system_check
        self.base_window.system_check = SystemCheckFrame(self.base_window)
        self.base_window.frame_list.append(self.base_window.system_check)

        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_qualitative_1)
        self.base_window.switch_page()


    def next_clicked(self):
        self.number_well_active = 0
        for i in range(0,48):
            if(self.well_button[i]['text'] != '#'):
                self.number_well_active += 1

        if(self.number_well_active == 0):
            msg = messagebox.showwarning("","You need to enter at least 1 sample name !")
        else:
            self.base_window.forget_page()
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_qualitative_3)
            self.base_window.switch_page()
            self.base_window.update_idletasks()
            self.base_window.new_qualitative_3.serial_handle()


class NewQuantitativeFrame2(NewQualitativeFrame2):
    def __init__(self, container):
        super().__init__(container)

        # In title frame
        self.title_label['text'] = "CALIBRATION"

        self.concentration = list(range(48))

    def well_button_clicked(self,n):
        if(self.well_button[n]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR):
            for k in range (0,48):
                if(self.well_button[k]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR and self.well_button[k]['bg'] != SAMPLE_BUTTON_TMP_BGD_COLOR):
                    self.well_button[k]['bg'] = SAMPLE_BUTTON_BGD_COLOR
                else:
                    self.well_button[k]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
            self.well_button[n]['bg'] = SAMPLE_BUTTON_CHOOSE_BGD_COLOR
        else:
            for k in range (0,48):
                if(self.well_button[k]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR and self.well_button[k]['bg'] != SAMPLE_BUTTON_TMP_BGD_COLOR):
                    self.well_button[k]['bg'] = SAMPLE_BUTTON_BGD_COLOR
                if(self.well_button[k]['bg'] == SAMPLE_BUTTON_TMP_BGD_COLOR):
                    self.well_button[k]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
            self.well_button[n]['bg'] = SAMPLE_BUTTON_TMP_BGD_COLOR


        def ok_clicked(event=None):
            if(self.sample_name_entry.get()==''):
                self.well_button[n]['bg'] = SAMPLE_BUTTON_CHOOSE_BGD_COLOR
                self.well_button[n]['text'] = '#'+str(n+1)
                msgbox = messagebox.showwarning("","Please enter sample name !")
            else:
                self.well_button[n]['text'] = self.sample_name_entry.get()
                self.well_button[n]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
                self.concentration[n] = self.concentration_combobox.current()
                print("concentration[%d] = %s"%(n,self.concentration[n]))
                try:
                    if(n==42):
                        self.well_button_clicked(1)
                    elif(n==43):
                        self.well_button_clicked(2)
                    elif(n==44):
                        self.well_button_clicked(3)
                    elif(n==45):
                        self.well_button_clicked(4)
                    elif(n==46):
                        self.well_button_clicked(5)
                    elif(n==47):
                        self.well_button_clicked(0)
                    else:
                        self.well_button_clicked(n+6)
                except:
                    self.well_button_clicked(0)


        sample_name_label = Label(self.property_labelframe,
                                    text = "Sample Name",
                                    font = LABEL_TXT_FONT,
                                    bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
                                    fg = LABEL_TXT_COLOR)
        sample_name_label.grid(row=1, column=0, padx=78, pady=2, sticky=SE)

        concentration_label = Label(self.property_labelframe,
                                    text = "Concentration (copies)",
                                    font = LABEL_TXT_FONT,
                                    bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
                                    fg = LABEL_TXT_COLOR)
        concentration_label.grid(row=1, column=1, padx=78, pady=2, sticky=S)

        self.sample_name_entry = Entry(self.property_labelframe, width=20, font=ENTRY_TXT_FONT)
        if(self.well_button[n]['bg'] == SAMPLE_BUTTON_TMP_BGD_COLOR):
            self.sample_name_entry.insert(0, self.well_button[n]['text'])
        #id_entry.bind("<Button-1>", enter_entry)
        self.sample_name_entry.bind("<Return>", ok_clicked)
        self.sample_name_entry.grid(row=2, column=0, padx=0, pady=0, sticky=NE)
        self.sample_name_entry.focus_set()


        self.concentration_value = StringVar()
        self.concentration_combobox = ttk.Combobox(self.property_labelframe,
                                                state = "readonly",
                                                width = 8,
                                                textvariable = self.concentration_value)
        self.concentration_combobox['values'] = ('0',
                                                '10e1',
                                                '10e2',
                                                '10e3',
                                                '10e4',
                                                '10e5',
                                                '10e6',
                                                '10e7',
                                                '10e8',
                                                '10e9')
        if(self.well_button[n]['text'][0] != '#'):
            self.concentration_combobox.current(self.concentration[n])
        else:
            self.concentration_combobox.current(0)

        self.concentration_combobox.grid(row=2, column=1, padx=0, pady=0, sticky=N)


        if(n<6):
            self.well_name_label['text'] = "A" + str(n+1)
        elif(n<12):
            self.well_name_label['text'] = "B" + str(n+1-6)
        elif(n<18):
            self.well_name_label['text'] = "C" + str(n+1-12)
        elif(n<24):
            self.well_name_label['text'] = "D" + str(n+1-18)
        elif(n<30):
            self.well_name_label['text'] = "E" + str(n+1-24)
        elif(n<36):
            self.well_name_label['text'] = "F" + str(n+1-30)
        elif(n<42):
            self.well_name_label['text'] = "G" + str(n+1-36)
        else:
            self.well_name_label['text'] = "H" + str(n+1-42)


        self.ok_button = Button(self.property_labelframe,
                                text = "OK",
                                bg = CONFIRM_BUTTON_BGD_COLOR,
                                fg = CONFIRM_BUTTON_TXT_COLOR,
                                font = CONFIRM_BUTTON_TXT_FONT,
                                borderwidth = 0,
                                command = ok_clicked)
        self.ok_button.grid(row=3, column=0, columnspan=2, ipadx=30, ipady=10)


    def back_clicked(self):
        self.base_window.frame_list.remove(self.base_window.system_check)
        del self.base_window.system_check
        self.base_window.system_check = SystemCheckFrame(self.base_window)
        self.base_window.frame_list.append(self.base_window.system_check)

        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_quantitative_1)
        self.base_window.switch_page()

    def next_clicked(self):
        self.number_well_active = 0
        self.number_concentration = 0
        concen0_flag = 0
        concen1_flag = 0
        concen2_flag = 0
        concen3_flag = 0
        concen4_flag = 0
        concen5_flag = 0
        concen6_flag = 0
        concen7_flag = 0
        concen8_flag = 0
        concen9_flag = 0
        for i in range(0,48):
            if(self.well_button[i]['bg'] == SAMPLE_BUTTON_DONE_BGD_COLOR):
                self.number_well_active += 1
                if(self.concentration[i] == 0):
                    if(concen0_flag == 0):
                        self.number_concentration += 1
                        concen0_flag = 1
                if(self.concentration[i] == 1):
                    if(concen1_flag == 0):
                        self.number_concentration += 1
                        concen1_flag = 1
                if(self.concentration[i] == 2):
                    if(concen2_flag == 0):
                        self.number_concentration += 1
                        concen2_flag = 1
                if(self.concentration[i] == 3):
                    if(concen3_flag == 0):
                        self.number_concentration += 1
                        concen3_flag = 1
                if(self.concentration[i] == 4):
                    if(concen4_flag == 0):
                        self.number_concentration += 1
                        concen4_flag = 1
                if(self.concentration[i] == 5):
                    if(concen5_flag == 0):
                        self.number_concentration += 1
                        concen5_flag = 1
                if(self.concentration[i] == 6):
                    if(concen6_flag == 0):
                        self.number_concentration += 1
                        concen6_flag = 1
                if(self.concentration[i] == 7):
                    if(concen7_flag == 0):
                        self.number_concentration += 1
                        concen7_flag = 1
                if(self.concentration[i] == 8):
                    if(concen8_flag == 0):
                        self.number_concentration += 1
                        concen8_flag = 1
                if(self.concentration[i] == 9):
                    if(concen9_flag == 0):
                        self.number_concentration += 1
                        concen9_flag = 1

        if(self.number_well_active == 0):
            msg = messagebox.showwarning("","Please enter sample name !")
        elif(self.number_concentration <= 1 or concen0_flag == 0): # it nhat phai co 2 concentratiobn khac nhau
            msg = messagebox.showwarning("","Must have at least 3 different concentrations (include 0 copy)!")
        else:
            self.base_window.forget_page()
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_quantitative_3)
            self.base_window.switch_page()
            self.base_window.update_idletasks()
            self.base_window.new_quantitative_3.serial_handle()



class NewQualitativeFrame1(Frame):
    def __init__(self, container):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container

        self.experiment_name = StringVar()
        self.user_name = StringVar()
        self.comments = StringVar()

        # 3 main frame
        self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=5, fill=X)
        self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X)

        # In title frame
        self.title_label = Label(self.title_frame,
                                text = "CALIBRATION",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

        # In work frame
        setup_labelframe = LabelFrame(self.work_frame,
                                    text = "Properties",
                                    font = LABEL_FRAME_TXT_FONT,
                                    bg = LABEL_FRAME_BGD_COLOR,
                                    fg = LABEL_FRAME_TXT_COLOR)
        setup_labelframe.pack(expand=TRUE)

        experiment_name_label = Label(setup_labelframe,
                                    text = "Kit Name",
                                    font = LABEL_TXT_FONT,
                                    bg = LABEL_BGD_COLOR,
                                    fg = LABEL_TXT_COLOR)
        experiment_name_label.grid(row=0, column=1, sticky=E, pady=20, padx=30)

        user_name_label = Label(setup_labelframe,
                                text = "User Name",
                                font = LABEL_TXT_FONT,
                                bg = LABEL_BGD_COLOR,
                                fg = LABEL_TXT_COLOR)
        user_name_label.grid(row=1, column=1, sticky=E, pady=20, padx=30)

        comment_label = Label(setup_labelframe,
                                text = "Comment",
                                font = LABEL_TXT_FONT,
                                bg = LABEL_BGD_COLOR,
                                fg = LABEL_TXT_COLOR)
        comment_label.grid(row=2, column=1, sticky=NE, pady=20, padx=30)

        self.experiment_name_entry = Entry(setup_labelframe, width=30, font=ENTRY_TXT_FONT)
        self.experiment_name_entry.grid(row=0, column=2, sticky=W, pady=20, padx=35)
        self.user_name_entry = Entry(setup_labelframe, width=30, font=ENTRY_TXT_FONT)
        self.user_name_entry.grid(row=1, column=2, sticky=W, pady=20, padx=35)
        self.comments_text = Text(setup_labelframe, width=30, height=9, font=ENTRY_TXT_FONT)
        self.comments_text.grid(row=2, column=2, sticky=E, pady=20, padx=35)

        # In button frame
        self.back_button = Button(self.button_frame,
                                text = "Back",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.back_clicked)
        self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

        self.next_button = Button(self.button_frame,
                                text = "Next",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.next_clicked)
        self.next_button.pack(ipadx=30, ipady=10, side=RIGHT)

    def back_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_calib_list)
        self.base_window.switch_page()

    def next_clicked(self):
        msg = messagebox.askokcancel("","Please make sure no sample is placed in the device !")
        if(msg == True):
            self.experiment_name = self.experiment_name_entry.get()
            self.user_name = self.user_name_entry.get()
            self.comments = self.comments_text.get("1.0",'end-1c')
            if(self.experiment_name==''):
                messagebox.showwarning("","Plese enter Experiment Name !")
            else:
                if (os.path.exists(programs_qualitative_path + self.experiment_name + ".xlsx")):
                    msg = messagebox.askquestion("","This Experiment Name already exists.\n Do you want to replace it?")
                    if(msg == 'yes'):
                        if os.path.exists(results_programs_qualitative_path + self.experiment_name):
                            f = results_programs_qualitative_path + self.experiment_name
                            shutil.rmtree(f)
                            os.mkdir(f)
                        else:
                            f = os.path.join(results_programs_qualitative_path + self.base_window.experiment_name)
                            os.mkdir(f)

                        self.base_window.system_check.mode_check = 0
                        self.base_window.forget_page()
                        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
                        self.base_window.switch_page()
                        self.base_window.update_idletasks()
                        self.base_window.system_check.serial_handle()
                else:
                    if os.path.exists(results_programs_qualitative_path + self.experiment_name):
                        f = results_programs_qualitative_path + self.experiment_name
                        shutil.rmtree(f)
                        os.mkdir(f)
                    else:
                        f = os.path.join(results_programs_qualitative_path + self.experiment_name)
                        os.mkdir(f)
                    self.base_window.system_check.mode_check = 0
                    self.base_window.forget_page()
                    self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
                    self.base_window.switch_page()
                    self.base_window.update_idletasks()
                    self.base_window.system_check.serial_handle()


class NewQuantitativeFrame1(NewQualitativeFrame1):
    def __init__(self, container):
        super().__init__(container)
        self.title_label['text'] = "CALIBRATION"

    def back_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_calib_list)
        self.base_window.switch_page()

    def next_clicked(self):
        msg = messagebox.askokcancel("","Please make sure no sample is placed in the device !")
        if(msg == True):
            self.experiment_name = self.experiment_name_entry.get()
            self.user_name = self.user_name_entry.get()
            self.comments = self.comments_text.get("1.0",'end-1c')
            if(self.experiment_name==''):
                messagebox.showwarning("","Plese enter Experiment Name !")
            else:
                if (os.path.exists(programs_quantitative_path + self.experiment_name + ".xlsx")):
                    msg = messagebox.askquestion("","This Experiment Name already exists.\n Do you want to replace it?")
                    if(msg == 'yes'):
                        if os.path.exists(results_programs_quantitative_path + self.experiment_name):
                            f = results_programs_quantitative_path + self.experiment_name
                            shutil.rmtree(f)
                            os.mkdir(f)
                        else:
                            f = os.path.join(results_programs_quantitative_path + self.base_window.experiment_name)
                            os.mkdir(f)

                        self.base_window.system_check.mode_check = 2
                        self.base_window.forget_page()
                        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
                        self.base_window.switch_page()
                        self.base_window.update_idletasks()
                        self.base_window.system_check.serial_handle()
                else:
                    if os.path.exists(results_programs_quantitative_path + self.experiment_name):
                        f = results_programs_quantitative_path + self.experiment_name
                        shutil.rmtree(f)
                        os.mkdir(f)
                    else:
                        f = os.path.join(results_programs_quantitative_path + self.experiment_name)
                        os.mkdir(f)
                    self.base_window.system_check.mode_check = 2
                    self.base_window.forget_page()
                    self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
                    self.base_window.switch_page()
                    self.base_window.update_idletasks()
                    self.base_window.system_check.serial_handle()

class QualitativeAnalysisFrame3(Frame):
    def __init__(self, container):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container

        self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=5, fill=X)
        self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X)

        # In title frame
        self.title_label = Label(self.title_frame,
                                text = "ANALYSIS",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

    def serial_handle(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("1.Horizontal.TProgressbar", troughcolor ='grey85', background='green3')
        self.progressbar = ttk.Progressbar(self.work_frame,
                                    style="1.Horizontal.TProgressbar",
                                    length = 200,
                                    mode = 'determinate')
        self.progressbar.pack(ipadx=2, ipady=2)

        self.process_label = Label(self.work_frame,
                        bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
                        fg = 'black',
                        text = 'Processing...',
                        font = LABEL_TXT_FONT)
        self.process_label.pack(ipadx=2, ipady=2, anchor=N)

        ser.flushInput()
        ser.flushOutput()

        self.progressbar['value']=5
        self.base_window.update_idletasks()
        sleep(0.5)

        data_send = 'P'
        print("Data send:", data_send)
        ser.write(data_send.encode())

        receive_data = StringVar()
        count = 0
        bled_ready = 0
        while(receive_data != 'C'):
            if(ser.in_waiting>0):
                receive_data = ser.readline().decode('utf-8').rstrip()
                print("Data received:", receive_data)

                self.progressbar['value']=10
                self.base_window.update_idletasks()
                sleep(0.5)

                if(receive_data == 'C'):
                    self.progressbar['value']=20
                    self.base_window.update_idletasks()
                    sleep(0.5)

                    bled_ready = 1
                    break;
            else:
                sleep(1)
                count += 1
                if(count > 15):
                    msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value),
                                                ERROR_LIST(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value).name,
                                                icon = "error")
                    break;

        if(bled_ready):
            self.progressbar['value']=45
            self.base_window.update_idletasks()
            try:
                #camera_capture(self.base_window.qualitative_analysis_1.analysis_result_folder + '/raw.jpg')
                camera_capture(self.base_window.qualitative_analysis_0.analysis_result_folder + '/raw.jpg')
                
                self.progressbar['value']=65
                self.base_window.update_idletasks()
            except Exception as e :
                msg = messagebox.showerror("ERR "+ str(ERROR_LIST['CAMERA_ERROR'].value),
                                        ERROR_LIST(ERROR_LIST['CAMERA_ERROR'].value).name,
                                        icon = "error")
                if(msg=='ok'):
                    self.base_window.destroy()

            #self.result, self.image = Process_Image(self.base_window.qualitative_analysis_1.analysis_result_folder + '/raw.jpg').process(coefficient, mode=1, well_list=self.base_window.qualitative_analysis_2.id_list)
            self.result, self.image = Process_Image(self.base_window.qualitative_analysis_0.analysis_result_folder + '/raw.jpg').process(coefficient, mode=1, well_list=self.base_window.qualitative_analysis_2.id_list)            
            
            self.progressbar['value']=80
            self.base_window.update_idletasks()
            sleep(0.5)

            #cv2.imwrite(self.base_window.qualitative_analysis_1.analysis_result_folder + '/process.jpg', self.image)
            cv2.imwrite(self.base_window.qualitative_analysis_0.analysis_result_folder + '/process.jpg', self.image)

            sum_value = 0
            active_well_number = 0

            # Save analysis file
            wb = Workbook()
            sheet = wb.active

            sheet["A2"] = "A"
            sheet["A3"] = "B"
            sheet["A4"] = "C"
            sheet["A5"] = "D"
            sheet["A6"] = "E"
            sheet["A7"] = "F"
            sheet["A8"] = "G"
            sheet["A9"] = "H"
            sheet["B1"] = "1"
            sheet["C1"] = "2"
            sheet["D1"] = "3"
            sheet["E1"] = "4"
            sheet["F1"] = "5"
            sheet["G1"] = "6"
            for i in range(0,48):
                if(i<6):
                    pos = str(chr(65+i+1)) + "2"
                if(i>=6 and i<12):
                    pos = str(chr(65+i-5)) + "3"
                if(i>=12 and i<18):
                    pos = str(chr(65+i-11)) + "4"
                if(i>=18 and i<24):
                    pos = str(chr(65+i-17)) + "5"
                if(i>=24 and i<30):
                    pos = str(chr(65+i-23)) + "6"
                if(i>=30 and i<36):
                    pos = str(chr(65+i-29)) + "7"
                if(i>=36 and i<42):
                    pos = str(chr(65+i-35)) + "8"
                if(i>=42):
                    pos = str(chr(65+i-41)) + "9"

                if(self.base_window.qualitative_analysis_2.id_list[i] != 'N/A'):
                    sheet[pos] = round(self.result[i]/self.base_window.system_check.threshold,2)
                else:
                    sheet[pos] = "N/A"

            #wb.save(self.base_window.qualitative_analysis_1.analysis_result_folder + "/analysis_value.xlsx")
            wb.save(self.base_window.qualitative_analysis_0.analysis_result_folder + "/analysis_value.xlsx")
            
            wb.close()

            self.progressbar['value']=95
            self.base_window.update_idletasks()
            sleep(0.5)

            self.progressbar['value']=100
            self.base_window.update_idletasks()
            sleep(0.5)

            self.progressbar.destroy()
            self.process_label.destroy()

            tab_control = ttk.Notebook(self.work_frame)
            result_tab = Frame(tab_control, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
            report_tab = Frame(tab_control, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
            image_tab = Frame(tab_control, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)

            tab_control.add(result_tab, text="Result")
            tab_control.add(report_tab, text="Report")
            tab_control.add(image_tab, text="Images")
            tab_control.grid(row=0, column=0, padx=0, pady=1, sticky=EW)

            self.check_result_frame = Frame(result_tab, bg=RESULT_TABLE_FRAME_BGD_COLOR)
            self.check_result_frame.grid(row=0, column=0, padx=76)

            # ~ Pmw.initialise(self.base_window)
            # ~ self.tooltip = list(range(48))
            
            
            self.pfi_value = round(self.result[47]/self.base_window.system_check.threshold,2)

            result_label = list(range(48))
            r=0
            c=-1
            
            for i in range(0,48):
                c+=1
                if(c>5):
                    c=0
                    r+=1
                result_label[i] = Label(self.check_result_frame,
                                        width=6,
                                        height=3,
                                        # ~ bg = RESULT_LABEL_BGD_COLOR,
                                        font = RESULT_LABEL_TXT_FONT)

                # ~ self.tooltip[i] = Pmw.Balloon(self.base_window)
                # ~ self.tooltip[i].bind(result_label[i], self.base_window.qualitative_analysis_2.id_list[i])
                
                if(self.base_window.qualitative_analysis_2.id_list[i] != 'N/A'):
                    result_label[i]['text'] = round(self.result[i]/self.base_window.system_check.threshold,2)
                    if(float(result_label[i]['text']) < self.pfi_value * NUM_1):
                        result_label[i]['bg'] = NEGATIVE_COLOR
                    elif(float(result_label[i]['text']) < self.pfi_value * NUM_2):
                        result_label[i]['bg'] = LOW_COPY_COLOR
                    else:
                        result_label[i]['bg'] = POSITIVE_COLOR

                else:
                    result_label[i]['text'] = "N/A"
                    result_label[i]['bg'] = NA_COLOR
                result_label[i].grid(row=r,column=c, padx=1, pady=1)
            result_label[47]['text'] = 'B'
            result_label[47]['bg'] = 'blue'
            
                        
            self.annotate_result_frame = Frame(result_tab, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
            self.annotate_result_frame.grid(row=0, column=1, padx=60)
            
            negative_label = Label(self.annotate_result_frame, bg=NEGATIVE_COLOR, width=4, height=2)
            negative_label.grid(row=0, column=0, padx=20, pady=10)
            negative_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="NEGATIVE (N)", height=2)
            negative_text_label.grid(row=0, column=1, padx=20, pady=10)
            low_copy_label = Label(self.annotate_result_frame, bg=LOW_COPY_COLOR, width=4, height=2)
            low_copy_label.grid(row=1, column=0, padx=20, pady=10)
            low_copy_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="LOW COPY (P_L)", height=2)
            low_copy_text_label.grid(row=1, column=1, padx=20, pady=10)
            positive_label = Label(self.annotate_result_frame, bg=POSITIVE_COLOR, width=4, height=2)
            positive_label.grid(row=2, column=0, padx=20, pady=10)
            positive_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="POSITIVE (P_H)", height=2)
            positive_text_label.grid(row=2, column=1, padx=20, pady=10)
            na_label = Label(self.annotate_result_frame, bg=NA_COLOR, width=4, height=2)
            na_label.grid(row=3, column=0, padx=20, pady=10)
            na_copy_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="NO SAMPLE (N/A)", height=2)
            na_copy_text_label.grid(row=3, column=1, padx=20, pady=10)
            

            
            # images tab
            img_labelframe_1 = LabelFrame(image_tab,
                                        bg="black",
                                        text="Raw Image",
                                        fg="cyan",
                                        font = LABELFRAME_TXT_FONT)
            img_labelframe_1.pack(fill=BOTH, expand=TRUE, side=LEFT)

            img_labelframe_2 = LabelFrame(image_tab,
                                        bg="black",
                                        text="Analysis Image",
                                        fg="cyan",
                                        font = LABELFRAME_TXT_FONT)
            img_labelframe_2.pack(fill=BOTH, expand=TRUE, side=RIGHT)

            a1 = Image.open(self.base_window.qualitative_analysis_0.analysis_result_folder + '/raw.jpg')
            a1_crop = a1.crop((x1-10, y1-10, x2+10, y2+10))
            crop_width, crop_height = a1_crop.size
            scale_percent = 100
            width = int(crop_width * scale_percent / 100)
            height = int(crop_height * scale_percent / 100)
            display_img = a1_crop.resize((width,height))
            a1_display = ImageTk.PhotoImage(display_img)
            a1_label = Label(img_labelframe_1, image=a1_display, bg="black")
            a1_label.image = a1_display
            a1_label.pack(fill=BOTH, expand=TRUE)

            a2 = Image.open(self.base_window.qualitative_analysis_0.analysis_result_folder + '/process.jpg')
            a2_crop = a2.crop((x1-10, y1-10, x2+10, y2+10))
            crop_width, crop_height = a2_crop.size
            scale_percent = 100
            width = int(crop_width * scale_percent / 100)
            height = int(crop_height * scale_percent / 100)
            display_img = a2_crop.resize((width,height))
            a2_display = ImageTk.PhotoImage(display_img)
            a2_label = Label(img_labelframe_2, image=a2_display, bg="black")
            a2_label.image = a2_display
            a2_label.pack(fill=BOTH, expand=TRUE)

            # In button frame
            self.finish_button = Button(self.button_frame,
                                    text = "FINISH",
                                    font = SWITCH_PAGE_BUTTON_FONT,
                                    width = 10,
                                    # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                    bg = 'dodger blue',
                                    fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.finish_clicked)
            self.finish_button.pack(ipady=10)

            self.base_window.update_idletasks()
            # ~ subprocess.call(["scrot", self.base_window.qualitative_analysis_1.analysis_result_folder + "/result_capture.jpg"])
            subprocess.call(["scrot", self.base_window.qualitative_analysis_0.result_folder_path + "/result_capture.jpg"])

            self.base_window.server_setting.check_status()
            if(self.base_window.server_setting.server_active==1):
                ftp = FTP(self.base_window.server_setting.ip_set, self.base_window.server_setting.user_set, self.base_window.server_setting.password_set, timeout=30)
                ftp.cwd(self.base_window.server_setting.path_set + '/TemplateBlank')
                
                localfolder = os.path.join(working_dir, 'template.xlsm')
                file = open(localfolder,'wb')
                ftp.retrbinary('RETR ' + 'template.xlsm', file.write)
                file.close()
                
                ftp.quit()
                
                #wb = load_workbook(working_dir + '/template.xlsm', keep_vba = True)
                wb = load_workbook(self.base_window.qualitative_analysis_2.id_file_path, keep_vba = True)
                sheet = wb.active
            else:
                wb = Workbook()
                sheet = wb.active

            font0 = Font(bold=False)
            font1 = Font(size='14', bold=True, color='00FF0000')
            font2 = Font(bold=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            sheet.protection.sheet = True
            sheet.protection.enable()

            sheet["B8"].protection = Protection(locked=False, hidden=False)

            for i in range(12,60):
                sheet["B"+str(i)].font = font0
                sheet["D"+str(i)].font = font0

            img = Img(working_dir + "/logo.png")
            img.height = 39
            img.width = 215
            img.anchor = 'B2'
            sheet.add_image(img)

            sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=6)
            sheet["B5"] = 'ANALYSIS RESULTS'
            sheet["B5"].font = font1
            sheet.cell(row=5,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
            #global foldername
            sheet["B7"] = 'Test name: ' + self.base_window.qualitative_analysis_0.experiment_name
            sheet["B7"].font = font2
            sheet['B8'] = 'Test technician name: ' + self.base_window.qualitative_analysis_0.user_name
            sheet["B8"].font = font2
            #global covid19dir_old
            sheet['B9'] = 'Date: ' + self.base_window.qualitative_analysis_0.create_time
            sheet["B9"].font = font2
            sheet['B60'] = 'Note:'
            sheet["B60"].font = font2
            sheet['B61'] = '+ N/A: No sample'
            sheet['B62'] = '+ N: Negative'
            sheet['C61'] = '+ P_L: Low copy'
            sheet['C62'] = '+ P_H: Positive'

            sheet.merge_cells(start_row=64, start_column=4, end_row=64, end_column=6)
            sheet.merge_cells(start_row=65, start_column=4, end_row=65, end_column=6)
            sheet['B64'] = 'Test Technician'
            sheet['B65'] = ''
            sheet['D64'] = 'Head of the Laboratory'
            sheet['D65'] = ''
            sheet["B64"].font = font2
            sheet["D64"].font = font2
            sheet["B64"].protection = Protection(locked=False, hidden=False)
            sheet["D64"].protection = Protection(locked=False, hidden=False)
            sheet.cell(row=64,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
            sheet.cell(row=65,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
            sheet.cell(row=64,column=4).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
            sheet.cell(row=65,column=4).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)

            for r in range(11,60):
                for c in range(2,7):
                    sheet.cell(row=r,column=c).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                    sheet.cell(row=r,column=c).border = thin_border

            sheet.column_dimensions['B'].width = 26
            sheet.column_dimensions['C'].width = 12
            sheet.column_dimensions['D'].width = 12
            sheet.column_dimensions['E'].width = 12
            sheet.column_dimensions['F'].width = 12

            sheet.row_dimensions[11].height = 40

            sheet['B11'] = 'Sample name'
            sheet["B11"].font = font2
            sheet["B11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
            sheet['C11'] = 'Sample positon'
            sheet["C11"].font = font2
            sheet["C11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
            sheet['D11'] = 'Spotcheck Result'
            sheet["D11"].font = font2
            sheet["D11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
            sheet['E11'] = 'Gel Result'
            sheet["E11"].font = font2
            sheet["E11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
            sheet['F11'] = 'Final Result'
            sheet["F11"].font = font2
            sheet["F11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')

            for i in range (12,60):
                if(i<20):
                    sheet['C'+str(i)] = str(chr(65+i-12)) + '1'
                if(i>=20 and i<28):
                    sheet['C'+str(i)] = str(chr(65+i-20)) + '2'
                if(i>=28 and i<36):
                    sheet['C'+str(i)] = str(chr(65+i-28)) + '3'
                if(i>=36 and i<44):
                    sheet['C'+str(i)] = str(chr(65+i-36)) + '4'
                if(i>=44 and i<52):
                    sheet['C'+str(i)] = str(chr(65+i-44)) + '5'
                if(i>=52):
                    sheet['C'+str(i)] = str(chr(65+i-52)) + '6'

            c1=-6
            c2=-5
            c3=-4
            c4=-3
            c5=-2
            c6=-1
            for i in range(0,8):
                c1=c1+6
                sheet['B'+str(i+12)] = self.base_window.qualitative_analysis_2.id_list[c1]
                if(self.base_window.qualitative_analysis_2.id_list[c1]=='N/A'):
                    sheet['D'+str(i+12)] = 'N/A'
                else:
                    if(float(result_label[c1]['text']) < self.pfi_value * NUM_1):
                        sheet['D'+str(i+12)] = 'N'
                        sheet['D'+str(i+12)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                    elif(float(result_label[c1]['text']) < self.pfi_value * NUM_2):
                            sheet['D'+str(i+12)] = 'P_L'
                            sheet['D'+str(i+12)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
                            sheet['D'+str(i+12)].font = font2
                            sheet['B'+str(i+12)].font = font2
                    else:
                        sheet['D'+str(i+12)] = 'P_H'
                        sheet['D'+str(i+12)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
                        sheet['D'+str(i+12)].font = font2
                        sheet['B'+str(i+12)].font = font2
                    
                sheet['E'+str(i+12)].protection = Protection(locked=False, hidden=False)
                sheet['F'+str(i+12)].protection = Protection(locked=False, hidden=False)

                c2=c2+6
                sheet['B'+str(i+20)] = self.base_window.qualitative_analysis_2.id_list[c2]
                if(self.base_window.qualitative_analysis_2.id_list[c2]=='N/A'):
                    sheet['D'+str(i+20)] = 'N/A'
                else:
                    if(float(result_label[c2]['text']) < self.pfi_value * NUM_1):
                        sheet['D'+str(i+20)] = 'N'
                        sheet['D'+str(i+20)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                    elif(float(result_label[c2]['text']) < self.pfi_value * NUM_2):
                        sheet['D'+str(i+20)] = 'P_L'
                        sheet['D'+str(i+20)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
                        sheet['D'+str(i+20)].font = font2
                        sheet['B'+str(i+20)].font = font2
                    else:
                        sheet['D'+str(i+20)] = 'P_H'
                        sheet['D'+str(i+20)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
                        sheet['D'+str(i+20)].font = font2
                        sheet['B'+str(i+20)].font = font2
                                                
                sheet['E'+str(i+20)].protection = Protection(locked=False, hidden=False)
                sheet['F'+str(i+20)].protection = Protection(locked=False, hidden=False)

                c3=c3+6
                sheet['B'+str(i+28)] = self.base_window.qualitative_analysis_2.id_list[c3]
                if(self.base_window.qualitative_analysis_2.id_list[c3]=='N/A'):
                    sheet['D'+str(i+28)] = 'N/A'
                else:
                    if(float(result_label[c3]['text']) < self.pfi_value * NUM_1):
                        sheet['D'+str(i+28)] = 'N'
                        sheet['D'+str(i+28)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                    elif(float(result_label[c3]['text']) < self.pfi_value * NUM_2):
                        sheet['D'+str(i+28)] = 'P_L'
                        sheet['D'+str(i+28)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
                        sheet['D'+str(i+28)].font = font2
                        sheet['B'+str(i+28)].font = font2
                    else:
                        sheet['D'+str(i+28)] = 'P_H'
                        sheet['D'+str(i+28)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
                        sheet['D'+str(i+28)].font = font2
                        sheet['B'+str(i+28)].font = font2
                        
                sheet['E'+str(i+28)].protection = Protection(locked=False, hidden=False)
                sheet['F'+str(i+28)].protection = Protection(locked=False, hidden=False)

                c4=c4+6
                sheet['B'+str(i+36)] = self.base_window.qualitative_analysis_2.id_list[c4]
                if(self.base_window.qualitative_analysis_2.id_list[c4]=='N/A'):
                    sheet['D'+str(i+36)] = 'N/A'
                else:
                    if(float(result_label[c4]['text']) < self.pfi_value * NUM_1):
                        sheet['D'+str(i+36)] = 'N'
                        sheet['D'+str(i+36)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                    elif(float(result_label[c4]['text']) < self.pfi_value * NUM_2):
                        sheet['D'+str(i+36)] = 'P_L'
                        sheet['D'+str(i+36)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
                        sheet['D'+str(i+36)].font = font2
                        sheet['B'+str(i+36)].font = font2
                    else:
                        sheet['D'+str(i+36)] = 'P_H'
                        sheet['D'+str(i+36)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
                        sheet['D'+str(i+36)].font = font2
                        sheet['B'+str(i+36)].font = font2
                   
                sheet['E'+str(i+36)].protection = Protection(locked=False, hidden=False)
                sheet['F'+str(i+36)].protection = Protection(locked=False, hidden=False)

                c5=c5+6
                sheet['B'+str(i+44)] = self.base_window.qualitative_analysis_2.id_list[c5]
                if(self.base_window.qualitative_analysis_2.id_list[c5]=='N/A'):
                    sheet['D'+str(i+44)] = 'N/A'
                else:
                    if(float(result_label[c5]['text']) < self.pfi_value * NUM_1):
                        sheet['D'+str(i+44)] = 'N'
                        sheet['D'+str(i+44)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                    elif(float(result_label[c5]['text']) < self.pfi_value * NUM_2):
                        sheet['D'+str(i+44)] = 'P_L'
                        sheet['D'+str(i+44)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
                        sheet['D'+str(i+44)].font = font2
                        sheet['B'+str(i+4)].font = font2
                    else:
                        sheet['D'+str(i+44)] = 'P_H'
                        sheet['D'+str(i+44)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
                        sheet['D'+str(i+44)].font = font2
                        sheet['B'+str(i+44)].font = font2

                sheet['E'+str(i+44)].protection = Protection(locked=False, hidden=False)
                sheet['F'+str(i+44)].protection = Protection(locked=False, hidden=False)

                c6=c6+6
                sheet['B'+str(i+52)] = self.base_window.qualitative_analysis_2.id_list[c6]
                if(self.base_window.qualitative_analysis_2.id_list[c6]=='N/A'):
                    sheet['D'+str(i+52)] = 'N/A'
                else:
                    if(result_label[c6]['text'] != 'B'):
                        if(float(result_label[c6]['text']) < self.pfi_value * NUM_1):
                            sheet['D'+str(i+52)] = 'N'
                            sheet['D'+str(i+52)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                        elif(float(result_label[c6]['text']) < self.pfi_value * NUM_2):
                            sheet['D'+str(i+52)] = 'P_L'
                            sheet['D'+str(i+52)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
                            sheet['D'+str(i+52)].font = font2
                            sheet['B'+str(i+52)].font = font2
                        else:
                            sheet['D'+str(i+52)] = 'P_H'
                            sheet['D'+str(i+52)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
                            sheet['D'+str(i+52)].font = font2
                            sheet['B'+str(i+52)].font = font2
                    
                sheet['E'+str(i+52)].protection = Protection(locked=False, hidden=False)
                sheet['F'+str(i+52)].protection = Protection(locked=False, hidden=False)
            
            sheet['D59'] = 'B'
            sheet['D'+str(i+52)].fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
            sheet.print_area = 'A1:G70'
            # ~ wb.save(self.base_window.qualitative_analysis_1.analysis_result_folder + '/' + self.base_window.qualitative_analysis_0.experiment_name + '.xlsm')
            wb.save(self.base_window.qualitative_analysis_0.result_folder_path +  '/' + self.base_window.qualitative_analysis_0.experiment_name + '.xlsm')
            
            
            
            # Report tab
            self.report_frame = ScrollableFrame2(report_tab)
            self.report_frame.pack(pady=5)
            
            wb = load_workbook(self.base_window.qualitative_analysis_0.result_folder_path +  '/' + self.base_window.qualitative_analysis_0.experiment_name + '.xlsm')
            sheet = wb.active
            
            sample_button_list = list(range(49))
            result_button_list = list(range(49))
            position_button_list = list(range(49))

            for i in range(0,48):
                sample_pos = 'B' + str(i+11)
                sample_button_list[i] = Button(self.report_frame.scrollable_frame,
                        fg = LABEL_TXT_COLOR,
                        font = LABEL_TXT_FONT,
                        text= sheet[sample_pos].value,
                        width=30,
                        bg = 'lavender',
                        borderwidth = 0)
                sample_button_list[i].grid(row=i, column=0, sticky=EW, padx=1, pady=1)

                position_pos = 'C' + str(i+11)
                position_button_list[i] = Button(self.report_frame.scrollable_frame,
                        fg = LABEL_TXT_COLOR,
                        font = LABEL_TXT_FONT,
                        text= sheet[position_pos].value,
                        width=10,
                        bg = 'lavender',
                        borderwidth = 0)
                position_button_list[i].grid(row=i, column=1, sticky=EW, padx=1, pady=1)

                result_pos = 'D' + str(i+11)
                result_button_list[i] = Button(self.report_frame.scrollable_frame,
                        fg = LABEL_TXT_COLOR,
                        font = LABEL_TXT_FONT,
                        text= sheet[result_pos].value,
                        width=22,
                        bg = 'lavender',
                        borderwidth = 0)
                result_button_list[i].grid(row=i, column=3, sticky=EW, padx=1, pady=1)
                
                if(i>0 and result_button_list[i]['text'] != 'N/A'):
                    if(result_button_list[i]['text'] == 'N'):
                        result_button_list[i]['bg'] = NEGATIVE_COLOR
                    elif(result_button_list[i]['text'] == 'P_L'):
                        result_button_list[i]['bg'] = LOW_COPY_COLOR
                    else:
                        result_button_list[i]['bg'] = POSITIVE_COLOR
                
            wb.close()
            
            
            
            
            
            if(os.path.exists(id_path + self.base_window.qualitative_analysis_2.id_file_name_label['text'] + '.xlsm')):
                try:
                    shutil.move(id_path + self.base_window.qualitative_analysis_2.id_file_name_label['text'] + '.xlsm', id_old_path)
                except:
                    pass

            self.automail_check()
            self.server_check()

            self.title_label['text'] = "ANALYSIS RESULTS"

            msg = messagebox.showinfo("","COMPLETED")

        else:
            self.base_window.forget_page()
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
            self.base_window.switch_page()
            self.base_window.main_menu.reset()

    def automail_check(self):
        if(self.base_window.email_setting.account_active == 1 and self.base_window.qualitative_analysis_2.automail_is_on == 1):
            shutil.make_archive(self.base_window.qualitative_analysis_0.result_folder_path,
                                format='zip',
                                root_dir = self.base_window.qualitative_analysis_0.result_folder_path)
            # ~ try:
            AutoMail(
                self.base_window.email_setting.email_address,
                self.base_window.email_setting.email_password,
                self.base_window.qualitative_analysis_2.recipient_email,
                "Spotcheck Result",
                "This is an automatic email from Spotcheck device.",
                self.base_window.qualitative_analysis_0.result_folder_path + '.zip',
                self.base_window.qualitative_analysis_0.result_folder_name).send()
            # ~ except:
                # ~ messagebox.showerror("","ERR 04")
                # ~ pass
    
    def server_check(self):
        self.base_window.server_setting.check_status()
        if(self.base_window.server_setting.server_active==1):
            try:
                ftp = FTP(self.base_window.server_setting.ip_set, self.base_window.server_setting.user_set, self.base_window.server_setting.password_set, timeout=30)
                ftp.cwd(self.base_window.server_setting.path_set + '/Processed_Data/Screening')
                file = open(self.base_window.qualitative_analysis_0.result_folder_path + '/' + self.base_window.qualitative_analysis_0.experiment_name + '.xlsm','rb')
                ftp.storbinary('STOR ' + self.base_window.qualitative_analysis_0.experiment_name + ".xlsm", file)
                ftp.quit()
            except Exception as e :
                messagebox.showwarning("There was an error while syncing the server",str(e))
                pass
    
    def finish_clicked(self):
        msg = messagebox.askquestion("","Do you want to go back to the main screen ?")
        if(msg=="yes"):
            self.base_window.forget_page()
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
            self.base_window.switch_page()
            self.base_window.main_menu.reset()


class QuantitativeAnalysisFrame3(QualitativeAnalysisFrame3):
    def __init__(self, container):
        super().__init__(container)
        self.title_label['text'] = "ANALYSIS"

    def serial_handle(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("1.Horizontal.TProgressbar", troughcolor ='grey85', background='green3')
        self.progressbar = ttk.Progressbar(self.work_frame,
                                    style="1.Horizontal.TProgressbar",
                                    length = 200,
                                    mode = 'determinate')
        self.progressbar.pack(ipadx=2, ipady=2)

        self.process_label = Label(self.work_frame,
                        bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
                        fg = 'black',
                        text = 'Processing...',
                        font = LABEL_TXT_FONT)
        self.process_label.pack(ipadx=2, ipady=2, anchor=N)

        ser.flushInput()
        ser.flushOutput()

        self.progressbar['value']=5
        self.base_window.update_idletasks()
        sleep(0.5)

        data_send = 'P'
        print("Data send:", data_send)
        ser.write(data_send.encode())

        receive_data = StringVar()
        count = 0
        bled_ready = 0
        while(receive_data != 'C'):
            if(ser.in_waiting>0):
                receive_data = ser.readline().decode('utf-8').rstrip()
                print("Data received:", receive_data)

                self.progressbar['value']=10
                self.base_window.update_idletasks()
                sleep(0.5)

                if(receive_data == 'C'):
                    self.progressbar['value']=20
                    self.base_window.update_idletasks()
                    sleep(0.5)

                    bled_ready = 1
                    break;
            else:
                sleep(1)
                count += 1
                if(count > 15):
                    msg = messagebox.showerror("ERR "+ str(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value),
                                                ERROR_LIST(ERROR_LIST['SERIAL_TIMEOUT_ERROR'].value).name,
                                                icon = "error")
                    break;

        if(bled_ready):
            self.progressbar['value']=45
            self.base_window.update_idletasks()
            try:
                camera_capture(self.base_window.quantitative_analysis_1.analysis_result_folder + '/raw.jpg')

                self.progressbar['value']=65
                self.base_window.update_idletasks()
            except Exception as e :
                msg = messagebox.showerror("ERR "+ str(ERROR_LIST['CAMERA_ERROR'].value),
                                        ERROR_LIST(ERROR_LIST['CAMERA_ERROR'].value).name,
                                        icon = "error")
                if(msg=='ok'):
                    self.base_window.destroy()

            self.result, self.image = Process_Image(self.base_window.quantitative_analysis_1.analysis_result_folder + '/raw.jpg').process(coefficient, mode=1, well_list=self.base_window.quantitative_analysis_2.id_list)

            self.x_result_list = list(range(48))
            self.y_result_list = list(range(48))
            self.concen_result_list = list(range(48))
            for i in range(0,48):
                if(self.base_window.quantitative_analysis_2.id_list[i] != 'N/A'):
                    self.y_result_list[i] = self.result[i]/self.base_window.system_check.threshold
                    self.x_result_list[i] = (self.y_result_list[i] - self.base_window.quantitative_analysis_1.b_value)/self.base_window.quantitative_analysis_1.a_value
                    self.concen_result_list[i] = round(10**self.x_result_list[i])
                    # ~ self.concen_result_list[i] = round((1 + (self.x_result_list[i] - round(self.x_result_list[i])))*(10**round(self.x_result_list[i])))
                    # ~ self.concen_result_list[i] = self.x_result_list[i]
            self.progressbar['value']=80
            self.base_window.update_idletasks()
            sleep(0.5)

            cv2.imwrite(self.base_window.quantitative_analysis_1.analysis_result_folder + '/process.jpg', self.image)

            sum_value = 0
            active_well_number = 0

            # Save analysis file
            wb = Workbook()
            sheet = wb.active

            sheet["A2"] = "A"
            sheet["A3"] = "B"
            sheet["A4"] = "C"
            sheet["A5"] = "D"
            sheet["A6"] = "E"
            sheet["A7"] = "F"
            sheet["A8"] = "G"
            sheet["A9"] = "H"
            sheet["B1"] = "1"
            sheet["C1"] = "2"
            sheet["D1"] = "3"
            sheet["E1"] = "4"
            sheet["F1"] = "5"
            sheet["G1"] = "6"
            for i in range(0,48):
                if(i<6):
                    pos = str(chr(65+i+1)) + "2"
                if(i>=6 and i<12):
                    pos = str(chr(65+i-5)) + "3"
                if(i>=12 and i<18):
                    pos = str(chr(65+i-11)) + "4"
                if(i>=18 and i<24):
                    pos = str(chr(65+i-17)) + "5"
                if(i>=24 and i<30):
                    pos = str(chr(65+i-23)) + "6"
                if(i>=30 and i<36):
                    pos = str(chr(65+i-29)) + "7"
                if(i>=36 and i<42):
                    pos = str(chr(65+i-35)) + "8"
                if(i>=42):
                    pos = str(chr(65+i-41)) + "9"

                if(self.base_window.quantitative_analysis_2.id_list[i] != 'N/A'):
                    sheet[pos] = round(self.result[i]/self.base_window.system_check.threshold,2)
                else:
                    sheet[pos] = "N/A"

            wb.save(self.base_window.quantitative_analysis_1.analysis_result_folder + "/analysis_value.xlsx")
            wb.close()

            self.progressbar['value']=95
            self.base_window.update_idletasks()
            sleep(0.5)

            self.progressbar['value']=100
            self.base_window.update_idletasks()
            sleep(0.5)

            self.progressbar.destroy()
            self.process_label.destroy()

            tab_control = ttk.Notebook(self.work_frame)
            result_tab = Frame(tab_control, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
            report_tab = Frame(tab_control, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
            image_tab = Frame(tab_control, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)

            tab_control.add(result_tab, text="Result")
            tab_control.add(report_tab, text="Report")
            tab_control.add(image_tab, text="Images")
            tab_control.grid(row=0, column=0, padx=0, pady=1, sticky=EW)

            self.check_result_frame = Frame(result_tab, bg=RESULT_TABLE_FRAME_BGD_COLOR)
            self.check_result_frame.grid(row=0, column=0, padx=0)

            # ~ Pmw.initialise(self.base_window)
            # ~ self.tooltip = list(range(48))

            result_label = list(range(48))
            r=0
            c=-1
            for i in range(0,48):
                c+=1
                if(c>5):
                    c=0
                    r+=1
                result_label[i] = Label(self.check_result_frame,
                                        width=12,
                                        height=3,
                                        # ~ bg = RESULT_LABEL_BGD_COLOR,
                                        font = RESULT_LABEL_TXT_FONT)

                # ~ self.tooltip[i] = Pmw.Balloon(self.base_window)
                # ~ self.tooltip[i].bind(result_label[i], self.base_window.quantitative_analysis_2.id_list[i])

                if(self.base_window.quantitative_analysis_2.id_list[i] != 'N/A'):
                    if(self.concen_result_list[i] <= 10000):
                        result_label[i]['bg'] = LOW_COPY_COLOR
                        result_label[i]['text'] = self.concen_result_list[i]
                    else:
                        result_label[i]['bg'] = POSITIVE_COLOR
                        result_label[i]['text'] = ">10000"
                    
                    if(round(self.result[i]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
                        result_label[i]['bg'] = NEGATIVE_COLOR
                        result_label[i]['text'] = '0'
                else:
                    result_label[i]['text'] = "N/A"
                    result_label[i]['bg'] = NA_COLOR
                result_label[i].grid(row=r,column=c, padx=1, pady=1)

            self.annotate_result_frame = Frame(result_tab, bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
            self.annotate_result_frame.grid(row=0, column=1, padx=2)

            nagative_label = Label(self.annotate_result_frame, bg=NEGATIVE_COLOR, width=4, height=2)
            nagative_label.grid(row=0, column=0, padx=20, pady=10)
            nagative_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="NEGATIVE (N)", height=2)
            nagative_text_label.grid(row=0, column=1, padx=20, pady=10)
            low_copy_label = Label(self.annotate_result_frame, bg=LOW_COPY_COLOR, width=4, height=2)
            low_copy_label.grid(row=1, column=0, padx=20, pady=10)
            low_copy_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="LOW COPY (P_L)", height=2)
            low_copy_text_label.grid(row=1, column=1, padx=20, pady=10)
            positive_label = Label(self.annotate_result_frame, bg=POSITIVE_COLOR, width=4, height=2)
            positive_label.grid(row=2, column=0, padx=20, pady=10)
            positive_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="POSITIVE (P_H)", height=2)
            positive_text_label.grid(row=2, column=1, padx=20, pady=10)
            na_label = Label(self.annotate_result_frame, bg=NA_COLOR, width=4, height=2)
            na_label.grid(row=3, column=0, padx=20, pady=10)
            na_copy_text_label = Label(self.annotate_result_frame, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text="NO SAMPLE (N/A)", height=2)
            na_copy_text_label.grid(row=3, column=1, padx=20, pady=10)

            #Report Tab
            self.report_frame = ScrollableFrame2(report_tab)
            self.report_frame.pack(pady=5)
            n=1
            sample_button_list = list(range(48))
            result_button_list = list(range(48))
            number_button_list = list(range(48))

            number0_button_list = Button(self.report_frame.scrollable_frame,
                            fg = LABEL_TXT_COLOR,
                            font = LABEL_TXT_FONT,
                            text= 'No',
                            width=2,
                            borderwidth = 0)
            number0_button_list.grid(row=0, column=0, sticky=EW, padx=1, pady=1)

            sample0_button_list = Button(self.report_frame.scrollable_frame,
                            fg = LABEL_TXT_COLOR,
                            font = LABEL_TXT_FONT,
                            text= 'Samples name',
                            width=30,
                            borderwidth = 0)
            sample0_button_list.grid(row=0, column=1, sticky=EW, padx=1, pady=1)

            result0_button_list = Button(self.report_frame.scrollable_frame,
                            fg = LABEL_TXT_COLOR,
                            font = LABEL_TXT_FONT,
                            text= 'Results',
                            width=5,
                            borderwidth = 0)
            result0_button_list.grid(row=0, column=2, sticky=EW, padx=1, pady=1)

            for i in range(0,48):
                if(self.base_window.quantitative_analysis_2.id_list[i] != 'N/A'):

                    number_button_list[i] = Button(self.report_frame.scrollable_frame,
                            fg = LABEL_TXT_COLOR,
                            font = LABEL_TXT_FONT,
                            bg='lavender',
                            text= str(n),
                            width=2,
                            borderwidth = 0)
                    
                    number_button_list[i].grid(row=n, column=0, sticky=EW, padx=1, pady=1)

                    sample_button_list[i] = Button(self.report_frame.scrollable_frame,
                            fg = LABEL_TXT_COLOR,
                            font = LABEL_TXT_FONT,
                            bg='lavender',
                            width=35,
                            borderwidth = 0,
                            anchor=W)
                    # ~ sample_button_list[i].pack(pady=1, ipady=8, fill=BOTH, expand=TRUE)
                    sample_button_list[i].grid(row=n, column=1, sticky=EW, padx=1, pady=1)

                    result_button_list[i] = Button(self.report_frame.scrollable_frame,
                            fg = LABEL_TXT_COLOR,
                            font = LABEL_TXT_FONT,
                            width=25,
                            borderwidth = 0)
                    result_button_list[i].grid(row=n, column=2,padx=1, pady=1)

                    n+=1

                    if(self.concen_result_list[i] <= 10000):
                        sample_button_list[i]['text'] = self.base_window.quantitative_analysis_2.id_list[i]
                        result_button_list[i]['text'] = self.concen_result_list[i]
                        result_button_list[i]['bg'] = LOW_COPY_COLOR
                    elif(self.concen_result_list[i] > 10000):
                        sample_button_list[i]['text'] = self.base_window.quantitative_analysis_2.id_list[i]
                        result_button_list[i]['text'] = ">10000"
                        result_button_list[i]['bg'] = POSITIVE_COLOR
                    
                    if(round(self.result[i]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
                        sample_button_list[i]['text'] = self.base_window.quantitative_analysis_2.id_list[i]
                        result_button_list[i]['text'] = '0'
                        result_button_list[i]['bg'] = NEGATIVE_COLOR



            img_labelframe_1 = LabelFrame(image_tab,
                                        bg="black",
                                        text="Raw Image",
                                        fg="cyan",
                                        font = LABELFRAME_TXT_FONT)
            img_labelframe_1.pack(fill=BOTH, expand=TRUE, side=LEFT)

            img_labelframe_2 = LabelFrame(image_tab,
                                        bg="black",
                                        text="Analysis Image",
                                        fg="cyan",
                                        font = LABELFRAME_TXT_FONT)
            img_labelframe_2.pack(fill=BOTH, expand=TRUE, side=RIGHT)

            a1 = Image.open(self.base_window.quantitative_analysis_1.analysis_result_folder + '/raw.jpg')
            a1_crop = a1.crop((x1-10, y1-10, x2+10, y2+10))
            crop_width, crop_height = a1_crop.size
            scale_percent = 100
            width = int(crop_width * scale_percent / 100)
            height = int(crop_height * scale_percent / 100)
            display_img = a1_crop.resize((width,height))
            a1_display = ImageTk.PhotoImage(display_img)
            a1_label = Label(img_labelframe_1, image=a1_display, bg="black")
            a1_label.image = a1_display
            a1_label.pack(fill=BOTH, expand=TRUE)

            a2 = Image.open(self.base_window.quantitative_analysis_1.analysis_result_folder + '/process.jpg')
            a2_crop = a2.crop((x1-10, y1-10, x2+10, y2+10))
            crop_width, crop_height = a2_crop.size
            scale_percent = 100
            width = int(crop_width * scale_percent / 100)
            height = int(crop_height * scale_percent / 100)
            display_img = a2_crop.resize((width,height))
            a2_display = ImageTk.PhotoImage(display_img)
            a2_label = Label(img_labelframe_2, image=a2_display, bg="black")
            a2_label.image = a2_display
            a2_label.pack(fill=BOTH, expand=TRUE)

            # In button frame
            self.finish_button = Button(self.button_frame,
                                    text = "FINISH",
                                    font = SWITCH_PAGE_BUTTON_FONT,
                                    width = 10,
                                    # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                    bg = 'dodger blue',
                                    fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.finish_clicked)
            self.finish_button.pack(ipady=10)

            self.base_window.update_idletasks()


            # ~ self.base_window.server_setting.check_status()
            # ~ if(self.base_window.server_setting.server_active==1):
                # ~ ftp = FTP(self.base_window.server_setting.ip_set, self.base_window.server_setting.user_set, self.base_window.server_setting.password_set, timeout=30)
                # ~ ftp.cwd(self.base_window.server_setting.path_set + '/TemplateBlank')
                
                # ~ localfolder = os.path.join(working_dir, 'template.xlsm')
                # ~ file = open(localfolder,'wb')
                # ~ ftp.retrbinary('RETR ' + 'template.xlsm', file.write)
                # ~ file.close()
                
                # ~ ftp.quit()
                
                # ~ #wb = load_workbook(working_dir + '/template.xlsm', keep_vba = True)
                # ~ wb = load_workbook(self.base_window.quantitative_analysis_2.id_file_path, keep_vba = True)
                # ~ sheet = wb.active
            # ~ else:
                # ~ wb = Workbook()
                # ~ sheet = wb.active
            
            wb = Workbook()
            sheet = wb.active
            
            font0 = Font(bold=False)
            font1 = Font(size='14', bold=True, color='00FF0000')
            font2 = Font(bold=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            sheet.protection.sheet = True
            sheet.protection.enable()

            sheet["B8"].protection = Protection(locked=False, hidden=False)

            for i in range(12,60):
                sheet["B"+str(i)].font = font0
                sheet["D"+str(i)].font = font0

            img = Img(working_dir + "/logo.png")
            img.height = 39
            img.width = 215
            img.anchor = 'B2'
            sheet.add_image(img)

            sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=6)
            sheet["B5"] = 'ANALYSIS RESULTS'
            sheet["B5"].font = font1
            sheet.cell(row=5,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
            #global foldername
            sheet["B7"] = 'Test name: ' + self.base_window.quantitative_analysis_1.program_name
            sheet["B7"].font = font2
            sheet['B8'] = 'Test technician name: ' + self.base_window.quantitative_analysis_0.user_name
            sheet["B8"].font = font2
            #global covid19dir_old
            sheet['B9'] = 'Date: '
            sheet["B9"].font = font2
            sheet['B60'] = 'Note:'
            sheet["B60"].font = font2
            sheet['B61'] = '+ N/A: No sample'
            sheet['B62'] = '+ N: Negative'
            sheet['C61'] = '+ P_L: Low copy'
            sheet['C62'] = '+ P_H: Positive'

            sheet.merge_cells(start_row=64, start_column=4, end_row=64, end_column=6)
            sheet.merge_cells(start_row=65, start_column=4, end_row=65, end_column=6)
            sheet['B64'] = 'Test Technician'
            sheet['B65'] = ''
            sheet['D64'] = 'Head of the Laboratory'
            sheet['D65'] = ''
            sheet["B64"].font = font2
            sheet["D64"].font = font2
            sheet["B64"].protection = Protection(locked=False, hidden=False)
            sheet["D64"].protection = Protection(locked=False, hidden=False)
            sheet.cell(row=64,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
            sheet.cell(row=65,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
            sheet.cell(row=64,column=4).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
            sheet.cell(row=65,column=4).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)

            for r in range(11,60):
                for c in range(2,7):
                    sheet.cell(row=r,column=c).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                    sheet.cell(row=r,column=c).border = thin_border

            sheet.column_dimensions['B'].width = 26
            sheet.column_dimensions['C'].width = 12
            sheet.column_dimensions['D'].width = 15
            sheet.column_dimensions['E'].width = 12
            sheet.column_dimensions['F'].width = 12

            sheet.row_dimensions[11].height = 40

            sheet['B11'] = 'Sample name'
            sheet["B11"].font = font2
            sheet["B11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
            sheet['C11'] = 'Sample position'
            sheet["C11"].font = font2
            sheet["C11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
            sheet['D11'] = 'Spotcheck result'
            sheet["D11"].font = font2
            sheet["D11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
            sheet['E11'] = 'Gel result'
            sheet["E11"].font = font2
            sheet["E11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
            sheet['F11'] = 'Final result'
            sheet["F11"].font = font2
            sheet["F11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')

            for i in range (12,60):
                if(i<20):
                    sheet['C'+str(i)] = str(chr(65+i-12)) + '1'
                if(i>=20 and i<28):
                    sheet['C'+str(i)] = str(chr(65+i-20)) + '2'
                if(i>=28 and i<36):
                    sheet['C'+str(i)] = str(chr(65+i-28)) + '3'
                if(i>=36 and i<44):
                    sheet['C'+str(i)] = str(chr(65+i-36)) + '4'
                if(i>=44 and i<52):
                    sheet['C'+str(i)] = str(chr(65+i-44)) + '5'
                if(i>=52):
                    sheet['C'+str(i)] = str(chr(65+i-52)) + '6'

            c1=-6
            c2=-5
            c3=-4
            c4=-3
            c5=-2
            c6=-1
            for i in range(0,8):
                c1=c1+6
                sheet['B'+str(i+12)] = self.base_window.quantitative_analysis_2.id_list[c1]
                if(self.base_window.quantitative_analysis_2.id_list[c1]=='N/A'):
                    sheet['D'+str(i+12)] = 'N/A'
                else:
                    
                    if(self.concen_result_list[c1] <= 10000):
                        sheet['D'+str(i+12)] = self.concen_result_list[c1]
                        sheet['D'+str(i+12)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
                        # ~ sheet['E'+str(i+12)] = self.concen_result_list[c1]
                        sheet['D'+str(i+12)].font = font2
                        sheet['B'+str(i+12)].font = font2
                    elif(self.concen_result_list[c1] > 10000):
                        sheet['D'+str(i+12)] = ">10000"
                        sheet['D'+str(i+12)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
                        # ~ sheet['E'+str(i+12)] = self.concen_result_list[c1]
                        sheet['D'+str(i+12)].font = font2
                        sheet['B'+str(i+12)].font = font2
                        
                    if(round(self.result[c1]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
                        sheet['D'+str(i+12)] = '0'
                        sheet['D'+str(i+12)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                        # ~ sheet['E'+str(i+12)] = '0'

                sheet['F'+str(i+12)].protection = Protection(locked=False, hidden=False)

                c2=c2+6
                sheet['B'+str(i+20)] = self.base_window.quantitative_analysis_2.id_list[c2]
                if(self.base_window.quantitative_analysis_2.id_list[c2]=='N/A'):
                    sheet['D'+str(i+20)] = 'N/A'
                else:
                    if(self.concen_result_list[c2] <= 10000):
                        sheet['D'+str(i+20)] = self.concen_result_list[c2]
                        sheet['D'+str(i+20)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
                        # ~ sheet['E'+str(i+20)] = self.concen_result_list[c2]
                        sheet['D'+str(i+20)].font = font2
                        sheet['B'+str(i+20)].font = font2
                    elif(self.concen_result_list[c2] > 10000):
                        sheet['D'+str(i+20)] =  ">10000"
                        sheet['D'+str(i+20)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
                        # ~ sheet['E'+str(i+20)] = self.concen_result_list[c2]
                        sheet['D'+str(i+20)].font = font2
                        sheet['B'+str(i+20)].font = font2
                        
                    if(round(self.result[c2]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
                        sheet['D'+str(i+20)] = '0'
                        sheet['D'+str(i+20)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                        # ~ sheet['E'+str(i+20)] = '0'

                sheet['F'+str(i+20)].protection = Protection(locked=False, hidden=False)

                c3=c3+6
                sheet['B'+str(i+28)] = self.base_window.quantitative_analysis_2.id_list[c3]
                if(self.base_window.quantitative_analysis_2.id_list[c3]=='N/A'):
                    sheet['D'+str(i+28)] = 'N/A'
                else:
                    if(self.concen_result_list[c3] <= 10000):
                        sheet['D'+str(i+28)] = self.concen_result_list[c3]
                        sheet['D'+str(i+28)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
                        # ~ sheet['E'+str(i+28)] = self.concen_result_list[c3]
                        sheet['D'+str(i+28)].font = font2
                        sheet['B'+str(i+28)].font = font2
                    elif(self.concen_result_list[c3] > 10000):
                        sheet['D'+str(i+28)] = ">10000"
                        sheet['D'+str(i+28)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
                        # ~ sheet['E'+str(i+28)] = self.concen_result_list[c3]
                        sheet['D'+str(i+28)].font = font2
                        sheet['B'+str(i+28)].font = font2
                    if(round(self.result[c3]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
                        sheet['D'+str(i+28)] = '0'
                        sheet['D'+str(i+28)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                        # ~ sheet['E'+str(i+28)] = '0'

                sheet['F'+str(i+28)].protection = Protection(locked=False, hidden=False)

                c4=c4+6
                sheet['B'+str(i+36)] = self.base_window.quantitative_analysis_2.id_list[c4]
                if(self.base_window.quantitative_analysis_2.id_list[c4]=='N/A'):
                    sheet['D'+str(i+36)] = 'N/A'
                else:
                    
                    if(self.concen_result_list[c4] <= 10000):
                        sheet['D'+str(i+36)] = self.concen_result_list[c4]
                        sheet['D'+str(i+36)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
                        # ~ sheet['E'+str(i+36)] = self.concen_result_list[c4]
                        sheet['D'+str(i+36)].font = font2
                        sheet['B'+str(i+36)].font = font2
                    elif(self.concen_result_list[c4] > 10000):
                        sheet['D'+str(i+36)] = ">10000"
                        sheet['D'+str(i+36)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
                        # ~ sheet['E'+str(i+36)] = self.concen_result_list[c4]
                        sheet['D'+str(i+36)].font = font2
                        sheet['B'+str(i+36)].font = font2
                    if(round(self.result[c4]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
                        sheet['D'+str(i+36)] = '0'
                        sheet['D'+str(i+36)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                        # ~ sheet['E'+str(i+36)] = '0'

                sheet['F'+str(i+36)].protection = Protection(locked=False, hidden=False)

                c5=c5+6
                sheet['B'+str(i+44)] = self.base_window.quantitative_analysis_2.id_list[c5]
                if(self.base_window.quantitative_analysis_2.id_list[c5]=='N/A'):
                    sheet['D'+str(i+44)] = 'N/A'
                else:
                    if(self.concen_result_list[c5] <= 10000):
                        sheet['D'+str(i+44)] = ">10000"
                        sheet['D'+str(i+44)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
                        # ~ sheet['E'+str(i+44)] = self.concen_result_list[c5]
                        sheet['D'+str(i+44)].font = font2
                        sheet['B'+str(i+44)].font = font2
                    elif(self.concen_result_list[c5] > 10000):
                        sheet['D'+str(i+44)] = self.concen_result_list[c5]
                        sheet['D'+str(i+44)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
                        # ~ sheet['E'+str(i+44)] = self.concen_result_list[c5]
                        sheet['D'+str(i+44)].font = font2
                        sheet['B'+str(i+44)].font = font2
                        
                    if(round(self.result[c5]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
                        sheet['D'+str(i+44)] = '0'
                        sheet['D'+str(i+44)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                        # ~ sheet['E'+str(i+44)] = '0'

                sheet['F'+str(i+44)].protection = Protection(locked=False, hidden=False)

                c6=c6+6
                sheet['B'+str(i+52)] = self.base_window.quantitative_analysis_2.id_list[c6]
                if(self.base_window.quantitative_analysis_2.id_list[c6]=='N/A'):
                    sheet['D'+str(i+52)] = 'N/A'
                else:
                    if(self.concen_result_list[c6] <= 10000):
                        sheet['D'+str(i+52)] = self.concen_result_list[c6]
                        sheet['D'+str(i+52)].fill = PatternFill(start_color='00FFCCFF', end_color='00FFCCFF', fill_type='solid')
                        # ~ sheet['E'+str(i+52)] = self.concen_result_list[c6]
                        sheet['D'+str(i+52)].font = font2
                        sheet['B'+str(i+52)].font = font2
                    elif(self.concen_result_list[c6] > 10000):
                        sheet['D'+str(i+52)] = ">10000"
                        sheet['D'+str(i+52)].fill = PatternFill(start_color='00FF9999', end_color='00FF9999', fill_type='solid')
                        # ~ sheet['E'+str(i+52)] = self.concen_result_list[c6]
                        sheet['D'+str(i+52)].font = font2
                        sheet['B'+str(i+52)].font = font2
                        
                    if(round(self.result[c6]/self.base_window.system_check.threshold,2) <= self.base_window.quantitative_analysis_1.n_base_value):
                        sheet['D'+str(i+52)] = '0'
                        sheet['D'+str(i+52)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
                        # ~ sheet['E'+str(i+52)] = '0'

                sheet['F'+str(i+52)].protection = Protection(locked=False, hidden=False)

            sheet.print_area = 'A1:G70'
            wb.save(self.base_window.quantitative_analysis_1.result_folder_path + '/' + self.base_window.quantitative_analysis_0.experiment_name + '.xlsm')

            if(os.path.exists(id_path + self.base_window.quantitative_analysis_2.id_file_name_label['text'] + '.xlsm')):
                try:
                    shutil.move(id_path + self.base_window.quantitative_analysis_2.id_file_name_label['text'] + '.xlsm', id_old_path)
                except:
                    pass

            self.automail_check()
            self.server_check()

            self.title_label['text'] = "ANALYSIS RESULTS"

            msg = messagebox.showinfo("","COMPLETED")
            self.base_window.update_idletasks()
            subprocess.call(["scrot", self.base_window.quantitative_analysis_1.result_folder_path + "/result_capture.jpg"])

        else:
            self.base_window.forget_page()
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
            self.base_window.switch_page()
            self.base_window.main_menu.reset()

    def finish_clicked(self):
        msg = messagebox.askquestion("","Do you want to go back to the main screen ?")
        if(msg=="yes"):
            self.base_window.forget_page()
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_option)
            self.base_window.switch_page()
            self.base_window.main_menu.reset()

    def automail_check(self):
        if(self.base_window.email_setting.account_active == 1 and self.base_window.quantitative_analysis_2.automail_is_on == 1):
            shutil.make_archive(self.base_window.quantitative_analysis_1.result_folder_path,
                                format='zip',
                                root_dir = self.base_window.quantitative_analysis_1.result_folder_path)
            # ~ try:
            AutoMail(
                self.base_window.email_setting.email_address,
                self.base_window.email_setting.email_password,
                self.base_window.quantitative_analysis_2.recipient_email,
                "Spotcheck Result",
                "This is an automatic email from Spotcheck device.",
                self.base_window.quantitative_analysis_1.result_folder_path + '.zip',
                self.base_window.quantitative_analysis_1.result_folder_name).send()
            # ~ except:
                # ~ messagebox.showerror("","ERR 04")
                # ~ pass
                
                
    def server_check(self):
        self.base_window.server_setting.check_status()
        if(self.base_window.server_setting.server_active==1):
            try:
                ftp = FTP(self.base_window.server_setting.ip_set, self.base_window.server_setting.user_set, self.base_window.server_setting.password_set, timeout=30)
                ftp.cwd(self.base_window.server_setting.path_set + '/Processed_Data/Quantitative')
                file = open(self.base_window.quantitative_analysis_1.result_folder_path + '/' + self.base_window.quantitative_analysis_0.experiment_name + '.xlsm','rb')
                ftp.storbinary('STOR ' + self.base_window.quantitative_analysis_0.experiment_name + ".xlsm", file)
                ftp.quit()
            except Exception as e :
                messagebox.showwarning("There was an error while syncing the server",str(e))
                pass
class QualitativeAnalysisFrame2(Frame):
    def __init__(self, container):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container

        self.id_list = list(range(48))

        self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=5, fill=X)
        self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.work_frame.pack_propagate(0)
        self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X, expand=TRUE)

        # In title frame
        self.title_label = Label(self.title_frame,
                                text = "ANALYSIS",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

        # In work frame
        id_load_frame = LabelFrame(self.work_frame,
                                    text="ID File",
                                    fg = LABEL_FRAME_TXT_COLOR,
                                    font = LABEL_FRAME_TXT_FONT,
                                    bg = LABEL_FRAME_BGD_COLOR)

        id_load_frame.grid(row=0, column=0, pady=146, padx=50)

        self.id_pos_frame = Frame(self.work_frame,
                            bg = LABEL_FRAME_BGD_COLOR)

        self.id_pos_frame.grid(row=0, column=1, padx=50)

        self.id_file_name_label = Label(id_load_frame,
                                   bg = LABEL_BGD_COLOR,
                                   fg = LABEL_TXT_COLOR,
                                   font = ('Helvetica', 13))
        self.id_file_name_label.pack(expand=TRUE, fill=BOTH, pady=10, padx=10)

        self.load_button = Button(id_load_frame,
                                  text = "Load",
                                  bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
                                  fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
                                  font = MAIN_FUCNTION_BUTTON_FONT,
                                  width = 15,
                                  height = 3,
                                  borderwidth = 0,
                                  command = self.load_clicked)
        self.load_button.pack(expand=TRUE, side=LEFT)

        self.create_button = Button(id_load_frame,
                                  text = "Create",
                                  bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
                                  fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
                                  font = MAIN_FUCNTION_BUTTON_FONT,
                                  width = 15,
                                  height = 3,
                                  borderwidth = 0,
                                  command = self.create_clicked)
        self.create_button.pack(expand=TRUE, side=RIGHT)



        # In button frame
        self.back_button = Button(self.button_frame,
                                text = "Back",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.back_clicked)
        self.back_button.pack(side=LEFT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)

        self.next_button = Button(self.button_frame,
                                text = "Next",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.next_clicked)
        self.next_button.pack(side=RIGHT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)
        
    def back_clicked(self):
        self.base_window.frame_list.remove(self.base_window.system_check)
        del self.base_window.system_check
        self.base_window.system_check = SystemCheckFrame(self.base_window)
        self.base_window.frame_list.append(self.base_window.system_check)

        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
        self.base_window.switch_page()

    def next_clicked(self):
        try:
             self.email_label_frame.place_forget()
        except:
            pass

        if(self.id_file_name_label['text'] == ""):
            messagebox.showwarning("","You haven't loaded the ID file.")
        else:
            self.base_window.server_setting.check_status()
            self.base_window.email_setting.check_status()
            if(self.base_window.email_setting.account_active == 1):
                msg = messagebox.askquestion("", "Do you want the device to automatically email the results ?")
                if(msg=='yes'):
                    self.email_label_frame = LabelFrame(self.work_frame,
                                                        width = 100,
                                                        height = 50,
                                                        text = "Recipient email",
                                                        bg = 'dodger blue')
                    self.email_label_frame.place(x=200, y=150)

                    self.email_entry = Entry(self.email_label_frame, width=30, justify='right', font=('Courier',14))
                    self.email_entry.pack()

                    self.ok_button = Button(self.email_label_frame,
                                text = "OK",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                width = 5,
                                height = 2,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.ok_clicked)
                    self.ok_button.pack(side=LEFT, padx=20, pady=2, ipady=10, ipadx=20)

                    self.cancel_button = Button(self.email_label_frame,
                                text = "Cancel",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                width = 5,
                                height = 2,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.cancel_clicked)
                    self.cancel_button.pack(side=RIGHT, padx=20, pady=2, ipady=10, ipadx=20)

                else:
                    self.automail_is_on = 0
                    self.base_window.forget_page()
                    self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_3)
                    self.base_window.switch_page()
                    self.base_window.qualitative_analysis_3.serial_handle()
            else:
                self.automail_is_on = 0
                self.base_window.forget_page()
                self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_3)
                self.base_window.switch_page()
                self.base_window.qualitative_analysis_3.serial_handle()

    def ok_clicked(self):
        if(self.email_entry.get() == ''):
            messagebox.showwarning("","Please enter the recipient email !")
        else:
            addressToVerify = self.email_entry.get()
            # ~ match = re.match('^[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})$', addressToVerify)
            # ~ if (match == None):
                # ~ messagebox.showerror("","Email syntax error")
            # ~ else:
            self.recipient_email = addressToVerify
            self.automail_is_on = 1
            self.base_window.forget_page()
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_3)
            self.base_window.switch_page()
            self.base_window.qualitative_analysis_3.serial_handle()

    def cancel_clicked(self):
        self.email_label_frame.place_forget()

    def load_clicked(self):
        self.server_check()
        path = filedialog.askopenfilename(initialdir=id_path, filetypes=[('Excel file','*.xlsm *.xlsx *.xls')])
        self.id_file_path = path
        if path is not None:
            try:
                wb = load_workbook(path)
                sheet = wb.active
                # ~ for i in range(0,48):
                    # ~ pos = 'B' + str(i+12)
                    # ~ self.id_list[i] = sheet[pos].value
                    
                self.id_list[0] = sheet["B12"].value
                self.id_list[1] = sheet["B20"].value
                self.id_list[2] = sheet["B28"].value
                self.id_list[3] = sheet["B36"].value
                self.id_list[4] = sheet["B44"].value
                self.id_list[5] = sheet["B52"].value
                
                self.id_list[6] = sheet["B13"].value
                self.id_list[7] = sheet["B21"].value
                self.id_list[8] = sheet["B29"].value
                self.id_list[9] = sheet["B37"].value
                self.id_list[10] = sheet["B45"].value
                self.id_list[11] = sheet["B53"].value
                
                self.id_list[12] = sheet["B14"].value
                self.id_list[13] = sheet["B22"].value
                self.id_list[14] = sheet["B30"].value
                self.id_list[15] = sheet["B38"].value
                self.id_list[16] = sheet["B46"].value
                self.id_list[17] = sheet["B54"].value
                
                self.id_list[18] = sheet["B15"].value
                self.id_list[19] = sheet["B23"].value
                self.id_list[20] = sheet["B31"].value
                self.id_list[21] = sheet["B39"].value
                self.id_list[22] = sheet["B47"].value
                self.id_list[23] = sheet["B55"].value
                
                self.id_list[24] = sheet["B16"].value
                self.id_list[25] = sheet["B24"].value
                self.id_list[26] = sheet["B32"].value
                self.id_list[27] = sheet["B40"].value
                self.id_list[28] = sheet["B48"].value
                self.id_list[29] = sheet["B56"].value
                
                self.id_list[30] = sheet["B17"].value
                self.id_list[31] = sheet["B25"].value
                self.id_list[32] = sheet["B33"].value
                self.id_list[33] = sheet["B41"].value
                self.id_list[34] = sheet["B49"].value
                self.id_list[35] = sheet["B57"].value
                
                self.id_list[36] = sheet["B18"].value
                self.id_list[37] = sheet["B26"].value
                self.id_list[38] = sheet["B34"].value
                self.id_list[39] = sheet["B42"].value
                self.id_list[40] = sheet["B50"].value
                self.id_list[41] = sheet["B58"].value
                
                self.id_list[42] = sheet["B19"].value
                self.id_list[43] = sheet["B27"].value
                self.id_list[44] = sheet["B35"].value
                self.id_list[45] = sheet["B43"].value
                self.id_list[46] = sheet["B51"].value
                self.id_list[47] = sheet["B59"].value

                tmp = 0
                for i in range(len(path)):
                    if(path[i]=='/'):
                        tmp = i+1
                file_name = path[tmp:(len(path)-5)]

                self.id_file_name_label['text'] = file_name
                self.id_file_name_label['bg'] = 'lawn green'

                try:
                    for i in range(0,48):
                        self.id_label[i].destroy()
                except:
                    pass

                # ~ Pmw.initialise(self.base_window)
                # ~ self.tooltip = list(range(48))

                self.id_label = list(range(48))
                r=0
                c=-1
                for i in range(0,48):
                    c+=1
                    if(c>5):
                        c=0
                        r+=1
                    self.id_label[i] = Label(self.id_pos_frame,
                                            width=6,
                                            height=3,
                                            text = self.id_list[i],
                                            # ~ bg = RESULT_LABEL_BGD_COLOR,
                                            font = RESULT_LABEL_TXT_FONT)

                    # ~ self.tooltip[i] = Pmw.Balloon(self.base_window)
                    # ~ self.tooltip[i].bind(self.id_label[i], self.id_list[i])

                    if(self.id_list[i] != 'N/A'):
                        self.id_label[i]['bg'] = "lawn green"
                    else:
                        self.id_label[i]['bg'] = "grey80"

                    self.id_label[i].grid(row=r, column=c, padx=1, pady=1)
                self.id_label[47]['bg'] = "blue"
                
                msg = messagebox.askokcancel("","Now you can put samples in and press Next to analyze.")

            except:
                pass

    def create_clicked(self):
        del self.base_window.id_create
        self.base_window.id_create = IDCreateFrame(self.base_window)
        self.base_window.frame_list.append(self.base_window.id_create)

        self.base_window.id_create.direct_create = 1
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.id_create)
        self.base_window.switch_page()
    
    def server_check(self):
        # ~ #Check if the server is connected
        self.base_window.email_setting.check_status()
        if(self.base_window.server_setting.server_active==1):
            try:
                #FTP:
                print("Connect to " + self.base_window.server_setting.ip_set + '/' + self.base_window.server_setting.path_set + '...')
                ftp = FTP(self.base_window.server_setting.ip_set, self.base_window.server_setting.user_set, self.base_window.server_setting.password_set, timeout=15)
                print("Done")
                ftp.cwd(self.base_window.server_setting.path_set + '/Unprocessed_Data')
                ftp_files = ftp.nlst()
                for ftp_file in ftp_files:
                    if(os.path.exists(id_path + ftp_file)):
                        pass
                    elif(os.path.exists(id_old_path + ftp_file)):
                        pass
                    else:
                        local_folder = os.path.join(id_path, ftp_file)
                        file = open(local_folder,'wb')
                        ftp.retrbinary('RETR ' + ftp_file, file.write)
                        file.close()
                        print(ftp_file, "download done!")
                ftp.quit()
                
            except Exception as e:
                messagebox.showerror("There was an error while syncing the server",str(e))
                pass


class QuantitativeAnalysisFrame2(QualitativeAnalysisFrame2):
    def __init__(self, container):
        super().__init__(container)
        self.title_label['text'] = "ANALYSIS"

    def back_clicked(self):
        self.base_window.frame_list.remove(self.base_window.system_check)
        del self.base_window.system_check
        self.base_window.system_check = SystemCheckFrame(self.base_window)
        self.base_window.frame_list.append(self.base_window.system_check)

        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_1)
        self.base_window.switch_page()

    def next_clicked(self):
        try:
             self.email_label_frame.place_forget()
        except:
            pass

        if(self.id_file_name_label['text'] == ""):
            messagebox.showwarning("","You haven't loaded the ID file.")
        else:
            self.base_window.email_setting.check_status()
            if(self.base_window.email_setting.account_active == 1):
                msg = messagebox.askquestion("", "Do you want the device to automatically email the results ?")
                if(msg=='yes'):
                    self.email_label_frame = LabelFrame(self.work_frame,
                                                        width = 100,
                                                        height = 50,
                                                        text = "Recipient email",
                                                        bg = 'dodger blue')
                    self.email_label_frame.place(x=200, y=150)

                    self.email_entry = Entry(self.email_label_frame, width=30, justify='right', font=('Courier',14))
                    self.email_entry.pack()

                    self.ok_button = Button(self.email_label_frame,
                                text = "OK",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                width = 5,
                                height = 2,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.ok_clicked)
                    self.ok_button.pack(side=LEFT, padx=20, pady=2, ipady=10, ipadx=20)

                    self.cancel_button = Button(self.email_label_frame,
                                text = "Cancel",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                width = 5,
                                height = 2,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.cancel_clicked)
                    self.cancel_button.pack(side=RIGHT, padx=20, pady=2, ipady=10, ipadx=20)

                else:
                    self.automail_is_on = 0
                    self.base_window.forget_page()
                    self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_3)
                    self.base_window.switch_page()
                    self.base_window.quantitative_analysis_3.serial_handle()
            else:
                self.automail_is_on = 0
                self.base_window.forget_page()
                self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_3)
                self.base_window.switch_page()
                self.base_window.quantitative_analysis_3.serial_handle()

    def ok_clicked(self):
        if(self.email_entry.get() == ''):
            messagebox.showwarning("","Please enter the recipient email !")
        else:
            addressToVerify = self.email_entry.get()
            # ~ match = re.match('^[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})$', addressToVerify)
            # ~ if (match == None):
                # ~ messagebox.showerror("","Email syntax error")
            # ~ else:
            self.recipient_email = addressToVerify
            self.automail_is_on = 1
            self.base_window.forget_page()
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_3)
            self.base_window.switch_page()
            self.base_window.quantitative_analysis_3.serial_handle()

    def cancel_clicked(self):
        self.email_label_frame.place_forget()


    def create_clicked(self):
        del self.base_window.id_create
        self.base_window.id_create = IDCreateFrame(self.base_window)
        self.base_window.frame_list.append(self.base_window.id_create)

        self.base_window.id_create.direct_create = 2
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.id_create)
        self.base_window.switch_page()

class QualitativeAnalysisFrame1(Frame):
    def __init__(self, container):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container

        self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=5, fill=X)
        self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.work_frame.pack_propagate(0)
        self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X, expand=TRUE)

        # In title frame
        self.title_label = Label(self.title_frame,
                                text = "ANALYSIS",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

        # In work frame
        self.program_frame = ScrollableFrame(self.work_frame)
        self.program_frame.grid(row=0, column=0, pady=20)

        self.info_labelframe = LabelFrame(self.work_frame,
                                text = "Information",
                                font = LABELFRAME_TXT_FONT,
                                bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.info_labelframe.grid(row=0, column=1, rowspan=2, ipadx=10, ipady=5, padx=10, pady=28)

        experiment_name_label = Label(self.info_labelframe,
                            bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
                            text = "Kit name:",
                            fg = LABEL_TXT_COLOR,
                            font = LABEL_TXT_FONT)
        experiment_name_label.grid(row=0, column=0, padx=10, sticky=E)


        self.experiment_name_text = Text(self.info_labelframe,
                            width = 30,
                            height = 1,
                            bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
                            fg = LABEL_TXT_COLOR,
                            font = LABEL_TXT_FONT,
                            state = 'disabled')
        self.experiment_name_text.grid(row=0, column=1, padx=5, pady=10)

        # ~ user_name_label = Label(self.info_labelframe,
                            # ~ bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
                            # ~ text = "User name:",
                            # ~ fg = LABEL_TXT_COLOR,
                            # ~ font = LABEL_TXT_FONT)
        # ~ user_name_label.grid(row=1, column=0, padx=10, sticky=E)

        # ~ self.user_name_text= Text(self.info_labelframe,
                            # ~ width = 30,
                            # ~ height = 1,
                            # ~ bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
                            # ~ fg = LABEL_TXT_COLOR,
                            # ~ font = LABEL_TXT_FONT,
                            # ~ state = 'disabled')
        # ~ self.user_name_text.grid(row=1, column=1, padx=5, pady=10)

        # ~ pfi_label = Label(self.info_labelframe,
                            # ~ bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
                            # ~ text = "PFi value:",
                            # ~ fg = LABEL_TXT_COLOR,
                            # ~ font = LABEL_TXT_FONT)
        # ~ pfi_label.grid(row=2, column=0, padx=10, pady=10, sticky=NE)

        # ~ self.pfi_text = Text(self.info_labelframe,
                            # ~ width = 30,
                            # ~ height = 1,
                            # ~ bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
                            # ~ fg = LABEL_TXT_COLOR,
                            # ~ font = LABEL_TXT_FONT,
                            # ~ state = 'disabled')
        # ~ self.pfi_text.grid(row=2, column=1, padx=5, pady=10)

        status_label = Label(self.info_labelframe,
                            bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
                            text = "Comment:",
                            fg = LABEL_TXT_COLOR,
                            font = LABEL_TXT_FONT)
        status_label.grid(row=3, column=0, padx=10, pady=10, sticky=NE)

        self.status_text = Text(self.info_labelframe,
                            width = 30,
                            height = 12,
                            bg = MAIN_FUNCTION_FRAME_BGD_COLOR,
                            fg = LABEL_TXT_COLOR,
                            font = LABEL_TXT_FONT,
                            state = 'disabled')
        self.status_text.grid(row=3, column=1, padx=5, pady=10)

        # In button frame
        self.back_button = Button(self.button_frame,
                                text = "Back",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.back_clicked)
        self.back_button.pack(side=LEFT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)

        self.next_button = Button(self.button_frame,
                                text = "Next",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.next_clicked)
        self.next_button.pack(side=RIGHT, padx=0, pady=0, ipady=10, ipadx=30, anchor=E)

    def back_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
        self.base_window.switch_page()
    def next_clicked(self):
        self.program_name = self.experiment_name_text.get("1.0","end-1c")
        # ~ self.program_name = self.base_window.qualitative_analysis_0.experiment_name
        if(len(self.program_name) != 0):
            msg = messagebox.askokcancel("","Please make sure no sample is placed in the device !")
            if(msg == True):
                if os.path.exists(results_qualitative_path + self.program_name):
                    self.program_path = results_qualitative_path + self.program_name
                else:
                    self.program_path = os.path.join(results_qualitative_path, self.program_name)
                    os.mkdir(self.program_path)

                self.create_time = strftime(" %y-%m-%d %H.%M.%S")
                self.result_folder_name = self.base_window.qualitative_analysis_0.experiment_name + self.create_time
                self.result_folder_path = os.path.join(self.program_path + '/', self.result_folder_name)
                os.mkdir(self.result_folder_path)
                print(self.result_folder_path)

                self.system_check_folder = os.path.join(self.result_folder_path, "System_Check")
                os.mkdir(self.system_check_folder)

                self.analysis_result_folder = os.path.join(self.result_folder_path, "Analysis Results")
                os.mkdir(self.analysis_result_folder)

                self.base_window.system_check.mode_check = 1
                self.base_window.forget_page()
                self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
                self.base_window.switch_page()
                self.base_window.system_check.serial_handle()

        else:
            messagebox.showwarning("","Please select the kit !")

    def load_program(self):
        try:
            for i in range(len(self.program_button)):
                self.program_button[i].destroy()
        except:
            pass
        self.program_button = list(range(100))
        for file in os.listdir(programs_qualitative_path):
            self.program_button[os.listdir(programs_qualitative_path).index(file)] = Button(self.program_frame.scrollable_frame,
                                    text=file[:(len(file)-5)],
                                    font = PROGRAM_BUTTON_TXT_FONT,
                                    bg = PROGRAM_BUTTON_BGD_COLOR,
                                    fg = PROGRAM_BUTTON_TXT_COLOR,
                                    width = 51,
                                    borderwidth = 0)
            self.program_button[os.listdir(programs_qualitative_path).index(file)]['command'] = partial(self.program_clicked, os.listdir(programs_qualitative_path).index(file))
            self.program_button[os.listdir(programs_qualitative_path).index(file)].pack(pady=2, ipady=5, fill=BOTH, expand=TRUE)

    def program_clicked(self, button_index):
        wb = load_workbook(programs_qualitative_path + self.program_button[button_index]['text'] + ".xlsx")
        sheet = wb.active

        self.program_base_value = float(sheet["E2"].value)

        self.experiment_name_text['state'] = "normal"
        self.experiment_name_text.delete('1.0',END)
        try:
            self.experiment_name_text.insert('1.0', self.program_button[button_index]['text'])
        except:
            pass
        self.experiment_name_text['state'] = "disabled"

        # ~ self.user_name_text['state'] = "normal"
        # ~ self.user_name_text.delete('1.0',END)
        # ~ try:
            # ~ self.user_name_text.insert('1.0', sheet["C2"].value)
        # ~ except:
            # ~ pass
        # ~ self.user_name_text['state'] = "disabled"

        # ~ self.pfi_text['state'] = "normal"
        # ~ self.pfi_text.delete('1.0',END)
        # ~ try:
            # ~ self.pfi_text.insert('1.0', sheet["E2"].value)
        # ~ except:
            # ~ pass
        # ~ self.pfi_text['state'] = "disabled"

        self.comment_text['state'] = "normal"
        self.comment_text.delete('1.0',END)
        try:
            # self.comment_text.insert('1.0', sheet["C3"].value)
            self.comment_text.insert(END, sheet["B2"].value + ': ')
            self.comment_text.insert(END, sheet["C2"].value + '\n')
            self.comment_text.insert(END, sheet["B3"].value + ':')
            self.comment_text.insert(END, sheet["C3"].value + '\n')
            self.comment_text.insert(END, sheet["B4"].value + ':')
            self.comment_text.insert(END, sheet["C4"].value + '\n')
            self.comment_text.insert(END, sheet["B5"].value + ':')
            self.comment_text.insert(END, sheet["C5"].value + '\n')
        except:
            pass
        self.comment_text['state'] = "disabled"

        wb.close()

class QuantitativeAnalysisFrame1(QualitativeAnalysisFrame1):
    def __init__(self, container):
        super().__init__(container)
        self.title_label['text'] = "ANALYSIS"

    def back_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_0)
        self.base_window.switch_page()

    def next_clicked(self):
        self.program_name = self.experiment_name_text.get("1.0","end-1c")
        if(len(self.program_name) != 0):
            msg = messagebox.askokcancel("","Please make sure no sample is placed in the device !")
            if(msg == True):
                if os.path.exists(results_quantitative_path + self.program_name):
                    self.program_path = results_quantitative_path + self.program_name
                else:
                    self.program_path = os.path.join(results_quantitative_path, self.program_name)
                    os.mkdir(self.program_path)

                self.create_time = strftime(" %y-%m-%d %H.%M.%S")
                self.result_folder_name = self.program_name + self.create_time
                self.result_folder_path = os.path.join(self.program_path + '/', self.result_folder_name)
                os.mkdir(self.result_folder_path)
                print(self.result_folder_path)

                self.system_check_folder = os.path.join(self.result_folder_path, "System_Check")
                os.mkdir(self.system_check_folder)

                self.analysis_result_folder = os.path.join(self.result_folder_path, "Analysis Results")
                os.mkdir(self.analysis_result_folder)

                self.base_window.system_check.mode_check = 3
                self.base_window.forget_page()
                self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
                self.base_window.switch_page()
                self.base_window.system_check.serial_handle()

        else:
            messagebox.showwarning("","Please select the program !")

    def load_program(self):
        try:
            for i in range(len(self.program_button)):
                self.program_button[i].destroy()
        except:
            pass
        self.program_button = list(range(100))
        for file in os.listdir(programs_quantitative_path):
            self.program_button[os.listdir(programs_quantitative_path).index(file)] = Button(self.program_frame.scrollable_frame,
                                    text=file[:(len(file)-5)],
                                    font = PROGRAM_BUTTON_TXT_FONT,
                                    bg = PROGRAM_BUTTON_BGD_COLOR,
                                    fg = PROGRAM_BUTTON_TXT_COLOR,
                                    width = 51,
                                    borderwidth = 0)
            self.program_button[os.listdir(programs_quantitative_path).index(file)]['command'] = partial(self.program_clicked, os.listdir(programs_quantitative_path).index(file))
            self.program_button[os.listdir(programs_quantitative_path).index(file)].pack(pady=2, ipady=5, fill=BOTH, expand=TRUE)

    def program_clicked(self, button_index):
        wb = load_workbook(programs_quantitative_path + self.program_button[button_index]['text'] + ".xlsx")
        sheet = wb.active

        # ~ self.n_base_value = float(sheet["I2"].value)
        # ~ self.a_value = float(sheet["E2"].value)
        # ~ self.b_value = float(sheetn["G2"].value)
        self.n_base_value = float(sheet["A1"].value)
        self.value1 = float(sheet["C2"].value)
        self.concen1 = float(sheet["B2"].value)
        self.value2 = float(sheet["C3"].value)
        self.concen2 = float(sheet["B3"].value)
        self.value3 = float(sheet["C4"].value)
        self.concen3 = float(sheet["B4"].value)
        self.value4 = float(sheet["C5"].value)
        self.concen4 = float(sheet["B5"].value)
        
        concen_1_pt = [self.concen1, self.value1]
        concen_2_pt = [self.concen2, self.value2]
        concen_3_pt = [self.concen3, self.value3]
        concen_4_pt = [self.concen4, self.value4]
        
        # ~ concen_1_pt = [1.3, 1.405]
        # ~ concen_2_pt = [2, 1.46]
        # ~ concen_3_pt = [3, 1.7]
        # ~ concen_4_pt = [4, 1.87]
        
        pts_list = []
        pts_list.append(concen_1_pt)
        pts_list.append(concen_2_pt)
        pts_list.append(concen_3_pt)
        pts_list.append(concen_4_pt)
        
        pts_list = np.array(pts_list)
        print("pts_list: ", pts_list)
        x = pts_list[:,0]
        y = pts_list[:,1]
        self.a_value, self.b_value = np.polyfit(x,y,1)
        self.a_value = round(self.a_value, 4)
        self.b_value = round(self.b_value, 4)
        print("a_value: ", self.a_value)
        print("b_value: ", self.b_value)

        self.experiment_name_text['state'] = "normal"
        self.experiment_name_text.delete('1.0',END)
        try:
            self.experiment_name_text.insert('1.0', self.program_button[button_index]['text'])
        except:
            pass
        self.experiment_name_text['state'] = "disabled"

        # ~ self.user_name_text['state'] = "normal"
        # ~ self.user_name_text.delete('1.0',END)
        # ~ try:
            # ~ self.user_name_text.insert('1.0', sheet["C2"].value)
        # ~ except:
            # ~ pass
        # ~ self.user_name_text['state'] = "disabled"

        # ~ self.pfi_text['state'] = "normal"
        # ~ self.pfi_text.delete('1.0',END)
        # ~ try:
            # ~ self.pfi_text.insert('1.0', sheet["I2"].value)
        # ~ except:
            # ~ pass
        # ~ self.pfi_text['state'] = "disabled"

        self.status_text['state'] = "normal"
        self.status_text.delete('1.0',END)
        try:
            self.status_text.insert(END, str(sheet["B2"].value) + ': ')
            self.status_text.insert(END, str(sheet["C2"].value) + '\n')
            self.status_text.insert(END, str(sheet["B3"].value) + ': ')
            self.status_text.insert(END, str(sheet["C3"].value) + '\n')
            self.status_text.insert(END, str(sheet["B4"].value) + ': ')
            self.status_text.insert(END, str(sheet["C4"].value) + '\n')
            self.status_text.insert(END, str(sheet["B5"].value) + ': ')
            self.status_text.insert(END, str(sheet["C5"].value) + '\n')
            self.status_text.insert(END, "a_value: " + str(self.a_value) + '\n')
            self.status_text.insert(END, "b_value: " + str(self.b_value) + '\n')
            self.status_text.insert(END, "n_base_value: " + str(self.n_base_value) + '\n')

        except:
            pass
        self.status_text['state'] = "disabled"

        wb.close()

class QualitativeAnalysisFrame0(Frame):
    def __init__(self, container):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container

        self.experiment_name = StringVar()
        self.user_name = StringVar()
        self.comments = StringVar()

        # 3 main frame
        self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=5, fill=X)
        self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X)

        # In title frame
        self.title_label = Label(self.title_frame,
                                text = "ANALYSIS",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

        # In work frame
        setup_labelframe = LabelFrame(self.work_frame,
                                    font = LABEL_FRAME_TXT_FONT,
                                    bg = LABEL_BGD_COLOR,
                                    fg = LABEL_FRAME_TXT_COLOR)
        setup_labelframe.pack(expand=TRUE)

        experiment_name_label = Label(setup_labelframe,
                                    text = "Test name :",
                                    font = ('Helvetica', 10, 'bold'),
                                    bg = LABEL_BGD_COLOR,
                                    fg = LABEL_TXT_COLOR)
        experiment_name_label.grid(row=0, column=1, sticky=E, pady=20, padx=10)

        user_name_label = Label(setup_labelframe,
                                text = "Test technician name :",
                                font = ('Helvetica', 10, 'bold'),
                                bg = LABEL_BGD_COLOR,
                                fg = LABEL_TXT_COLOR)
        user_name_label.grid(row=1, column=1, sticky=E, pady=20, padx=10)

        # ~ comment_label = Label(setup_labelframe,
                                # ~ text = "Comment",
                                # ~ font = LABEL_TXT_FONT,
                                # ~ bg = LABEL_BGD_COLOR,
                                # ~ fg = LABEL_TXT_COLOR)
        # ~ comment_label.grid(row=2, column=1, sticky=NE, pady=20, padx=30)

        self.experiment_name_entry = Entry(setup_labelframe, width=30, font=ENTRY_TXT_FONT)
        self.experiment_name_entry.grid(row=0, column=2, sticky=W, pady=20, padx=20)
        self.user_name_entry = Entry(setup_labelframe, width=30, font=ENTRY_TXT_FONT)
        self.user_name_entry.grid(row=1, column=2, sticky=W, pady=20, padx=20)
        # ~ self.comments_text = Text(setup_labelframe, width=30, height=9, font=ENTRY_TXT_FONT)
        # ~ self.comments_text.grid(row=2, column=2, sticky=E, pady=20, padx=35)

        # In button frame
        self.back_button = Button(self.button_frame,
                                text = "Back",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.back_clicked)
        self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

        self.next_button = Button(self.button_frame,
                                text = "Next",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.next_clicked)
        self.next_button.pack(ipadx=30, ipady=10, side=RIGHT)

    def back_clicked(self):
        self.base_window.forget_page()
        #self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_option)
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
        
        self.base_window.switch_page()

    def next_clicked(self):
        self.experiment_name = self.experiment_name_entry.get()
        self.user_name = self.user_name_entry.get()
        if(self.experiment_name==''):
            messagebox.showwarning("","Plese enter Folder Name !")
        elif (self.user_name==''):
            messagebox.showwarning("","Plese enter User Name !")
        else:
            msg = messagebox.askokcancel("","Please make sure no sample is placed in the device !")
            if(msg == True):
                ### NEW ADD - START ###
                self.create_time = strftime(" %y-%m-%d %H.%M.%S")
                self.result_folder_name = self.experiment_name + self.create_time
                self.result_folder_path = os.path.join(results_qualitative_path, self.result_folder_name)
                os.mkdir(self.result_folder_path)
                print(self.result_folder_path)
                
                self.system_check_folder = os.path.join(self.result_folder_path, "System_Check")
                os.mkdir(self.system_check_folder)

                self.analysis_result_folder = os.path.join(self.result_folder_path, "Analysis Results")
                os.mkdir(self.analysis_result_folder)
                
                self.base_window.system_check.mode_check = 1
                ### NEW ADD - END ###
                
                self.base_window.forget_page()
                #self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_1)
                self.base_window.page_num = self.base_window.frame_list.index(self.base_window.system_check)
                self.base_window.switch_page()
                
                ### NEW ADD - START ###
                self.base_window.system_check.serial_handle()
                ### NEW ADD - END ###

class QuantitativeAnalysisFrame0(QualitativeAnalysisFrame0):
    def __init__(self, container):
        super().__init__(container)
        self.title_label ['text'] = "ANALYSIS"

    def back_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_option)
        self.base_window.switch_page()

    def next_clicked(self):
        self.experiment_name = self.experiment_name_entry.get()
        self.user_name = self.user_name_entry.get()
        if(self.experiment_name==''):
            messagebox.showwarning("","Plese enter Folder Name !")
        elif (self.user_name==''):
            messagebox.showwarning("","Plese enter User Name !")
        else:
            self.base_window.forget_page()
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_1)
            self.base_window.switch_page()
            self.base_window.quantitative_analysis_1.load_program()


class IDCreateFrame(Frame):
    def __init__(self, container):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container

        self.direct_create = 0

        # 3 main frame
        self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=5, fill=X)
        self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X)

        # In title frame
        self.title_label = Label(self.title_frame,
                                text = "SAMPLES NAME FILE",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

        # In work frame
        # Sample button frame
        self.well_button_table_frame = Frame(self.work_frame, bg=SAMPLE_BUTTON_FRAME_BDG_COLOR)
        self.well_button_table_frame.pack(side=LEFT)
        self.well_button = list(range(48))
        r=0
        c=-1
        for i in range(0,48):
            c+=1
            if(c>5):
                c=0
                r+=1
            self.well_button[i] = Button(self.well_button_table_frame,
                                        bg = SAMPLE_BUTTON_BGD_COLOR,
                                        fg = SAMPLE_BUTTON_TXT_COLOR,
                                        activebackground = SAMPLE_BUTTON_ACTIVE_BGD_COLOR,
                                        justify = 'left',
                                        borderwidth = 0,
                                        text = '#',
                                        width = SAMPLE_BUTTON_WIDTH,
                                        height = SAMPLE_BUTTON_HEIGHT)
            if(i!=47):
                self.well_button[i]['command'] = partial(self.well_button_clicked, i)
            self.well_button[i].grid(row=r, column=c, padx=2, pady=2)
        
        self.well_button[47]['bg'] = "blue"
        self.well_button[47]['text'] = "B"
        self.well_button[47]['state'] = "disabled"
        # Properties frame
        self.property_frame = Frame(self.work_frame, bg=SAMPLE_BUTTON_FRAME_BDG_COLOR, width=495)
        self.property_frame.pack(fill=BOTH, expand=TRUE, side=RIGHT)

        self.property_labelframe = LabelFrame(self.property_frame,
                                        text = "Sample Properties",
                                        font  = LABEL_FRAME_TXT_FONT,
                                        bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
                                        fg = LABEL_FRAME_TXT_COLOR)
        self.property_labelframe.pack(fill=BOTH, expand=TRUE, padx=10, pady=10)

        self.property_labelframe.rowconfigure(0, weight=1)
        self.property_labelframe.rowconfigure(1, weight=1)
        self.property_labelframe.rowconfigure(2, weight=1)
        self.property_labelframe.rowconfigure(3, weight=4)

        self.well_name_label = Label(self.property_labelframe,
                        bg = SAMPLE_BUTTON_CHOOSE_BGD_COLOR,
                        fg = LABEL_TXT_COLOR,
                        font = SAMPLE_LABEL_TXT_FONT)
        self.well_name_label.grid(row=0, column=0, columnspan=2, sticky=EW)

        # In button frame
        self.back_button = Button(self.button_frame,
                                text = "Back",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.back_clicked)
        self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

        self.create_button = Button(self.button_frame,
                                text = "Create",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.create_clicked)
        self.create_button.pack(ipadx=30, ipady=10, side=RIGHT)

        self.load_button = Button(self.button_frame,
                                text = "Load",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.load_clicked)
        self.load_button.pack(ipadx=30, ipady=10, anchor=CENTER)

    def well_button_clicked(self,n):
        if(self.well_button[n]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR):
            for k in range (0,47):
                if(self.well_button[k]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR and self.well_button[k]['bg'] != SAMPLE_BUTTON_TMP_BGD_COLOR):
                    self.well_button[k]['bg'] = SAMPLE_BUTTON_BGD_COLOR
                else:
                    self.well_button[k]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
            self.well_button[n]['bg'] = SAMPLE_BUTTON_CHOOSE_BGD_COLOR
        else:
            for k in range (0,47):
                if(self.well_button[k]['bg'] != SAMPLE_BUTTON_DONE_BGD_COLOR and self.well_button[k]['bg'] != SAMPLE_BUTTON_TMP_BGD_COLOR):
                    self.well_button[k]['bg'] = SAMPLE_BUTTON_BGD_COLOR
                if(self.well_button[k]['bg'] == SAMPLE_BUTTON_TMP_BGD_COLOR):
                    self.well_button[k]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
            self.well_button[n]['bg'] = SAMPLE_BUTTON_TMP_BGD_COLOR


        def ok_clicked(event=None):
            if(self.sample_name_entry.get()==''):
                self.well_button[n]['bg'] = SAMPLE_BUTTON_CHOOSE_BGD_COLOR
                self.well_button[n]['text'] = '#'
                msgbox = messagebox.showwarning("","Please enter sample name !")
            else:
                self.well_button[n]['text'] = self.sample_name_entry.get()
                self.well_button[n]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
                try:
                    if(n==42):
                        self.well_button_clicked(1)
                    elif(n==43):
                        self.well_button_clicked(2)
                    elif(n==44):
                        self.well_button_clicked(3)
                    elif(n==45):
                        self.well_button_clicked(4)
                    elif(n==46):
                        self.well_button_clicked(5)
                    elif(n==41):
                        self.well_button_clicked(0)
                    else:
                        self.well_button_clicked(n+6)
                except:
                    self.well_button_clicked(0)


        sample_name_label = Label(self.property_labelframe,
                                    text = "Sample Name",
                                    font = LABEL_TXT_FONT,
                                    bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
                                    fg = LABEL_TXT_COLOR)
        sample_name_label.grid(row=1, column=0, padx=78, pady=2, sticky=SE)

        self.sample_name_entry = Entry(self.property_labelframe, width=20, font=ENTRY_TXT_FONT)
        if(self.well_button[n]['bg'] == SAMPLE_BUTTON_TMP_BGD_COLOR):
            self.sample_name_entry.insert(0, self.well_button[n]['text'])
        #id_entry.bind("<Button-1>", enter_entry)
        self.sample_name_entry.bind("<Return>", ok_clicked)
        self.sample_name_entry.grid(row=2, column=0, padx=30, pady=0)
        self.sample_name_entry.focus_set()

        if(n<6):
            self.well_name_label['text'] = "A" + str(n+1)
        elif(n<12):
            self.well_name_label['text'] = "B" + str(n+1-6)
        elif(n<18):
            self.well_name_label['text'] = "C" + str(n+1-12)
        elif(n<24):
            self.well_name_label['text'] = "D" + str(n+1-18)
        elif(n<30):
            self.well_name_label['text'] = "E" + str(n+1-24)
        elif(n<36):
            self.well_name_label['text'] = "F" + str(n+1-30)
        elif(n<42):
            self.well_name_label['text'] = "G" + str(n+1-36)
        else:
            self.well_name_label['text'] = "H" + str(n+1-42)

        # ~ if(n<8):
            # ~ self.well_name_label['text'] = str(chr(65+n)) + '1'
        # ~ if(n>=8 and n<16):
            # ~ self.well_name_label['text'] = str(chr(65+n-8)) + '2'
        # ~ if(n>=16 and n<24):
            # ~ self.well_name_label['text'] = str(chr(65+n-16)) + '3'
        # ~ if(n>=24 and n<32):
            # ~ self.well_name_label['text'] = str(chr(65+n-24)) + '4'
        # ~ if(n>=32 and n<40):
            # ~ self.well_name_label['text'] = str(chr(65+n-32)) + '5'
        # ~ if(n>=40):
            # ~ self.well_name_label['text'] = str(chr(65+n-40)) + '6'

        self.ok_button = Button(self.property_labelframe,
                                text = "OK",
                                bg = CONFIRM_BUTTON_BGD_COLOR,
                                fg = CONFIRM_BUTTON_TXT_COLOR,
                                font = CONFIRM_BUTTON_TXT_FONT,
                                borderwidth = 0,
                                command = ok_clicked)
        self.ok_button.grid(row=3, column=0, columnspan=2, ipadx=30, ipady=10)

    def load_clicked(self):
        file = filedialog.askopenfilename(initialdir=id_path, filetypes=[('Excel file','*.xlsm *.xlsx *.xls')])
        if file is not None:
            wb = load_workbook(file)
            sheet = wb.active
            
            j=0
            for i in range(0,48):
                if(i==8):
                    j=1
                elif(i==16):
                    j=2
                elif(i==24):
                    j=3
                elif(i==32):
                    j=4
                elif(i==40):
                    j=5
                #elif(i==47):
                #   j=0
                    
                pos = 'B' + str(i+12)
                self.well_button[j]['text'] = sheet[pos].value
                if(self.well_button[j]['text'] == "N/A"):
                    self.well_button[j]['bg'] = SAMPLE_BUTTON_BGD_COLOR
                else:
                    if(i!=47):
                        self.well_button[j]['bg'] = SAMPLE_BUTTON_DONE_BGD_COLOR
                
                j=j+6
                
    def create_clicked(self):
        file_create_done = 0
        msg = messagebox.askquestion("","Do you want to create file ?")
        if(msg=="yes"):
            self.well_button_change_pos = list(range(48))
            wb = Workbook()
            sheet = wb.active
            
            self.well_button_change_pos[0] = self.well_button[0]['text']
            self.well_button_change_pos[1] = self.well_button[6]['text']
            self.well_button_change_pos[2] = self.well_button[12]['text']
            self.well_button_change_pos[3] = self.well_button[18]['text']
            self.well_button_change_pos[4] = self.well_button[24]['text']
            self.well_button_change_pos[5] = self.well_button[30]['text']
            self.well_button_change_pos[6] = self.well_button[36]['text']
            self.well_button_change_pos[7] = self.well_button[42]['text']
            
            self.well_button_change_pos[8] = self.well_button[1]['text']
            self.well_button_change_pos[9] = self.well_button[7]['text']
            self.well_button_change_pos[10] = self.well_button[13]['text']
            self.well_button_change_pos[11] = self.well_button[19]['text']
            self.well_button_change_pos[12] = self.well_button[25]['text']
            self.well_button_change_pos[13] = self.well_button[31]['text']
            self.well_button_change_pos[14] = self.well_button[37]['text']
            self.well_button_change_pos[15] = self.well_button[43]['text']
            
            self.well_button_change_pos[16] = self.well_button[2]['text']
            self.well_button_change_pos[17] = self.well_button[8]['text']
            self.well_button_change_pos[18] = self.well_button[14]['text']
            self.well_button_change_pos[19] = self.well_button[20]['text']
            self.well_button_change_pos[20] = self.well_button[26]['text']
            self.well_button_change_pos[21] = self.well_button[32]['text']
            self.well_button_change_pos[22] = self.well_button[38]['text']
            self.well_button_change_pos[23] = self.well_button[44]['text']
            
            self.well_button_change_pos[24] = self.well_button[3]['text']
            self.well_button_change_pos[25] = self.well_button[9]['text']
            self.well_button_change_pos[26] = self.well_button[15]['text']
            self.well_button_change_pos[27] = self.well_button[21]['text']
            self.well_button_change_pos[28] = self.well_button[27]['text']
            self.well_button_change_pos[29] = self.well_button[33]['text']
            self.well_button_change_pos[30] = self.well_button[39]['text']
            self.well_button_change_pos[31] = self.well_button[45]['text']
            
            self.well_button_change_pos[32] = self.well_button[4]['text']
            self.well_button_change_pos[33] = self.well_button[10]['text']
            self.well_button_change_pos[34] = self.well_button[16]['text']
            self.well_button_change_pos[35] = self.well_button[22]['text']
            self.well_button_change_pos[36] = self.well_button[28]['text']
            self.well_button_change_pos[37] = self.well_button[34]['text']
            self.well_button_change_pos[38] = self.well_button[40]['text']
            self.well_button_change_pos[39] = self.well_button[46]['text']
            
            self.well_button_change_pos[40] = self.well_button[5]['text']
            self.well_button_change_pos[41] = self.well_button[11]['text']
            self.well_button_change_pos[42] = self.well_button[17]['text']
            self.well_button_change_pos[43] = self.well_button[23]['text']
            self.well_button_change_pos[44] = self.well_button[29]['text']
            self.well_button_change_pos[45] = self.well_button[35]['text']
            self.well_button_change_pos[46] = self.well_button[41]['text']
            self.well_button_change_pos[47] = self.well_button[47]['text']
            
            for i in range(0,48):
                pos = "B" + str(i+12)
                if(self.well_button_change_pos[i] != "#"):
                    sheet[pos] = self.well_button_change_pos[i]
                else:
                    sheet[pos] = "N/A"
                    
            
            if(self.direct_create == 1):
                wb.save(id_path + '/' + self.base_window.qualitative_analysis_0.experiment_name + '.xlsm')
                msg = messagebox.askquestion("","File has been created.\n Do you want to go back to the previous step?")
                if(msg=="yes"):
                    file_name = self.base_window.qualitative_analysis_0.experiment_name
                    file_create_done = 1
            elif(self.direct_create == 1):
                wb.save(id_path + '/' + self.base_window.quantitative_analysis_0.experiment_name + '.xlsm')
                msg = messagebox.askquestion("","File has been created.\n Do you want to go back to the previous step?")
                if(msg=="yes"):
                    file_name = self.base_window.quantitative_analysis_0.experiment_name
                    file_create_done = 1
            else:
                path = filedialog.asksaveasfilename(initialdir = id_path + '/', defaultextension='.xlsm')
                if path is not None:
                    tmp = 0
                    for i in range(len(path)):
                        if(path[i]=='/'):
                            tmp = i+1
                    file_name = path[tmp:(len(path)-5)]
                    if(len(file_name)<30):
                        wb.save(path)
                        msg = messagebox.askquestion("","File has been created.\n Do you want to go back to the previous step?")
                        if(msg=="yes"):
                            file_create_done = 1
                    else:
                        messagebox.showerror("","File name cannot be more than 30 characters.")
            
            if(file_create_done):
                if(self.direct_create == 0): #create file with create module from main menu
                    self.back_clicked()
                elif(self.direct_create==1): #create file with create module from qualitative_analysis_2
                    for i in range(0,47):
                        if(self.well_button[i]['text'][0] != '#'):
                            self.base_window.qualitative_analysis_2.id_list[i] = self.well_button[i]['text']
                        else:
                            self.base_window.qualitative_analysis_2.id_list[i] = 'N/A'

                    self.base_window.qualitative_analysis_2.id_file_name_label['text'] = file_name
                    self.base_window.qualitative_analysis_2.id_file_name_label['bg'] = 'lawn green'

                    try:
                        for i in range(0,47):
                            self.base_window.qualitative_analysis_2.id_label[i].destroy()
                    except:
                        pass

                    # ~ Pmw.initialise(self.base_window)
                    # ~ self.base_window.qualitative_analysis_2.tooltip = list(range(48))

                    self.base_window.qualitative_analysis_2.id_label = list(range(48))
                    r=0
                    c=-1
                    for i in range(0,48):
                        c+=1
                        if(c>5):
                            c=0
                            r+=1
                        self.base_window.qualitative_analysis_2.id_label[i] = Label(self.base_window.qualitative_analysis_2.id_pos_frame,
                                                width=6,
                                                height=3,
                                                text = self.base_window.qualitative_analysis_2.id_list[i],
                                                # ~ bg = RESULT_LABEL_BGD_COLOR,
                                                font = RESULT_LABEL_TXT_FONT)

                        # ~ self.base_window.qualitative_analysis_2.tooltip[i] = Pmw.Balloon(self.base_window)
                        # ~ self.base_window.qualitative_analysis_2.tooltip[i].bind(self.base_window.qualitative_analysis_2.id_label[i], self.base_window.qualitative_analysis_2.id_list[i])


                        if(self.well_button[i]['text'][0] != '#'):
                            self.base_window.qualitative_analysis_2.id_label[i]['bg'] = "lawn green"
                            self.base_window.qualitative_analysis_2.id_label[i]['text'] = self.base_window.qualitative_analysis_2.id_list[i]
                        else:
                            self.base_window.qualitative_analysis_2.id_label[i]['bg'] = "grey80"
                            self.base_window.qualitative_analysis_2.id_label[i]['text'] = "N/A"


                        self.base_window.qualitative_analysis_2.id_label[i].grid(row=r, column=c, padx=1, pady=1)
                    
                    self.base_window.qualitative_analysis_2.id_label[47]['text'] = 'B'
                    self.base_window.qualitative_analysis_2.id_label[47]['bg'] = 'blue'
                    
                    self.back_clicked()
                    msg = messagebox.askokcancel("","Now you can put samples in and press Next to analyze.")

                elif(self.direct_create==2):
                    for i in range(0,47):
                        if(self.well_button[i]['text'][0] != '#'):
                            self.base_window.quantitative_analysis_2.id_list[i] = self.well_button[i]['text']
                        else:
                            self.base_window.quantitative_analysis_2.id_list[i] = 'N/A'

                    self.base_window.quantitative_analysis_2.id_file_name_label['text'] = file_name
                    self.base_window.quantitative_analysis_2.id_file_name_label['bg'] = 'lawn green'

                    try:
                        for i in range(0,48):
                            self.base_window.quantitative_analysis_2.id_label[i].destroy()
                    except:
                        pass

                    # ~ Pmw.initialise(self.base_window)
                    # ~ self.base_window.quantitative_analysis_2.tooltip = list(range(48))

                    self.base_window.quantitative_analysis_2.id_label = list(range(48))
                    r=0
                    c=-1
                    for i in range(0,48):
                        c+=1
                        if(c>5):
                            c=0
                            r+=1
                        self.base_window.quantitative_analysis_2.id_label[i] = Label(self.base_window.quantitative_analysis_2.id_pos_frame,
                                                width=6,
                                                height=3,
                                                text = self.base_window.quantitative_analysis_2.id_list[i],
                                                # ~ bg = RESULT_LABEL_BGD_COLOR,
                                                font = RESULT_LABEL_TXT_FONT)

                        # ~ self.base_window.quantitative_analysis_2.tooltip[i] = Pmw.Balloon(self.base_window)
                        # ~ self.base_window.quantitative_analysis_2.tooltip[i].bind(self.base_window.quantitative_analysis_2.id_label[i], self.base_window.quantitative_analysis_2.id_list[i])


                        if(self.well_button[i]['text'][0] != '#'):
                            self.base_window.quantitative_analysis_2.id_label[i]['bg'] = "lawn green"
                            self.base_window.quantitative_analysis_2.id_label[i]['text'] = self.base_window.quantitative_analysis_2.id_list[i]
                        else:
                            self.base_window.quantitative_analysis_2.id_label[i]['bg'] = "grey80"
                            self.base_window.quantitative_analysis_2.id_label[i]['text'] = "N/A"


                        self.base_window.quantitative_analysis_2.id_label[i].grid(row=r, column=c, padx=1, pady=1)

                    self.back_clicked()
                    msg = messagebox.askokcancel("","Now you can put samples in and press Next to analyze.")
            
    def back_clicked(self):
        try:
            self.well_button_table_frame.destroy()
            self.property_frame.destroy()
        except:
            pass

        self.well_button_table_frame = Frame(self.work_frame, bg=SAMPLE_BUTTON_FRAME_BDG_COLOR)
        self.well_button_table_frame.pack(side=LEFT)
        self.well_button = list(range(48))
        r=0
        c=-1
        for i in range(0,48):
            c+=1
            if(c>5):
                c=0
                r+=1
            self.well_button[i] = Button(self.well_button_table_frame,
                                        bg = SAMPLE_BUTTON_BGD_COLOR,
                                        fg = SAMPLE_BUTTON_TXT_COLOR,
                                        activebackground = SAMPLE_BUTTON_ACTIVE_BGD_COLOR,
                                        justify = 'left',
                                        borderwidth = 0,
                                        text = '#',
                                        width = SAMPLE_BUTTON_WIDTH,
                                        height = SAMPLE_BUTTON_HEIGHT)
            if(i!=47):
                self.well_button[i]['command'] = partial(self.well_button_clicked, i)
            self.well_button[i].grid(row=r, column=c, padx=2, pady=2)
        self.well_button[47]['bg'] = "blue"
        self.well_button[47]['text'] = "B"
        self.well_button[47]['state'] = "disabled"
        
        # Properties frame
        self.property_frame = Frame(self.work_frame, bg=SAMPLE_BUTTON_FRAME_BDG_COLOR, width=495)
        self.property_frame.pack(fill=BOTH, expand=TRUE, side=RIGHT)

        self.property_labelframe = LabelFrame(self.property_frame,
                                        text = "Sample Properties",
                                        font  = LABEL_FRAME_TXT_FONT,
                                        bg = SAMPLE_BUTTON_FRAME_BDG_COLOR,
                                        fg = LABEL_FRAME_TXT_COLOR)
        self.property_labelframe.pack(fill=BOTH, expand=TRUE, padx=10, pady=10)

        self.property_labelframe.rowconfigure(0, weight=1)
        self.property_labelframe.rowconfigure(1, weight=1)
        self.property_labelframe.rowconfigure(2, weight=1)
        self.property_labelframe.rowconfigure(3, weight=4)

        self.well_name_label = Label(self.property_labelframe,
                        bg = SAMPLE_BUTTON_CHOOSE_BGD_COLOR,
                        fg = LABEL_TXT_COLOR,
                        font = SAMPLE_LABEL_TXT_FONT)
        self.well_name_label.grid(row=0, column=0, columnspan=2, sticky=EW)




        self.base_window.forget_page()
        if(self.direct_create==0):
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
        elif(self.direct_create==1):
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_2)
        else:
            self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_2)
        self.base_window.switch_page()


class QualitativeSavedFrame(QualitativeAnalysisFrame1):
    def __init__(self, container):
        super().__init__(container)
        self.title_label['text'] = "Qualitative Program"

        self.next_button.destroy()

        self.delete_button = Button(self.button_frame,
                                text = "Delete",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.delete_clicked)
        self.delete_button.pack(side=RIGHT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)

    def delete_clicked(self):
        msg = messagebox.askquestion("","Are you sure you want to delete this program ?")
        if(msg == 'yes'):
            os.remove(programs_qualitative_path + self.experiment_name_text.get("1.0","end-1c") + '.xlsx')
            self.load_program()
            self.experiment_name_text['state'] = "normal"
            self.user_name_text['state'] = "normal"
            self.comment_text['state'] = "normal"
            self.experiment_name_text.delete('1.0',END)
            self.user_name_text.delete('1.0',END)
            self.comment_text.delete('1.0',END)
            self.experiment_name_text['state'] = "disabled"
            self.user_name_text['state'] = "disabled"
            self.comment_text['state'] = "disabled"
            messagebox.showinfo("","Deleted")


    def back_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.saved_program)
        self.base_window.switch_page()

class QuantitativeSavedFrame(QuantitativeAnalysisFrame1):
    def __init__(self, container):
        super().__init__(container)
        self.title_label['text'] = "Quantitative Program"

        self.next_button.destroy()

        self.delete_button = Button(self.button_frame,
                                text = "Delete",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.delete_clicked)
        self.delete_button.pack(side=RIGHT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)

    def delete_clicked(self):
        msg = messagebox.askquestion("","Are you sure you want to delete this program ?")
        if(msg == 'yes'):
            os.remove(programs_quantitative_path + self.experiment_name_text.get("1.0","end-1c") + '.xlsx')
            self.load_program()
            self.experiment_name_text['state'] = "normal"
            self.user_name_text['state'] = "normal"
            self.comment_text['state'] = "normal"
            self.experiment_name_text.delete('1.0',END)
            self.user_name_text.delete('1.0',END)
            self.comment_text.delete('1.0',END)
            self.experiment_name_text['state'] = "disabled"
            self.user_name_text['state'] = "disabled"
            self.comment_text['state'] = "disabled"
            messagebox.showinfo("","Deleted")

    def back_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.saved_program)
        self.base_window.switch_page()

class AnalysisFrame(Frame):
    def __init__(self, container, base_window):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = base_window

        qualitative_button = Button(self,
                                    text = "Qualitative",
                                    font = MAIN_FUCNTION_BUTTON_FONT,
                                    # ~ width = MAIN_FUNCTION_BUTTON_WIDTH,
                                    # ~ height = MAIN_FUNCTION_BUTTON_HEIGHT,
                                    bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
                                    fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.qualitative_clicked)
        qualitative_button.pack(fill=BOTH, expand=TRUE, side=LEFT, padx=90, pady=210)
        quantitative_button = Button(self,
                                    text = "Quantitative",
                                    font = MAIN_FUCNTION_BUTTON_FONT,
                                    # ~ width = MAIN_FUNCTION_BUTTON_WIDTH,
                                    # ~ height = MAIN_FUNCTION_BUTTON_HEIGHT,
                                    bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
                                    fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.quantitative_clicked)
        quantitative_button.pack(fill=BOTH, expand=TRUE, side=RIGHT, padx=90, pady=210)

    def qualitative_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_1)
        self.base_window.switch_page()
        self.base_window.qualitative_analysis_1.load_program()
    def quantitative_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_1)
        self.base_window.switch_page()
        self.base_window.quantitative_analysis_1.load_program()


class NewProgramFrame(Frame):
    def __init__(self, container):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container

        self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=5, fill=X)
        self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X)

        self.title_label = Label(self.title_frame,
                                text = "NEW PROGRAM",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

        qualitative_button = Button(self.work_frame,
                                    text = "Qualitative",
                                    font = MAIN_FUCNTION_BUTTON_FONT,
                                    # ~ width = MAIN_FUNCTION_BUTTON_WIDTH,
                                    # ~ height = MAIN_FUNCTION_BUTTON_HEIGHT,
                                    bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
                                    fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.qualitative_clicked)
        qualitative_button.pack(fill=BOTH, expand=TRUE, side=LEFT, padx=90, pady=100, ipadx=25, ipady=20)

        quantitative_button = Button(self.work_frame,
                                    text = "Quantitative",
                                    font = MAIN_FUCNTION_BUTTON_FONT,
                                    # ~ width = MAIN_FUNCTION_BUTTON_WIDTH,
                                    # ~ height = MAIN_FUNCTION_BUTTON_HEIGHT,
                                    bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
                                    fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.quantitative_clicked)
        quantitative_button.pack(fill=BOTH, expand=TRUE, side=LEFT, padx=90, pady=100, ipadx=25, ipady=20)

        self.back_button = Button(self.button_frame,
                                text = "Back",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.back_clicked)
        self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

    def qualitative_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_qualitative_1)
        self.base_window.switch_page()
        
    def quantitative_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_quantitative_1)
        self.base_window.switch_page()

    def back_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
        self.base_window.switch_page()
    
class SavedProgramFrame(NewProgramFrame):
    def __init__(self, container):
        super().__init__(container)
        self.title_label['text'] = "Saved Program"

    def qualitative_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_saved_program)
        self.base_window.switch_page()
        self.base_window.qualitative_saved_program.load_program()
    def quantitative_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_saved_program)
        self.base_window.switch_page()
        self.base_window.quantitative_saved_program.load_program()

class ServerSettingFrame(Frame):
    def __init__(self, container):
        super().__init__(container)
        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container
        self.account_active = 0

        # 3 main frame
        self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=5, fill=X)
        self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X)

        # In title frame
        self.title_label = Label(self.title_frame,
                                text = "Server",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

        # In work frame
        self.login_label = Label(self.work_frame,text ="LOGIN",
                                bg=LABEL_BGD_COLOR,
                                fg=LABEL_TXT_COLOR,
                                font=LOGIN_LABEL_TXT_FONT)
        self.login_label.grid(row=0, column=0, columnspan=3, pady=30, ipadx=1, ipady=1)

        ip_label = Label(self.work_frame, bg=LABEL_BGD_COLOR, text="IP address :")
        ip_label.grid(row=1, column=0, sticky=E, pady=15, padx=30)
        user_label = Label(self.work_frame, bg=LABEL_BGD_COLOR, text="User :")
        user_label.grid(row=2, column=0, sticky=E, pady=15, padx=30)
        password_label = Label(self.work_frame, bg=LABEL_BGD_COLOR, text="Password :")
        password_label.grid(row=3, column=0, sticky=E, pady=15, padx=30)
        path_label = Label(self.work_frame, bg=LABEL_BGD_COLOR, text="Folder path :")
        path_label.grid(row=4, column=0, sticky=E, pady=15, padx=30)

        self.ip_entry = Entry(self.work_frame, width=30, justify='left', font=('Courier',14))
        self.ip_entry.grid(row=1, column=1, sticky=W)
        self.user_entry = Entry(self.work_frame, width=30, justify='left', font=('Courier',14))
        self.user_entry.grid(row=2, column=1, sticky=W)
        self.pass_entry = Entry(self.work_frame, width=30, show='◼', justify='left', font=('Courier',14))
        self.pass_entry.grid(row=3, column=1, sticky=W)
        self.path_entry = Entry(self.work_frame, width=30, justify='left', font=('Courier',14))
        self.path_entry.grid(row=4, column=1, sticky=W)

        hide_var = IntVar()
        def hide_charaters():
            if(hide_var.get()==0):
                self.pass_entry['show']=""
            else:
                self.pass_entry['show']="◼"
        self.hidepass_checkbutton = Checkbutton(self.work_frame, variable=hide_var, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text = "Hide charaters",onvalue=1, offvalue=0, command=hide_charaters)
        self.hidepass_checkbutton.select()
        self.hidepass_checkbutton.grid(row=3, column=2, sticky=W, padx=10)

        self.connect_button = Button(self.work_frame,
                                text = "Connect",
                                bg = LOGIN_BUTTON_BGD_COLOR,
                                fg = LOGIN_BUTTON_TXT_COLOR,
                                font = LOGIN_BUTTON_TXT_FONT,
                                borderwidth = 0,
                                command = self.connect_clicked)
        self.connect_button.grid(row=5, column=0, columnspan=3, ipadx=30, ipady=10, pady=20)


        # In button frame
        self.back_button = Button(self.button_frame,
                                text = "Back",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.back_clicked)
        self.back_button.pack(ipadx=30, ipady=10, side=LEFT)

        # ~ self.save_button = Button(self.button_frame,
                                # ~ text = "Back",
                                # ~ font = SWITCH_PAGE_BUTTON_FONT,
                                # ~ # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # ~ # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                # ~ bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                # ~ fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                # ~ borderwidth = 0,
                                # ~ command = self.save_clicked)
        # ~ self.save_button.pack(ipadx=30, ipady=10, side=LEFT, sticky=W)

        self.check_status()

    def back_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.setting)
        self.base_window.switch_page()

    def connect_clicked(self):
        if(self.connect_button['text']=='Connect'):
            self.ip_set = self.ip_entry.get().strip()
            self.user_set = self.user_entry.get().strip()
            self.password_set = self.pass_entry.get().strip()
            self.path_set = self.path_entry.get().strip()
            if(self.ip_set==''):
                messagebox.showwarning("","Please enter yours IP address.")
            elif(self.user_set==''):
                messagebox.showwarning("","Please enter yours user name.")
            elif(self.password_set==''):
                messagebox.showwarning("","Please enter yours password.")
            # ~ elif(self.path_set==''):
                # ~ messagebox.showwarning("","Please enter yours folder path.")
            else:
                try:
                    #FTP:
                    ftp = FTP(self.ip_set, self.user_set, self.password_set)
                    ftp.cwd(self.path_set + '/')
                    ftp.quit()
                    
                    
                    #LAN:
                    # ~ os.system('sudo mount -t cifs -o username=' + self.user_set + ',password=' + self.password_set + ' //' + self.ip_set + '/' + self.path_set + ' /home/pi/Server')
                    
                    tc= open(working_dir + '/.server.txt',"w")
                    tc.writelines('1\n')
                    tc.writelines(self.ip_set+"\n")
                    tc.writelines(self.user_set+"\n")
                    tc.writelines(self.password_set+"\n")
                    tc.writelines(self.path_set+"\n")
                    tc.close()

                    self.check_status()

                    messagebox.showinfo("", "Connection successful.")
                except Exception as e :
                    error = messagebox.showerror("Could not connect to server.",str(e))
                    if(error=='ok'):
                        pass
        else:
            msg = messagebox.askquestion("","Do you want to disconnect from the server ?")
            if(msg=='yes'):
                tc= open(working_dir + '/.server.txt',"w")
                tc.writelines('0\n')
                tc.writelines("\n")
                tc.writelines("\n")
                tc.writelines("\n")
                tc.writelines("\n")
                tc.close()
                
                #LAN:
                # ~ try:
                    # ~ os.system('sudo umount //' + self.ip_set + '/' + self.path_set)
                # ~ except:
                    # ~ pass
                
                self.check_status()
                
    def check_status(self):
        fr = open(working_dir + "/.server.txt","r")
        self.server_active = int(fr.readline().strip())
        self.ip_set = fr.readline().strip()
        self.user_set = fr.readline().strip()
        self.password_set = fr.readline().strip()
        self.path_set = fr.readline().strip()
        fr.close()

        if(self.server_active == 0):
            self.connect_button['text'] = 'Connect'
            self.login_label['text'] = 'Server'
            self.login_label['fg'] = 'black'

            self.ip_entry['state'] = 'normal'
            self.user_entry['state'] = 'normal'
            self.pass_entry['state'] = 'normal'
            self.path_entry['state'] = 'normal'
            self.ip_entry.delete(0,END)
            self.user_entry.delete(0,END)
            self.pass_entry.delete(0,END)
            self.path_entry.delete(0,END)

            self.pass_entry['show']="◼"
            self.hidepass_checkbutton['state'] = 'normal'
            self.hidepass_checkbutton.select()
        else:
            self.connect_button['text'] = 'Disconnect'
            self.login_label['text'] = 'You are already connected to the server ✔️'
            self.login_label['fg'] = 'green3'

            self.ip_entry['state'] = 'normal'
            self.user_entry['state'] = 'normal'
            self.pass_entry['state'] = 'normal'
            self.path_entry['state'] = 'normal'
            self.ip_entry.delete(0,END)
            self.user_entry.delete(0,END)
            self.pass_entry.delete(0,END)
            self.path_entry.delete(0,END)
            self.ip_entry.insert(END, self.ip_set)
            self.user_entry.insert(END, self.user_set)
            self.pass_entry.insert(END, self.password_set)
            self.path_entry.insert(END, self.path_set)
            self.ip_entry['state'] = 'disabled'
            self.user_entry['state'] = 'disabled'
            self.pass_entry['state'] = 'disabled'
            self.path_entry['state'] = 'disabled'
            
            self.pass_entry['show']="◼"
            self.hidepass_checkbutton.select()
            self.hidepass_checkbutton['state'] = 'disable'
            

class EmailSettingFrame(Frame):
    def __init__(self, container):
        super().__init__(container)
        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container
        self.account_active = 0

        # 3 main frame
        self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=5, fill=X)
        self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X)

        # In title frame
        self.title_label = Label(self.title_frame,
                                text = "EMAIL",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

        # In work frame
        self.login_label = Label(self.work_frame,text ="LOGIN",
                                bg=LABEL_BGD_COLOR,
                                fg=LABEL_TXT_COLOR,
                                font=LOGIN_LABEL_TXT_FONT)
        self.login_label.grid(row=0, column=0, columnspan=2, pady=30, ipadx=1, ipady=1)

        user_label = Label(self.work_frame, bg=LABEL_BGD_COLOR, text="Email :")
        user_label.grid(row=1, column=0, sticky=E, pady=20, padx=30)
        pass_label = Label(self.work_frame, bg=LABEL_BGD_COLOR, text="Device Password :")
        pass_label.grid(row=2, column=0, sticky=NE, pady=3, padx=30)

        self.user_entry = Entry(self.work_frame, width=30, justify='right', font=('Courier',14))
        self.user_entry.grid(row=1, column=1, sticky=W)
        self.pass_entry = Entry(self.work_frame, width=30, show='◼', justify='right', font=('Courier',14))
        self.pass_entry.grid(row=2, column=1, sticky=NW)

        hide_var = IntVar()
        def hide_charaters():
            if(hide_var.get()==0):
                self.pass_entry['show']=""
            else:
                self.pass_entry['show']="◼"
        self.hidepass_checkbutton = Checkbutton(self.work_frame, variable=hide_var, bg=MAIN_FUNCTION_FRAME_BGD_COLOR, text = "Hide charaters",onvalue=1, offvalue=0, command=hide_charaters)
        self.hidepass_checkbutton.select()
        self.hidepass_checkbutton.grid(row=2, column=1, pady=32, sticky=SE)

        self.login_button = Button(self.work_frame,
                                text = "Login",
                                bg = LOGIN_BUTTON_BGD_COLOR,
                                fg = LOGIN_BUTTON_TXT_COLOR,
                                font = LOGIN_BUTTON_TXT_FONT,
                                borderwidth = 0,
                                command = self.login_clicked)
        self.login_button.grid(row=3, column=0, columnspan=2, ipadx=30, ipady=10, pady=30)


        # In button frame
        self.back_button = Button(self.button_frame,
                                text = "Back",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.back_clicked)
        self.back_button.pack(ipadx=30, ipady=10, side=LEFT)


    def back_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.setting)
        self.base_window.switch_page()

    def check_status(self):
        fr = open(working_dir + "/.email.txt","r")
        self.account_active = int(fr.readline())
        self.email_address = fr.readline()
        self.email_password = fr.readline()
        fr.close()

        if(self.account_active==1):
            self.login_label['text'] = "You are already logged in ✔️"
            self.login_label['fg'] = 'green2'

            self.user_entry['state'] = 'normal'
            self.pass_entry['state'] = 'normal'
            self.user_entry.delete(0,END)
            self.pass_entry.delete(0,END)
            self.user_entry.insert(END, self.email_address)
            self.pass_entry.insert(END, self.email_password)
            self.user_entry['state'] = 'disabled'
            self.pass_entry['state'] = 'disabled'

            self.pass_entry['show']="◼"
            self.hidepass_checkbutton.select()
            self.hidepass_checkbutton['state'] = 'disabled'

            self.login_button['text'] = 'Logout'
        else:
            self.login_label['text'] = "LOGIN"
            self.login_label['fg'] = LABEL_TXT_COLOR

            self.user_entry['state'] = 'normal'
            self.pass_entry['state'] = 'normal'
            self.user_entry.delete(0,END)
            self.pass_entry.delete(0,END)

            self.login_button['text'] = 'Login'

            self.hidepass_checkbutton['state'] = 'normal'

    def login_clicked(self):
        if(self.login_button['text'] == 'Login'):
            if(self.user_entry.get()==''):
                messagebox.showwarning("","Please enter the email address!")
            elif(self.pass_entry.get()==''):
                messagebox.showwarning("","Please enter the password !")
            else:
                self.email_address = self.user_entry.get()
                self.email_password = self.pass_entry.get()

                addressToVerify = self.email_address
                match = re.match('^[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})$', addressToVerify)
                if(match == None):
                    messagebox.showerror("","Email syntax error")
                else:
                    domain_name = self.email_address.split('@')[1]
                    records = dns.resolver.query(domain_name, 'MX')
                    mxRecord = records[0].exchange
                    mxRecord = str(mxRecord)

                    host = socket.gethostname()

                    server = smtplib.SMTP()
                    server.set_debuglevel(0)

                    server.connect(mxRecord)
                    server.helo(host)
                    server.mail('me@domain.com')
                    code, message = server.rcpt(str(addressToVerify))
                    server.quit()

                    if(code==250):
                        server=smtplib.SMTP('smtp.gmail.com:587')
                        server.starttls()
                        try:
                            server.login(self.email_address,self.email_password)
                            save_file = open(working_dir + '/.email.txt',"w")
                            save_file.writelines('1' + "\n")
                            save_file.writelines(self.email_address + "\n")
                            save_file.writelines(self.email_password + "\n")
                            save_file.close()
                            messagebox.showinfo("", "Login Success !")
                            self.check_status()
                        except:
                            messagebox.showerror("","Your password was incorrect\rPlease try again !")
                        server.quit()
                    else:
                        messagebox.showerror("","Your email address was incorrect\rPlease try again !")
        else:
            msg = messagebox.askquestion("","Do you want to logout ?")
            if(msg=='yes'):
                save_file = open(working_dir + '/.email.txt',"w")
                save_file.writelines('0' + "\n")
                save_file.writelines("\n")
                save_file.writelines("\n")
                save_file.close()

                self.account_active = 0

                self.login_label['text'] = "LOGIN"
                self.login_label['fg'] = LABEL_TXT_COLOR

                self.user_entry['state'] = 'normal'
                self.pass_entry['state'] = 'normal'
                self.user_entry.delete(0,END)
                self.pass_entry.delete(0,END)

                self.login_button['text'] = 'Login'
                self.hidepass_checkbutton['state'] = 'normal'

class ProgramFrame(Frame):
    def __init__(self, container, base_window):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = base_window

        new_program_button = Button(self,
                                    text = "New Program",
                                    font = MAIN_FUCNTION_BUTTON_FONT,
                                    # ~ width = MAIN_FUNCTION_BUTTON_WIDTH,
                                    # ~ height = MAIN_FUNCTION_BUTTON_HEIGHT,
                                    bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
                                    fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.new_program_clicked)
        new_program_button.pack(fill=BOTH, expand=TRUE, side=LEFT, padx=90, pady=210)
        saved_program_button = Button(self,
                                    text = "Saved Program",
                                    font = MAIN_FUCNTION_BUTTON_FONT,
                                    # ~ width = MAIN_FUNCTION_BUTTON_WIDTH,
                                    # ~ height = MAIN_FUNCTION_BUTTON_HEIGHT,
                                    bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
                                    fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.saved_program_clicked)
        saved_program_button.pack(fill=BOTH, expand=TRUE, side=RIGHT, padx=90, pady=210)

    def new_program_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_program)
        self.base_window.switch_page()
    def saved_program_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.saved_program)
        self.base_window.switch_page()

class SetIdFrame(Frame):
    def __init__(self, container, base_window):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = base_window

        set_button = Button(self,
                                    text = "Create ID file",
                                    font = MAIN_FUCNTION_BUTTON_FONT,
                                    # ~ width = MAIN_FUNCTION_BUTTON_WIDTH,
                                    # ~ height = MAIN_FUNCTION_BUTTON_HEIGHT,
                                    bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
                                    fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.create_clicked)
        set_button.pack(ipady=20, ipadx=20, expand=TRUE)

    def create_clicked(self):
        del self.base_window.id_create
        self.base_window.id_create = IDCreateFrame(self.base_window)
        self.base_window.frame_list.append(self.base_window.id_create)

        self.base_window.id_create.direct_create = 0
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.id_create)
        self.base_window.switch_page()

class ViewResultFrame(Frame):
    def __init__(self, container):
        super().__init__(container)
        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container
        self.account_active = 0

        # 3 main frame
        self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=5, fill=X)
        self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X)

        # In title frame
        self.title_label = Label(self.title_frame,
                                text = "View Results",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

        #In work frame
        self.infor_frame = LabelFrame(self.work_frame,
                                bg= LABEL_FRAME_BGD_COLOR,
                                font = LABELFRAME_TXT_FONT,
                                fg = LABEL_FRAME_TXT_COLOR,
                                width=150,
                                text = 'Information')
        self.infor_frame.pack(side=LEFT, anchor=W, pady=5, padx=5)

        self.report_frame = ScrollableFrame2(self.work_frame)
        self.report_frame.pack(side=RIGHT, pady=5)

        # In button frame
        self.back_button = Button(self.button_frame,
                                text = "Back",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.back_clicked)
        self.back_button.pack(ipadx=30, ipady=10,side=LEFT)

        self.open_button = Button(self.button_frame,
                                text = "Open",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.open_clicked)
        self.open_button.pack(ipadx=30, ipady=10, side=LEFT, padx=225)

        # ~ self.delete_button = Button(self.button_frame,
                                # ~ text = "Delete",
                                # ~ font = SWITCH_PAGE_BUTTON_FONT,
                                # ~ # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # ~ # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                # ~ bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                # ~ fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                # ~ borderwidth = 0,
                                # ~ state = 'disabled',
                                # ~ command = self.delete_clicked)
        # ~ self.delete_button.pack(ipadx=30, ipady=10, side=LEFT)

    def back_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
        self.base_window.switch_page()

    def open_clicked(self):
        sample_button_list = list(range(49))
        result_button_list = list(range(49))
        position_button_list = list(range(49))

        self.path = filedialog.askopenfilename(initialdir=results_path, filetypes=[('Excel file','.xlsm')])
        if self.path is not None:
            try:
                self.infor_frame.destroy()
                self.infor_frame = LabelFrame(self.work_frame,
                                bg= LABEL_FRAME_BGD_COLOR,
                                font = LABELFRAME_TXT_FONT,
                                fg = LABEL_FRAME_TXT_COLOR,
                                width=150,
                                text = 'Information')
                self.infor_frame.pack(side=LEFT, anchor=N, pady=5, padx=5)
            except:
                pass

            try:
                self.report_frame.destroy()
                self.report_frame = ScrollableFrame2(self.work_frame)
                self.report_frame.pack(side=RIGHT, pady=5)
            except:
                pass


            wb = load_workbook(self.path)
            sheet = wb.active

            info1_button_list = Button(self.infor_frame,
                    fg = LABEL_TXT_COLOR,
                    font = LABEL_TXT_FONT,
                    text= sheet['B7'].value,
                    width=30,
                    bg = 'lavender',
                    borderwidth = 0)
            info1_button_list.grid(row=0, column=0, sticky=EW, padx=1, pady=1)
            info2_button_list = Button(self.infor_frame,
                    fg = LABEL_TXT_COLOR,
                    font = LABEL_TXT_FONT,
                    text= sheet['B8'].value,
                    width=30,
                    bg = 'lavender',
                    borderwidth = 0)
            info2_button_list.grid(row=1, column=0, sticky=EW, padx=1, pady=1)
            info3_button_list = Button(self.infor_frame,
                    fg = LABEL_TXT_COLOR,
                    font = LABEL_TXT_FONT,
                    text= sheet['B9'].value,
                    width=30,
                    bg = 'lavender',
                    borderwidth = 0)
            info3_button_list.grid(row=2, column=0, sticky=EW, padx=1, pady=1)

            for i in range(0,49):
                sample_pos = 'B' + str(i+11)
                sample_button_list[i] = Button(self.report_frame.scrollable_frame,
                        fg = LABEL_TXT_COLOR,
                        font = LABEL_TXT_FONT,
                        text= sheet[sample_pos].value,
                        width=30,
                        bg = 'lavender',
                        borderwidth = 0)
                sample_button_list[i].grid(row=i, column=0, sticky=EW, padx=1, pady=1)

                position_pos = 'C' + str(i+11)
                position_button_list[i] = Button(self.report_frame.scrollable_frame,
                        fg = LABEL_TXT_COLOR,
                        font = LABEL_TXT_FONT,
                        text= sheet[position_pos].value,
                        width=10,
                        bg = 'lavender',
                        borderwidth = 0)
                position_button_list[i].grid(row=i, column=1, sticky=EW, padx=1, pady=1)

                result_pos = 'D' + str(i+11)
                result_button_list[i] = Button(self.report_frame.scrollable_frame,
                        fg = LABEL_TXT_COLOR,
                        font = LABEL_TXT_FONT,
                        text= sheet[result_pos].value,
                        width=22,
                        bg = 'lavender',
                        borderwidth = 0)
                result_button_list[i].grid(row=i, column=3, sticky=EW, padx=1, pady=1)
            wb.close()


class QualitativeCalibListFrame(QualitativeAnalysisFrame1):
    def __init__(self, container):
        super().__init__(container)
        self.title_label['text'] = "CALIBRATION"

        self.edit_frame = Frame(self.work_frame,
                                bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.edit_frame.grid(row=1, column=0, sticky=EW)

        new_button = Button(self.edit_frame,
                                text = "New",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.new_clicked)
        new_button.grid(row=0, column=0, ipadx=35, ipady=10)

        recalibration_button = Button(self.edit_frame,
                                text = "Recalibration",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.recalibration_clicked)
        recalibration_button.grid(row=0, column=1, padx=10, ipadx=10, ipady=10)

        delete_button = Button(self.edit_frame,
                                text = "Delete",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.delete_clicked)
        delete_button.grid(row=0, column=2, ipadx=30, ipady=10)

        self.next_button.pack_forget()

    def delete_clicked(self):
        self.program_name = self.experiment_name_text.get("1.0","end-1c")
        if(len(self.program_name) != 0):
            msg = messagebox.askquestion("","Are you sure you want to delete this kit ?")
            if(msg == 'yes'):
                os.remove(programs_qualitative_path + self.experiment_name_text.get("1.0","end-1c") + '.xlsx')
                self.load_program()
                self.experiment_name_text['state'] = "normal"
                self.user_name_text['state'] = "normal"
                self.pfi_text['state'] = "normal"
                self.comment_text['state'] = "normal"
                self.experiment_name_text.delete('1.0',END)
                self.user_name_text.delete('1.0',END)
                self.pfi_text.delete('1.0', END)
                self.comment_text.delete('1.0',END)
                self.experiment_name_text['state'] = "disabled"
                self.user_name_text['state'] = "disabled"
                self.pfi_text['state'] = "disabled"
                self.comment_text['state'] = "disabled"
                messagebox.showinfo("","Deleted")
        else:
            messagebox.showwarning("","Please select the kit !")


    def recalibration_clicked(self):
        self.program_name = self.experiment_name_text.get("1.0","end-1c")
        if(len(self.program_name) != 0):
            msg = messagebox.askquestion("","Are you sure you want to recalibrate this kit ?")
            if(msg == 'yes'):
                os.remove(programs_qualitative_path + self.experiment_name_text.get("1.0","end-1c") + '.xlsx')
                self.load_program()

                self.base_window.new_qualitative_1.experiment_name_entry.delete(0,END)
                self.base_window.new_qualitative_1.user_name_entry.delete(0,END)
                self.base_window.new_qualitative_1.comments_text.delete('1.0',END)

                self.base_window.forget_page()
                self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_qualitative_1)
                self.base_window.switch_page()

                self.base_window.new_qualitative_1.experiment_name_entry.insert(0, self.experiment_name_text.get("1.0",'end-1c'))
                self.base_window.new_qualitative_1.user_name_entry.insert(0, self.user_name_text.get("1.0",'end-1c'))
                self.base_window.new_qualitative_1.comments_text.insert('1.0', self.comment_text.get("1.0",'end-1c'))

                self.experiment_name_text['state'] = "normal"
                self.user_name_text['state'] = "normal"
                self.pfi_text['state'] = "normal"
                self.comment_text['state'] = "normal"
                self.experiment_name_text.delete('1.0',END)
                self.user_name_text.delete('1.0',END)
                self.pfi_text.delete('1.0', END)
                self.comment_text.delete('1.0',END)
                self.experiment_name_text['state'] = "disabled"
                self.user_name_text['state'] = "disabled"
                self.pfi_text['state'] = "disabled"
                self.comment_text['state'] = "disabled"

        else:
            messagebox.showwarning("","Please select the kit !")

    def new_clicked(self):
        self.base_window.new_qualitative_1.experiment_name_entry.delete(0,END)
        self.base_window.new_qualitative_1.user_name_entry.delete(0,END)
        self.base_window.new_qualitative_1.comments_text.delete('1.0',END)

        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_qualitative_1)
        self.base_window.switch_page()

    def back_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_option)
        self.base_window.switch_page()


class QuantitativeCalibListFrame(QuantitativeAnalysisFrame1):
    def __init__(self, container):
        super().__init__(container)
        self.title_label['text'] = "CALIBRATION"

        self.edit_frame = Frame(self.work_frame,
                                bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.edit_frame.grid(row=1, column=0, sticky=EW)

        new_button = Button(self.edit_frame,
                                text = "New",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.new_clicked)
        new_button.grid(row=0, column=0, ipadx=35, ipady=10)

        recalibration_button = Button(self.edit_frame,
                                text = "Recalibration",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.recalibration_clicked)
        recalibration_button.grid(row=0, column=1, padx=10, ipadx=10, ipady=10)

        delete_button = Button(self.edit_frame,
                                text = "Delete",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.delete_clicked)
        delete_button.grid(row=0, column=2, ipadx=30, ipady=10)

        self.next_button.pack_forget()

    def delete_clicked(self):
        self.program_name = self.experiment_name_text.get("1.0","end-1c")
        if(len(self.program_name) != 0):
            msg = messagebox.askquestion("","Are you sure you want to delete this kit ?")
            if(msg == 'yes'):
                os.remove(programs_quantitative_path + self.experiment_name_text.get("1.0","end-1c") + '.xlsx')
                self.load_program()
                self.experiment_name_text['state'] = "normal"
                self.user_name_text['state'] = "normal"
                self.pfi_text['state'] = "normal"
                self.comment_text['state'] = "normal"
                self.experiment_name_text.delete('1.0',END)
                self.user_name_text.delete('1.0',END)
                self.pfi_text.delete('1.0', END)
                self.comment_text.delete('1.0',END)
                self.experiment_name_text['state'] = "disabled"
                self.user_name_text['state'] = "disabled"
                self.pfi_text['state'] = "disabled"
                self.comment_text['state'] = "disabled"
                messagebox.showinfo("","Deleted")
        else:
            messagebox.showwarning("","Please select the kit !")


    def recalibration_clicked(self):
        self.program_name = self.experiment_name_text.get("1.0","end-1c")
        if(len(self.program_name) != 0):
            msg = messagebox.askquestion("","Are you sure you want to recalibrate this kit ?")
            if(msg == 'yes'):
                os.remove(programs_quantitative_path + self.experiment_name_text.get("1.0","end-1c") + '.xlsx')
                self.load_program()

                self.base_window.new_quantitative_1.experiment_name_entry.delete(0,END)
                self.base_window.new_quantitative_1.user_name_entry.delete(0,END)
                self.base_window.new_quantitative_1.comments_text.delete('1.0',END)

                self.base_window.forget_page()
                self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_quantitative_1)
                self.base_window.switch_page()

                self.base_window.new_quantitative_1.experiment_name_entry.insert(0, self.experiment_name_text.get("1.0",'end-1c'))
                self.base_window.new_quantitative_1.user_name_entry.insert(0, self.user_name_text.get("1.0",'end-1c'))
                self.base_window.new_quantitative_1.comments_text.insert('1.0', self.comment_text.get("1.0",'end-1c'))

                os.remove(programs_quantitative_path + self.experiment_name_text.get("1.0","end-1c") + '.xlsx')
                self.load_program()
                self.experiment_name_text['state'] = "normal"
                self.user_name_text['state'] = "normal"
                self.pfi_text['state'] = "normal"
                self.comment_text['state'] = "normal"
                self.experiment_name_text.delete('1.0',END)
                self.user_name_text.delete('1.0',END)
                self.pfi_text.delete('1.0', END)
                self.comment_text.delete('1.0',END)
                self.experiment_name_text['state'] = "disabled"
                self.user_name_text['state'] = "disabled"
                self.pfi_text['state'] = "disabled"
                self.comment_text['state'] = "disabled"
        else:
            messagebox.showwarning("","Please select the kit !")

    def new_clicked(self):
        self.base_window.new_quantitative_1.experiment_name_entry.delete(0,END)
        self.base_window.new_quantitative_1.user_name_entry.delete(0,END)
        self.base_window.new_quantitative_1.comments_text.delete('1.0',END)

        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.new_quantitative_1)
        self.base_window.switch_page()

    def back_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_option)
        self.base_window.switch_page()


class QualitativeOptionFrame(Frame):
    def __init__(self, container):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container

        self.title_frame = Frame(self, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=5, fill=X)
        self.work_frame = Frame(self, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.work_frame.pack_propagate(0)
        self.button_frame = Frame(self, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X, expand=TRUE)

        # In title frame
        self.title_label = Label(self.title_frame,
                                text = "SCREENING MODE",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

        # In work frame
        self.calibration_button = Button(self.work_frame,
                                    text = "Calibration",
                                    font = MAIN_FUCNTION_BUTTON_FONT,
                                    bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
                                    fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.calibration_clicked)
        self.calibration_button.grid(row=0, column=0, ipadx=30, ipady=20, padx=80, pady=168)

        self.analysis_button = Button(self.work_frame,
                                    text = "Analysis",
                                    font = MAIN_FUCNTION_BUTTON_FONT,
                                    bg = MAIN_FUNCTION_BUTTON_BGD_COLOR,
                                    fg = MAIN_FUNCTION_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.analysis_clicked)
        self.analysis_button.grid(row=0, column=1, ipadx=30, ipady=20, padx=80, pady=168)

        # In button frame
        self.back_button = Button(self.button_frame,
                                text = "Back",
                                font = SWITCH_PAGE_BUTTON_FONT,
                                # width = SWITCH_PAGE_BUTTON_WIDTH,
                                # height = SWITCH_PAGE_BUTTON_HEIGHT,
                                bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                                fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                                borderwidth = 0,
                                command = self.back_clicked)
        self.back_button.pack(side=LEFT, padx=0, pady=0, ipady=10, ipadx=30, anchor=W)

    def calibration_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_calib_list)
        self.base_window.switch_page()
        self.base_window.qualitative_calib_list.load_program()
    def analysis_clicked(self):
        self.dye_label_frame = LabelFrame(self.work_frame,
                                            width = 450,
                                            height = 50,
                                            text = "NUMBER OF DYES",
                                            bg = 'dodger blue')
        self.dye_label_frame.place(x=20, y=100)

        self.one_dye_button = Button(self.dye_label_frame,
                    text = "1 Dye",
                    font = SWITCH_PAGE_BUTTON_FONT,

                    bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                    fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                    borderwidth = 0,
                    command = self.one_dye_clicked)
        self.one_dye_button.pack(side=LEFT, padx=85, pady=50, ipady=10, ipadx=20)

        self.two_dye_button = Button(self.dye_label_frame,
                    text = "2 Dyes",
                    font = SWITCH_PAGE_BUTTON_FONT,
                    bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                    fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                    borderwidth = 0,
                    command = self.two_dye_clicked)
        self.two_dye_button.pack(side=LEFT, padx=85, pady=50, ipady=10, ipadx=20, anchor=W)
        
        self.cancel_button = Button(self.dye_label_frame,
                    text = "X",
                    font = SWITCH_PAGE_BUTTON_FONT,
                    width = 1,
                    height = 1,
                    bg = SWITCH_PAGE_BUTTON_BGD_COLOR,
                    fg = SWITCH_PAGE_BUTTON_TXT_COLOR,
                    borderwidth = 0,
                    command = self.cancel_clicked)
        self.cancel_button.place(x=516,y=-8)
        
    def back_clicked(self):
        self.cancel_clicked()
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
        self.base_window.switch_page()
    
    def one_dye_clicked(self):
        self.number_of_dyes = 1 
        self.cancel_clicked()
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
        self.base_window.switch_page()
    def two_dye_clicked(self):
        self.number_of_dyes = 2
        self.cancel_clicked()
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
        self.base_window.switch_page()
        
    def cancel_clicked(self):
        try:
            self.dye_label_frame.place_forget()
        except:
            pass

class QuantitativeOptionFrame(QualitativeOptionFrame):
    def __init__(self, container):
        super().__init__(container)

        self.title_label['text'] = "QUANTITATIVE MODE"

    def calibration_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_calib_list)
        self.base_window.switch_page()
        self.base_window.quantitative_calib_list.load_program()
    def analysis_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_analysis_0)
        self.base_window.switch_page()
    def back_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.main_menu)
        self.base_window.switch_page()


class SettingFrame(QualitativeOptionFrame):
    def __init__(self, container):
        super().__init__(container)

        self['bg'] = MAIN_FUNCTION_FRAME_BGD_COLOR
        self.base_window = container

        self.title_label['text'] = 'SETTING'

        self.calibration_button['text'] = "Email"
        self.analysis_button['text'] = "Server"

        self.calibration_button['command'] = self.email_clicked
        self.analysis_button['command'] = self.server_clicked


    def email_clicked(self):
        self.base_window.forget_page()
        self.base_window.email_setting.check_status()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.email_setting)
        self.base_window.switch_page()
    def server_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.server_setting)
        self.base_window.switch_page()



class MainMenu(Frame):
    def __init__(self, master):
        super().__init__(master)
        self.base_window = master

        # Base frame create
        self.base_frame = Frame(self,bg=MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.base_frame.pack(fill=BOTH, expand=TRUE)

        self.title_frame = Frame(self.base_frame, bg = TITILE_FRAME_BGD_COLOR)
        self.title_frame.pack(ipadx=0, ipady=20, fill=X)
        self.work_frame = Frame(self.base_frame, bg = MAIN_FUNCTION_FRAME_BGD_COLOR)
        self.work_frame.pack(expand=TRUE)
        self.button_frame = Frame(self.base_frame, bg = MAIN_MENU_BUTTON_FRAME_BGD_COLOR)
        self.button_frame.pack(fill=X)

        self.title_label = Label(self.title_frame,
                                text = "SPOTCHECK",
                                font = TITLE_TXT_FONT,
                                bg = TITILE_FRAME_BGD_COLOR,
                                fg = TITILE_FRAME_TXT_COLOR)
        self.title_label.pack(expand=TRUE)

        self.screening_button = Button(self.work_frame,
                                    text = "SCREENING MODE",
                                    font = MAIN_MENU_BUTTON_FONT,
                                    width = MAIN_MENU_BUTTON_WIDTH,
                                    height = MAIN_MENU_BUTTON_HEIGHT,
                                    bg = MAIN_MENU_BUTTON_BGD_COLOR,
                                    fg = MAIN_MENU_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.screening_clicked)
        self.screening_button.grid(row=0, column=0, ipadx=20, ipady=10, padx=100, pady=130)

        self.quantitative_button = Button(self.work_frame,
                                    text = "QUANTITATIVE MODE",
                                    font = MAIN_MENU_BUTTON_FONT,
                                    width = MAIN_MENU_BUTTON_WIDTH,
                                    height = MAIN_MENU_BUTTON_HEIGHT,
                                    bg = MAIN_MENU_BUTTON_BGD_COLOR,
                                    fg = MAIN_MENU_BUTTON_TXT_COLOR,
                                    borderwidth = 0,
                                    command = self.quantitative_clicked)
        self.quantitative_button.grid(row=0, column=1, ipadx=20, ipady=10, padx=100, pady=130)

        self.view_result_button = Button(self.button_frame,
                                    text = "View Results",
                                    font = MAIN_MENU_BUTTON_FONT,
                                    bg = MAIN_MENU_BUTTON_BGD_COLOR,
                                    fg = MAIN_MENU_BUTTON_TXT_COLOR,
                                    # ~ width = 16,
                                    # ~ height = 2,
                                    borderwidth = 0,
                                    command = self.view_result_clicked)
        self.view_result_button.pack(side=LEFT, fill=BOTH, expand=TRUE, ipady=5)

        self.create_file_button = Button(self.button_frame,
                                    text = "Create samples name file",
                                    font = MAIN_MENU_BUTTON_FONT,
                                    bg = MAIN_MENU_BUTTON_BGD_COLOR,
                                    fg = MAIN_MENU_BUTTON_TXT_COLOR,
                                    # ~ width = 16,
                                    # ~ height = 2,
                                    borderwidth = 0,
                                    command = self.set_id_clicked)
        self.create_file_button.pack(side=LEFT, fill=BOTH, expand=TRUE)

        self.setting_button = Button(self.button_frame,
                                    text = "Setting",
                                    font = MAIN_MENU_BUTTON_FONT,
                                    bg = MAIN_MENU_BUTTON_BGD_COLOR,
                                    fg = MAIN_MENU_BUTTON_TXT_COLOR,
                                    # ~ width = 16,
                                    # ~ height = 2,
                                    borderwidth = 0,
                                    command = self.setting_clicked)
        self.setting_button.pack(side=LEFT, fill=BOTH, expand=TRUE)

        self.exit_button = Button(self.button_frame,
                                    text = "Exit",
                                    font = MAIN_MENU_BUTTON_FONT,
                                    bg = MAIN_MENU_BUTTON_BGD_COLOR,
                                    fg = MAIN_MENU_BUTTON_TXT_COLOR,
                                    # ~ width = 16,
                                    # ~ height = 2,
                                    borderwidth = 0,
                                    command = self.exit_clicked)
        self.exit_button.pack(side=LEFT, fill=BOTH, expand=TRUE)


    def screening_clicked(self):
        self.base_window.forget_page()
        #self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_option)
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.qualitative_analysis_0)
        self.base_window.switch_page()
    
    def quantitative_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.quantitative_option)
        self.base_window.switch_page()

    def view_result_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.view_results)
        self.base_window.switch_page()

    def set_id_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.id_create)
        self.base_window.switch_page()

    def setting_clicked(self):
        self.base_window.forget_page()
        self.base_window.page_num = self.base_window.frame_list.index(self.base_window.setting)
        self.base_window.switch_page()

    def exit_clicked(self):
        msg = messagebox.askquestion("","Do you want to close the app ?")
        if(msg == "yes"):
            self.base_window.destroy()

    def reset(self):
        self.base_window.frame_list.remove(self.base_window.new_qualitative_1)
        self.base_window.frame_list.remove(self.base_window.new_qualitative_2)
        self.base_window.frame_list.remove(self.base_window.new_qualitative_3)
        self.base_window.frame_list.remove(self.base_window.system_check)
        self.base_window.frame_list.remove(self.base_window.qualitative_analysis_1)
        self.base_window.frame_list.remove(self.base_window.qualitative_analysis_2)
        self.base_window.frame_list.remove(self.base_window.qualitative_analysis_3)
        self.base_window.frame_list.remove(self.base_window.id_create)
        self.base_window.frame_list.remove(self.base_window.new_quantitative_1)
        self.base_window.frame_list.remove(self.base_window.new_quantitative_2)
        self.base_window.frame_list.remove(self.base_window.new_quantitative_3)
        self.base_window.frame_list.remove(self.base_window.quantitative_analysis_1)
        self.base_window.frame_list.remove(self.base_window.quantitative_analysis_2)
        self.base_window.frame_list.remove(self.base_window.quantitative_analysis_3)
        self.base_window.frame_list.remove(self.base_window.qualitative_calib_list)
        self.base_window.frame_list.remove(self.base_window.qualitative_analysis_0)
        self.base_window.frame_list.remove(self.base_window.quantitative_calib_list)
        self.base_window.frame_list.remove(self.base_window.quantitative_analysis_0)

        del self.base_window.new_qualitative_1
        del self.base_window.new_qualitative_2
        del self.base_window.new_qualitative_3
        del self.base_window.system_check
        del self.base_window.qualitative_analysis_1
        del self.base_window.qualitative_analysis_2
        del self.base_window.qualitative_analysis_3
        del self.base_window.id_create
        del self.base_window.new_quantitative_1
        del self.base_window.new_quantitative_2
        del self.base_window.new_quantitative_3
        del self.base_window.quantitative_analysis_1
        del self.base_window.quantitative_analysis_2
        del self.base_window.quantitative_analysis_3
        del self.base_window.qualitative_calib_list
        del self.base_window.qualitative_analysis_0
        del self.base_window.quantitative_calib_list
        del self.base_window.quantitative_analysis_0

        self.base_window.new_qualitative_1 = NewQualitativeFrame1(self.base_window)
        self.base_window.new_qualitative_2 = NewQualitativeFrame2(self.base_window)
        self.base_window.new_qualitative_3 = NewQualitativeFrame3(self.base_window)
        self.base_window.system_check = SystemCheckFrame(self.base_window)
        self.base_window.qualitative_analysis_1 = QualitativeAnalysisFrame1(self.base_window)
        self.base_window.qualitative_analysis_2 = QualitativeAnalysisFrame2(self.base_window)
        self.base_window.qualitative_analysis_3 = QualitativeAnalysisFrame3(self.base_window)
        self.base_window.id_create = IDCreateFrame(self.base_window)
        self.base_window.new_quantitative_1 = NewQuantitativeFrame1(self.base_window)
        self.base_window.new_quantitative_2 = NewQuantitativeFrame2(self.base_window)
        self.base_window.new_quantitative_3 = NewQuantitativeFrame3(self.base_window)
        self.base_window.quantitative_analysis_1 = QuantitativeAnalysisFrame1(self.base_window)
        self.base_window.quantitative_analysis_2 = QuantitativeAnalysisFrame2(self.base_window)
        self.base_window.quantitative_analysis_3 = QuantitativeAnalysisFrame3(self.base_window)
        self.base_window.qualitative_calib_list = QualitativeCalibListFrame(self.base_window)
        self.base_window.qualitative_analysis_0 = QualitativeAnalysisFrame0(self.base_window)
        self.base_window.quantitative_calib_list = QuantitativeCalibListFrame(self.base_window)
        self.base_window.quantitative_analysis_0 = QuantitativeAnalysisFrame0(self.base_window)

        self.base_window.frame_list.append(self.base_window.new_qualitative_1)
        self.base_window.frame_list.append(self.base_window.new_qualitative_2)
        self.base_window.frame_list.append(self.base_window.new_qualitative_3)
        self.base_window.frame_list.append(self.base_window.system_check)
        self.base_window.frame_list.append(self.base_window.qualitative_analysis_1)
        self.base_window.frame_list.append(self.base_window.qualitative_analysis_2)
        self.base_window.frame_list.append(self.base_window.qualitative_analysis_3)
        self.base_window.frame_list.append(self.base_window.id_create)
        self.base_window.frame_list.append(self.base_window.new_quantitative_1)
        self.base_window.frame_list.append(self.base_window.new_quantitative_2)
        self.base_window.frame_list.append(self.base_window.new_quantitative_3)
        self.base_window.frame_list.append(self.base_window.quantitative_analysis_1)
        self.base_window.frame_list.append(self.base_window.quantitative_analysis_2)
        self.base_window.frame_list.append(self.base_window.quantitative_analysis_3)
        self.base_window.frame_list.append(self.base_window.qualitative_calib_list)
        self.base_window.frame_list.append(self.base_window.qualitative_analysis_0)
        self.base_window.frame_list.append(self.base_window.quantitative_calib_list)
        self.base_window.frame_list.append(self.base_window.quantitative_analysis_0)

class MainWindow(Tk):
    def __init__(self):
        Tk.__init__(self)
        self.title('Spotcheck')
        self.geometry('800x480')
        self.configure(background = APP_BGD_COLOR)
        self.resizable(FALSE, FALSE)
        self.attributes('-fullscreen', True)

        self.page_num = 0
        self.frame_list = []

        self.main_menu = MainMenu(self)
        self.qualitative_option = QualitativeOptionFrame(self)
        self.quantitative_option = QuantitativeOptionFrame(self)
        self.new_program = NewProgramFrame(self)
        self.view_results = ViewResultFrame(self)
        self.new_qualitative_1 = NewQualitativeFrame1(self)
        self.new_qualitative_2 = NewQualitativeFrame2(self)
        self.new_qualitative_3 = NewQualitativeFrame3(self)
        self.qualitative_analysis_1 = QualitativeAnalysisFrame1(self)
        self.qualitative_analysis_2 = QualitativeAnalysisFrame2(self)
        self.qualitative_analysis_3 = QualitativeAnalysisFrame3(self)
        self.new_quantitative_1 = NewQuantitativeFrame1(self)
        self.new_quantitative_2 = NewQuantitativeFrame2(self)
        self.new_quantitative_3 = NewQuantitativeFrame3(self)
        self.quantitative_analysis_1 = QuantitativeAnalysisFrame1(self)
        self.quantitative_analysis_2 = QuantitativeAnalysisFrame2(self)
        self.quantitative_analysis_3 = QuantitativeAnalysisFrame3(self)
        self.system_check = SystemCheckFrame(self)
        self.id_create = IDCreateFrame(self)
        self.saved_program = SavedProgramFrame(self)
        self.qualitative_saved_program = QualitativeSavedFrame(self)
        self.quantitative_saved_program = QuantitativeSavedFrame(self)
        self.email_setting = EmailSettingFrame(self)
        self.server_setting = ServerSettingFrame(self)
        self.qualitative_calib_list = QualitativeCalibListFrame(self)
        self.qualitative_analysis_0 = QualitativeAnalysisFrame0(self)
        self.quantitative_calib_list = QuantitativeCalibListFrame(self)
        self.quantitative_analysis_0 = QuantitativeAnalysisFrame0(self)
        self.setting = SettingFrame(self)

        self.frame_list.append(self.main_menu)
        self.frame_list.append(self.qualitative_option)
        self.frame_list.append(self.quantitative_option)
        self.frame_list.append(self.new_program)
        self.frame_list.append(self.view_results)
        self.frame_list.append(self.new_qualitative_1)
        self.frame_list.append(self.new_qualitative_2)
        self.frame_list.append(self.new_qualitative_3)
        self.frame_list.append(self.qualitative_analysis_1)
        self.frame_list.append(self.qualitative_analysis_2)
        self.frame_list.append(self.qualitative_analysis_3)
        self.frame_list.append(self.new_quantitative_1)
        self.frame_list.append(self.new_quantitative_2)
        self.frame_list.append(self.new_quantitative_3)
        self.frame_list.append(self.quantitative_analysis_1)
        self.frame_list.append(self.quantitative_analysis_2)
        self.frame_list.append(self.quantitative_analysis_3)
        self.frame_list.append(self.system_check)
        self.frame_list.append(self.id_create)
        self.frame_list.append(self.saved_program)
        self.frame_list.append(self.qualitative_saved_program)
        self.frame_list.append(self.quantitative_saved_program)
        self.frame_list.append(self.email_setting)
        self.frame_list.append(self.server_setting)
        self.frame_list.append(self.qualitative_calib_list)
        self.frame_list.append(self.qualitative_analysis_0)
        self.frame_list.append(self.quantitative_calib_list)
        self.frame_list.append(self.quantitative_analysis_0)
        self.frame_list.append(self.setting)

        self.switch_page()

    def forget_page(self):
        self.frame_list[self.page_num].forget()
    def switch_page(self):
        self.frame_list[self.page_num].tkraise()
        self.frame_list[self.page_num].pack(expand=TRUE, fill=BOTH)

if __name__ == "__main__":
    app = MainWindow()
    app.mainloop()
